import os
import json
import re
import ast
import time
import html
import hashlib
import http.cookiejar
import urllib.parse
import urllib.request
from html.parser import HTMLParser
from functools import wraps
from contextlib import contextmanager

from dotenv import load_dotenv
from flask import Flask, jsonify, g, request
from openpyxl import load_workbook
import pymysql
from pymysql.cursors import DictCursor
from pymysql.err import IntegrityError
from pymysql.err import OperationalError
from werkzeug.exceptions import HTTPException
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(BASE_DIR)
load_dotenv(os.path.join(ROOT_DIR, ".env"))

OPPORTUNITY_TYPES = {"normal", "host"}
OPPORTUNITY_STATUSES = {"new", "assigned", "in_progress", "valid", "invalid"}
OPPORTUNITY_STAGES = {"cold", "interest", "need_defined", "bid_preparing", "ready_for_handoff"}
ORGANIZER_TYPES = {"foreign", "state_owned", "gov_joint", "government", "commercial"}
ACTIVITY_CHANNELS = {"phone", "email", "wechat", "onsite", "other"}
COMPANY_STATUSES = {"active", "inactive"}
USER_ROLES = {"group_admin", "subsidiary_admin", "sales", "marketing"}
USER_STATUSES = {"active", "inactive"}
DEFAULT_USER_PASSWORD = os.getenv("DEFAULT_USER_PASSWORD", "88888888")
DEFAULT_SEARCH_PROVIDER = os.getenv("SEARCH_PROVIDER", "")
DEFAULT_ANALYSIS_MODEL = os.getenv("AZURE_OPENAI_MODEL_ID", "gpt-5-chat")
DEFAULT_IMPORT_FILE = os.getenv("DEFAULT_IMPORT_FILE", "CPS参展商客户名单-分配表1219.xlsx")
UPLOAD_DIR = os.path.join(ROOT_DIR, "uploads")
ALLOWED_IMPORT_EXTENSIONS = {".xlsx"}
CONTACT_ROLE_COLUMN_READY = None
WORKFLOW_TABLES_READY = None
ORG_DIMENSION_TABLES_READY = None
HOST_POOL_TABLES_READY = None
WORKFLOW_TEMPLATE_STATUSES = {"active", "inactive"}
WORKFLOW_PROCESS_DEFAULT_STATUS = "inactive"
ORG_DIMENSION_STATUSES = {"active", "inactive"}
HOST_POOL_STATUSES = {"active", "converted", "archived"}
QUFAIR_DEFAULT_DOMESTIC_URL = os.getenv("QUFAIR_DEFAULT_DOMESTIC_URL", "https://www.qufair.com/flcn/")
WORKFLOW_FIELD_TYPES = {"text", "textarea", "number", "date", "select", "boolean", "attachment", "table"}
WORKFLOW_APPROVER_TYPES = {
  "user",
  "role",
  "manager",
  "department_manager",
  "position",
  "applicant_select",
  "previous_handler"
}
WORKFLOW_APPROVAL_MODES = {"any", "all"}
WORKFLOW_APPROVAL_TYPES = {"any", "all", "sequential"}
WORKFLOW_STEP_TYPES = {"approval", "cc", "condition", "subprocess", "parallel_start", "parallel_join"}
WORKFLOW_NODE_TYPES = {
  "start",
  "approval",
  "cc",
  "condition",
  "end",
  "parallel_start",
  "parallel_join",
  "subprocess"
}
WORKFLOW_CONDITION_LOGICS = {"and", "or"}
WORKFLOW_CONDITION_OPERATORS = {
  "eq",
  "neq",
  "gt",
  "gte",
  "lt",
  "lte",
  "in",
  "not_in",
  "contains",
  "is_true",
  "is_false",
  "is_empty",
  "not_empty"
}
WORKFLOW_INSTANCE_STATUSES = {"pending", "approved", "rejected", "withdrawn"}
WORKFLOW_TASK_STATUSES = {"pending", "waiting", "approved", "rejected", "skipped"}
WORKFLOW_INSTANCE_ACTIONS = {"approve", "reject", "withdraw", "return", "transfer", "add_sign", "remind"}
WORKFLOW_FIELD_KEY_PATTERN = re.compile(r"^[a-zA-Z][a-zA-Z0-9_]{1,63}$")
WORKFLOW_NODE_ID_PATTERN = re.compile(r"^[a-zA-Z0-9_-]{1,64}$")
WORKFLOW_IDEMPOTENCY_KEY_MAX_LEN = 128
DB_CONNECT_RETRIES = 3
DB_RETRY_DELAY_SECONDS = 0.35

HOST_PERSONA_DEFAULTS = {
  "primary_roles": ["采购经理", "市场/品牌经理", "招商主管", "展会项目经理"],
  "decision_makers": ["总经理/副总经理", "市场负责人", "采购负责人"],
  "goals": ["提升参展商数量", "提高观众质量", "扩大品牌影响力", "增加展会收入（展位/赞助）"],
  "pain_points": ["招商周期长", "宣传渠道成本高", "观众质量不稳定", "预算审批不确定", "执行资源紧张"],
  "budget_signals": "年度展会预算/政府补贴/赞助意向",
  "procurement_cycle": "3-6个月（展会筹备期）"
}
NORMAL_PERSONA_DEFAULTS = {
  "primary_roles": ["采购经理", "业务负责人", "运营负责人"],
  "decision_makers": ["总经理/副总经理", "采购负责人", "业务负责人"],
  "goals": ["提升效率", "降低成本", "保障交付与服务质量", "拓展业务"],
  "pain_points": ["预算审批周期长", "需求不明确", "替换供应商阻力大", "交付周期紧"],
  "budget_signals": "年度/季度采购预算或项目立项",
  "procurement_cycle": "1-3个月（视项目复杂度）"
}
HOST_SALES_ANGLES_DEFAULTS = [
  "提供过往展会成效数据与案例",
  "协助招商推广与观众引流",
  "灵活的赞助/展位组合方案",
  "一体化执行与落地服务"
]
NORMAL_SALES_ANGLES_DEFAULTS = [
  "对标行业成功案例",
  "量化ROI与成本节省",
  "试点/分阶段交付方案"
]
HOST_RISK_FLAGS_DEFAULTS = ["预算未确认", "关键决策人不明确", "展会档期与资源冲突"]
NORMAL_RISK_FLAGS_DEFAULTS = ["需求不明确", "预算审批周期长", "关键决策链不完整"]


def hash_password(password: str) -> str:
  # Use pbkdf2 to avoid hashlib.scrypt dependency in older Python builds.
  return generate_password_hash(password, method="pbkdf2:sha256")


def _to_int(value):
  try:
    return int(value or 0)
  except (TypeError, ValueError):
    return 0

ANALYSIS_SCHEMA = (
  "{\n"
  '  "organizer_profile": {\n'
  '    "name": null,\n'
  '    "background": null,\n'
  '    "business_scope": null,\n'
  '    "scale": null,\n'
  '    "locations": [],\n'
  '    "official_sites": []\n'
  "  },\n"
  '  "event_profile": {\n'
  '    "name": null,\n'
  '    "city": null,\n'
  '    "time": null,\n'
  '    "venue": null,\n'
  '    "theme": null,\n'
  '    "scale": null\n'
  "  },\n"
  '  "customer_persona": {\n'
  '    "primary_roles": [],\n'
  '    "decision_makers": [],\n'
  '    "goals": [],\n'
  '    "pain_points": [],\n'
  '    "budget_signals": null,\n'
  '    "procurement_cycle": null\n'
  "  },\n"
  '  "form_suggestions": {\n'
  '    "organizer_name": null,\n'
  '    "organizer_type": null,\n'
  '    "exhibition_name": null,\n'
  '    "exhibition_time": null,\n'
  '    "exhibition_start_date": null,\n'
  '    "exhibition_end_date": null,\n'
  '    "venue_name": null,\n'
  '    "venue_address": null,\n'
  '    "booth_count": null,\n'
  '    "exhibition_area_sqm": null,\n'
  '    "expected_visitors": null,\n'
  '    "exhibition_theme": null,\n'
  '    "budget_range": null,\n'
  '    "city": null,\n'
  '    "industry": null,\n'
  '    "source": null,\n'
  '    "company_name": null,\n'
  '    "company_phone": null,\n'
  '    "company_email": null,\n'
  '    "contact_department": null,\n'
  '    "contact_person": null,\n'
  '    "contact_address": null,\n'
  '    "website": null,\n'
  '    "country": null,\n'
  '    "hall_no": null,\n'
  '    "booth_no": null,\n'
  '    "booth_type": null,\n'
  '    "booth_area_sqm": null\n'
  "  },\n"
  '  "sales_angles": [],\n'
  '  "risk_flags": [],\n'
  '  "assumptions": [],\n'
  '  "contacts": [\n'
  '    {\n'
  '      "name": "",\n'
  '      "title": null,\n'
  '      "organization": null,\n'
  '      "source_url": "",\n'
  '      "confidence": "medium"\n'
  "    }\n"
  "  ]\n"
  "}"
)


class _TextExtractor(HTMLParser):
  def __init__(self):
    super().__init__()
    self.parts = []

  def handle_data(self, data):
    text = data.strip()
    if text:
      self.parts.append(text)


def _strip_html(html):
  if not html:
    return ""
  cleaned = re.sub(r"<(script|style)[^>]*>.*?</\\1>", " ", html, flags=re.I | re.S)
  parser = _TextExtractor()
  parser.feed(cleaned)
  text = " ".join(parser.parts)
  text = re.sub(r"\\s+", " ", text).strip()
  return text


def _fetch_url_text(url, timeout=12, max_chars=12000):
  try:
    req = urllib.request.Request(
      url,
      headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0 Safari/537.36"
      }
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
      content_type = resp.headers.get("Content-Type", "")
      if "text/html" not in content_type:
        return ""
      raw = resp.read(600000)
      html = raw.decode("utf-8", errors="ignore")
      text = _strip_html(html)
      return text[:max_chars]
  except Exception:
    return ""


def _search_serper(query, num=5):
  api_key = os.getenv("SERPER_API_KEY")
  if not api_key:
    raise RuntimeError("missing_serper_api_key")
  payload = json.dumps({"q": query, "num": num}).encode("utf-8")
  req = urllib.request.Request(
    "https://google.serper.dev/search",
    data=payload,
    headers={"Content-Type": "application/json", "X-API-KEY": api_key}
  )
  with urllib.request.urlopen(req, timeout=12) as resp:
    body = json.loads(resp.read().decode("utf-8"))
  results = []
  for item in body.get("organic", [])[:num]:
    results.append(
      {
        "title": item.get("title"),
        "url": item.get("link"),
        "snippet": item.get("snippet")
      }
    )
  return results


def _search_serpapi(query, num=5):
  api_key = os.getenv("SERPAPI_API_KEY")
  if not api_key:
    raise RuntimeError("missing_serpapi_api_key")
  params = urllib.parse.urlencode({"q": query, "api_key": api_key, "num": num})
  url = f"https://serpapi.com/search.json?{params}"
  with urllib.request.urlopen(url, timeout=12) as resp:
    body = json.loads(resp.read().decode("utf-8"))
  results = []
  for item in body.get("organic_results", [])[:num]:
    results.append(
      {
        "title": item.get("title"),
        "url": item.get("link"),
        "snippet": item.get("snippet")
      }
    )
  return results


def _search_searxng(query, num=5):
  base_url = os.getenv("SEARXNG_URL", "").rstrip("/")
  if not base_url:
    raise RuntimeError("missing_searxng_url")
  lang = os.getenv("SEARXNG_LANG", "zh-CN")
  params = urllib.parse.urlencode({"q": query, "format": "json", "language": lang})
  url = f"{base_url}/search?{params}"
  headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0 Safari/537.36"
  }
  api_key = os.getenv("SEARXNG_API_KEY")
  if api_key:
    headers["X-API-Key"] = api_key
  req = urllib.request.Request(url, headers=headers)
  with urllib.request.urlopen(req, timeout=12) as resp:
    body = json.loads(resp.read().decode("utf-8"))
  results = []
  for item in body.get("results", [])[:num]:
    results.append(
      {
        "title": item.get("title"),
        "url": item.get("url"),
        "snippet": item.get("content") or item.get("snippet")
      }
    )
  return results


def _search_web(query, num=5):
  provider = DEFAULT_SEARCH_PROVIDER
  if not provider:
    if os.getenv("SERPER_API_KEY"):
      provider = "serper"
    elif os.getenv("SERPAPI_API_KEY"):
      provider = "serpapi"
    elif os.getenv("SEARXNG_URL"):
      provider = "searxng"
  if provider == "serper":
    return _search_serper(query, num=num)
  if provider == "serpapi":
    return _search_serpapi(query, num=num)
  if provider == "searxng":
    return _search_searxng(query, num=num)
  raise RuntimeError("search_provider_not_configured")


def _call_chat_completion(messages, temperature=0.2):
  use_azure = os.getenv("AZURE_OPENAI_DEFAULT_FLAG", "").upper() == "Y"
  if use_azure:
    endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").rstrip("/")
    api_key = os.getenv("AZURE_OPENAI_API_KEY")
    api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
    if not endpoint or not api_key:
      raise RuntimeError("azure_openai_not_configured")
    if not endpoint.endswith("/chat/completions"):
      url = f"{endpoint}/chat/completions?api-version={api_version}"
    else:
      connector = "&" if "api-version=" in endpoint else "?"
      url = f"{endpoint}{connector}api-version={api_version}"
    payload = {
      "model": os.getenv("AZURE_OPENAI_MODEL_ID", DEFAULT_ANALYSIS_MODEL),
      "messages": messages,
      "temperature": temperature
    }
    headers = {"Content-Type": "application/json", "api-key": api_key}
  else:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
      raise RuntimeError("openai_not_configured")
    url = "https://api.openai.com/v1/chat/completions"
    payload = {"model": DEFAULT_ANALYSIS_MODEL, "messages": messages, "temperature": temperature}
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}

  req = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"), headers=headers)
  with urllib.request.urlopen(req, timeout=30) as resp:
    body = json.loads(resp.read().decode("utf-8"))

  choices = body.get("choices") or []
  if not choices:
    raise RuntimeError("model_no_response")
  return (choices[0].get("message") or {}).get("content", "")


def _safe_json_load(raw):
  if raw is None:
    return {}
  if isinstance(raw, (dict, list)):
    return raw
  if isinstance(raw, (bytes, bytearray)):
    raw = raw.decode("utf-8", errors="ignore")
  if not isinstance(raw, str):
    return {}
  raw = raw.strip()
  if not raw:
    return {}
  try:
    return json.loads(raw)
  except Exception:
    candidates = []
    start_obj = raw.find("{")
    end_obj = raw.rfind("}")
    if start_obj != -1 and end_obj != -1 and end_obj > start_obj:
      candidates.append(raw[start_obj : end_obj + 1])
    start_arr = raw.find("[")
    end_arr = raw.rfind("]")
    if start_arr != -1 and end_arr != -1 and end_arr > start_arr:
      candidates.append(raw[start_arr : end_arr + 1])
    for candidate in candidates:
      try:
        return json.loads(candidate)
      except Exception:
        continue
    return {}


def _normalize_contacts(raw):
  if not isinstance(raw, list):
    return []
  normalized = []
  for item in raw:
    if not isinstance(item, dict):
      continue
    contact = {
      "name": item.get("name") or item.get("contact_name"),
      "role": item.get("role") or item.get("contact_role"),
      "title": item.get("title") or item.get("contact_title"),
      "phone": item.get("phone") or item.get("contact_phone"),
      "email": item.get("email") or item.get("contact_email"),
      "wechat": item.get("wechat") or item.get("contact_wechat")
    }
    if any(contact.values()):
      normalized.append(contact)
  return normalized


def _ensure_contact_role_column(db):
  global CONTACT_ROLE_COLUMN_READY
  if CONTACT_ROLE_COLUMN_READY is not None:
    return
  try:
    with db.cursor() as cur:
      cur.execute("SHOW COLUMNS FROM opportunity_contacts LIKE 'role'")
      if cur.fetchone():
        CONTACT_ROLE_COLUMN_READY = True
        return
      cur.execute("ALTER TABLE opportunity_contacts ADD COLUMN role VARCHAR(50) NULL AFTER title")
    CONTACT_ROLE_COLUMN_READY = True
  except Exception:
    CONTACT_ROLE_COLUMN_READY = False


def _ensure_workflow_tables(db):
  global WORKFLOW_TABLES_READY
  if WORKFLOW_TABLES_READY is not None:
    return
  try:
    with db.cursor() as cur:
      cur.execute(
        "CREATE TABLE IF NOT EXISTS approval_form_templates ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "name VARCHAR(255) NOT NULL, "
        "description VARCHAR(500) NULL, "
        "company_id BIGINT UNSIGNED NULL, "
        "schema_json LONGTEXT NOT NULL, "
        "status ENUM('active', 'inactive') NOT NULL DEFAULT 'active', "
        "created_by BIGINT UNSIGNED NOT NULL, "
        "updated_by BIGINT UNSIGNED NOT NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "INDEX idx_form_tpl_company_status (company_id, status), "
        "INDEX idx_form_tpl_created_by (created_by), "
        "CONSTRAINT fk_form_tpl_company FOREIGN KEY (company_id) REFERENCES companies(id), "
        "CONSTRAINT fk_form_tpl_created_by FOREIGN KEY (created_by) REFERENCES users(id), "
        "CONSTRAINT fk_form_tpl_updated_by FOREIGN KEY (updated_by) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS approval_process_templates ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "name VARCHAR(255) NOT NULL, "
        "description VARCHAR(500) NULL, "
        "company_id BIGINT UNSIGNED NULL, "
        "form_template_id BIGINT UNSIGNED NOT NULL, "
        "steps_json LONGTEXT NOT NULL, "
        "current_version INT NOT NULL DEFAULT 1, "
        "published_version INT NULL, "
        "status ENUM('active', 'inactive') NOT NULL DEFAULT 'inactive', "
        "created_by BIGINT UNSIGNED NOT NULL, "
        "updated_by BIGINT UNSIGNED NOT NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "INDEX idx_proc_tpl_company_status (company_id, status), "
        "INDEX idx_proc_tpl_form_id (form_template_id), "
        "CONSTRAINT fk_proc_tpl_company FOREIGN KEY (company_id) REFERENCES companies(id), "
        "CONSTRAINT fk_proc_tpl_form FOREIGN KEY (form_template_id) REFERENCES approval_form_templates(id), "
        "CONSTRAINT fk_proc_tpl_created_by FOREIGN KEY (created_by) REFERENCES users(id), "
        "CONSTRAINT fk_proc_tpl_updated_by FOREIGN KEY (updated_by) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS approval_process_template_versions ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "process_template_id BIGINT UNSIGNED NOT NULL, "
        "version_no INT NOT NULL, "
        "form_template_id BIGINT UNSIGNED NOT NULL, "
        "definition_json LONGTEXT NOT NULL, "
        "status ENUM('draft', 'published', 'archived') NOT NULL DEFAULT 'draft', "
        "published_at TIMESTAMP NULL, "
        "created_by BIGINT UNSIGNED NOT NULL, "
        "updated_by BIGINT UNSIGNED NOT NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "UNIQUE KEY uniq_proc_tpl_version (process_template_id, version_no), "
        "INDEX idx_proc_tpl_version_status (process_template_id, status), "
        "CONSTRAINT fk_proc_tpl_ver_template FOREIGN KEY (process_template_id) REFERENCES approval_process_templates(id) ON DELETE CASCADE, "
        "CONSTRAINT fk_proc_tpl_ver_form FOREIGN KEY (form_template_id) REFERENCES approval_form_templates(id), "
        "CONSTRAINT fk_proc_tpl_ver_created_by FOREIGN KEY (created_by) REFERENCES users(id), "
        "CONSTRAINT fk_proc_tpl_ver_updated_by FOREIGN KEY (updated_by) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS approval_instances ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "process_template_id BIGINT UNSIGNED NOT NULL, "
        "form_template_id BIGINT UNSIGNED NOT NULL, "
        "process_name VARCHAR(255) NOT NULL, "
        "title VARCHAR(255) NOT NULL, "
        "company_id BIGINT UNSIGNED NULL, "
        "applicant_id BIGINT UNSIGNED NOT NULL, "
        "process_snapshot_json LONGTEXT NOT NULL, "
        "form_schema_json LONGTEXT NOT NULL, "
        "form_data_json LONGTEXT NOT NULL, "
        "status ENUM('pending', 'approved', 'rejected', 'withdrawn') NOT NULL DEFAULT 'pending', "
        "current_step INT NOT NULL DEFAULT 1, "
        "total_steps INT NOT NULL DEFAULT 1, "
        "current_step_name VARCHAR(255) NULL, "
        "current_node_id VARCHAR(64) NULL, "
        "finished_at TIMESTAMP NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "INDEX idx_instance_status (status), "
        "INDEX idx_instance_applicant (applicant_id), "
        "INDEX idx_instance_company (company_id), "
        "INDEX idx_instance_process (process_template_id), "
        "CONSTRAINT fk_instance_process FOREIGN KEY (process_template_id) REFERENCES approval_process_templates(id), "
        "CONSTRAINT fk_instance_form FOREIGN KEY (form_template_id) REFERENCES approval_form_templates(id), "
        "CONSTRAINT fk_instance_company FOREIGN KEY (company_id) REFERENCES companies(id), "
        "CONSTRAINT fk_instance_applicant FOREIGN KEY (applicant_id) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS approval_instance_tasks ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "instance_id BIGINT UNSIGNED NOT NULL, "
        "step_no INT NOT NULL, "
        "step_name VARCHAR(255) NOT NULL, "
        "approval_mode ENUM('any', 'all') NOT NULL DEFAULT 'any', "
        "approver_id BIGINT UNSIGNED NOT NULL, "
        "status ENUM('pending', 'waiting', 'approved', 'rejected', 'skipped') NOT NULL DEFAULT 'pending', "
        "decision ENUM('approve', 'reject') NULL, "
        "comment TEXT NULL, "
        "acted_at TIMESTAMP NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "UNIQUE KEY uniq_instance_step_approver (instance_id, step_no, approver_id), "
        "INDEX idx_task_instance_step (instance_id, step_no), "
        "INDEX idx_task_approver_status (approver_id, status), "
        "CONSTRAINT fk_task_instance FOREIGN KEY (instance_id) REFERENCES approval_instances(id) ON DELETE CASCADE, "
        "CONSTRAINT fk_task_approver FOREIGN KEY (approver_id) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS approval_instance_events ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "instance_id BIGINT UNSIGNED NOT NULL, "
        "task_id BIGINT UNSIGNED NULL, "
        "user_id BIGINT UNSIGNED NOT NULL, "
        "action VARCHAR(32) NOT NULL, "
        "detail_json LONGTEXT NULL, "
        "comment TEXT NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "INDEX idx_instance_event_instance (instance_id, created_at), "
        "INDEX idx_instance_event_action (action), "
        "CONSTRAINT fk_instance_event_instance FOREIGN KEY (instance_id) REFERENCES approval_instances(id) ON DELETE CASCADE, "
        "CONSTRAINT fk_instance_event_task FOREIGN KEY (task_id) REFERENCES approval_instance_tasks(id) ON DELETE SET NULL, "
        "CONSTRAINT fk_instance_event_user FOREIGN KEY (user_id) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS approval_action_idempotency ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "idem_key VARCHAR(128) NOT NULL, "
        "instance_id BIGINT UNSIGNED NOT NULL, "
        "actor_id BIGINT UNSIGNED NOT NULL, "
        "action VARCHAR(32) NOT NULL, "
        "response_json LONGTEXT NULL, "
        "status_code INT NOT NULL DEFAULT 200, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "UNIQUE KEY uniq_action_idem (idem_key, instance_id, actor_id, action), "
        "INDEX idx_action_idem_created (created_at), "
        "CONSTRAINT fk_action_idem_instance FOREIGN KEY (instance_id) REFERENCES approval_instances(id) ON DELETE CASCADE, "
        "CONSTRAINT fk_action_idem_actor FOREIGN KEY (actor_id) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute("SHOW COLUMNS FROM approval_process_templates LIKE 'current_version'")
      if not cur.fetchone():
        cur.execute("ALTER TABLE approval_process_templates ADD COLUMN current_version INT NOT NULL DEFAULT 1 AFTER steps_json")
      cur.execute("SHOW COLUMNS FROM approval_process_templates LIKE 'published_version'")
      if not cur.fetchone():
        cur.execute("ALTER TABLE approval_process_templates ADD COLUMN published_version INT NULL AFTER current_version")
      cur.execute(
        "UPDATE approval_process_templates "
        "SET current_version = 1 "
        "WHERE current_version IS NULL OR current_version <= 0"
      )
      cur.execute(
        "UPDATE approval_process_templates "
        "SET published_version = CASE WHEN status = 'active' THEN current_version ELSE published_version END "
        "WHERE published_version IS NULL AND status = 'active'"
      )
      cur.execute(
        "INSERT INTO approval_process_template_versions "
        "(process_template_id, version_no, form_template_id, definition_json, status, published_at, created_by, updated_by) "
        "SELECT apt.id, 1, apt.form_template_id, apt.steps_json, "
        "CASE WHEN apt.status = 'active' AND COALESCE(apt.published_version, 1) = 1 THEN 'published' ELSE 'draft' END, "
        "CASE WHEN apt.status = 'active' AND COALESCE(apt.published_version, 1) = 1 THEN CURRENT_TIMESTAMP ELSE NULL END, "
        "apt.created_by, apt.updated_by "
        "FROM approval_process_templates apt "
        "WHERE NOT EXISTS ("
        "SELECT 1 FROM approval_process_template_versions apv "
        "WHERE apv.process_template_id = apt.id AND apv.version_no = 1"
        ")"
      )
      cur.execute(
        "UPDATE approval_process_template_versions apv "
        "JOIN approval_process_templates apt ON apt.id = apv.process_template_id "
        "SET apv.status = CASE "
        "WHEN apt.published_version = apv.version_no AND apt.status = 'active' THEN 'published' "
        "ELSE IF(apv.status = 'published', 'archived', apv.status) "
        "END, "
        "apv.published_at = CASE "
        "WHEN apt.published_version = apv.version_no AND apt.status = 'active' AND apv.published_at IS NULL THEN CURRENT_TIMESTAMP "
        "ELSE apv.published_at "
        "END"
      )
      cur.execute("SHOW COLUMNS FROM approval_instances LIKE 'current_node_id'")
      if not cur.fetchone():
        cur.execute("ALTER TABLE approval_instances ADD COLUMN current_node_id VARCHAR(64) NULL AFTER current_step_name")
      cur.execute("SHOW COLUMNS FROM approval_instance_tasks LIKE 'status'")
      task_status_col = cur.fetchone()
      task_status_type = str((task_status_col or {}).get("Type") or (task_status_col or {}).get("type") or "").lower()
      if "waiting" not in task_status_type:
        cur.execute(
          "ALTER TABLE approval_instance_tasks "
          "MODIFY COLUMN status ENUM('pending', 'waiting', 'approved', 'rejected', 'skipped') "
          "NOT NULL DEFAULT 'pending'"
        )
    _repair_process_form_template_conflicts(db)
    _ensure_unique_process_form_binding_index(db)
    WORKFLOW_TABLES_READY = True
  except Exception:
    WORKFLOW_TABLES_READY = False


def _ensure_org_dimension_tables(db):
  global ORG_DIMENSION_TABLES_READY
  if ORG_DIMENSION_TABLES_READY is not None:
    return
  try:
    with db.cursor() as cur:
      cur.execute(
        "CREATE TABLE IF NOT EXISTS org_roles ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "name VARCHAR(100) NOT NULL, "
        "code VARCHAR(64) NULL, "
        "company_id BIGINT UNSIGNED NULL, "
        "status ENUM('active', 'inactive') NOT NULL DEFAULT 'active', "
        "created_by BIGINT UNSIGNED NOT NULL, "
        "updated_by BIGINT UNSIGNED NOT NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "UNIQUE KEY uniq_org_role_code (code), "
        "INDEX idx_org_role_scope (company_id, status), "
        "INDEX idx_org_role_name (name), "
        "CONSTRAINT fk_org_role_company FOREIGN KEY (company_id) REFERENCES companies(id), "
        "CONSTRAINT fk_org_role_created_by FOREIGN KEY (created_by) REFERENCES users(id), "
        "CONSTRAINT fk_org_role_updated_by FOREIGN KEY (updated_by) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS org_positions ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "name VARCHAR(100) NOT NULL, "
        "code VARCHAR(64) NULL, "
        "company_id BIGINT UNSIGNED NULL, "
        "status ENUM('active', 'inactive') NOT NULL DEFAULT 'active', "
        "created_by BIGINT UNSIGNED NOT NULL, "
        "updated_by BIGINT UNSIGNED NOT NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "UNIQUE KEY uniq_org_position_code (code), "
        "INDEX idx_org_position_scope (company_id, status), "
        "INDEX idx_org_position_name (name), "
        "CONSTRAINT fk_org_position_company FOREIGN KEY (company_id) REFERENCES companies(id), "
        "CONSTRAINT fk_org_position_created_by FOREIGN KEY (created_by) REFERENCES users(id), "
        "CONSTRAINT fk_org_position_updated_by FOREIGN KEY (updated_by) REFERENCES users(id)"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS user_org_roles ("
        "user_id BIGINT UNSIGNED NOT NULL, "
        "role_id BIGINT UNSIGNED NOT NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "PRIMARY KEY (user_id, role_id), "
        "INDEX idx_user_org_role_role (role_id, user_id), "
        "CONSTRAINT fk_user_org_role_user FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE, "
        "CONSTRAINT fk_user_org_role_role FOREIGN KEY (role_id) REFERENCES org_roles(id) ON DELETE CASCADE"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
      cur.execute(
        "CREATE TABLE IF NOT EXISTS user_org_positions ("
        "user_id BIGINT UNSIGNED NOT NULL, "
        "position_id BIGINT UNSIGNED NOT NULL, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "PRIMARY KEY (user_id, position_id), "
        "INDEX idx_user_org_position_position (position_id, user_id), "
        "CONSTRAINT fk_user_org_position_user FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE, "
        "CONSTRAINT fk_user_org_position_position FOREIGN KEY (position_id) REFERENCES org_positions(id) ON DELETE CASCADE"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
    ORG_DIMENSION_TABLES_READY = True
  except Exception:
    ORG_DIMENSION_TABLES_READY = False


def _ensure_host_pool_tables(db):
  global HOST_POOL_TABLES_READY
  if HOST_POOL_TABLES_READY is not None:
    return
  try:
    with db.cursor() as cur:
      cur.execute(
        "CREATE TABLE IF NOT EXISTS host_opportunity_pool_events ("
        "id BIGINT UNSIGNED AUTO_INCREMENT PRIMARY KEY, "
        "source_site VARCHAR(32) NOT NULL DEFAULT 'qufair', "
        "external_id VARCHAR(128) NOT NULL, "
        "name VARCHAR(255) NOT NULL, "
        "alias_name VARCHAR(255) NULL, "
        "industry VARCHAR(255) NULL, "
        "country VARCHAR(100) NULL, "
        "city VARCHAR(100) NULL, "
        "organizer_name VARCHAR(255) NULL, "
        "venue_name VARCHAR(255) NULL, "
        "venue_address VARCHAR(255) NULL, "
        "exhibition_start_date DATE NULL, "
        "exhibition_end_date DATE NULL, "
        "cycle_text VARCHAR(100) NULL, "
        "exhibition_area_sqm INT NULL, "
        "exhibitors_count INT NULL, "
        "visitors_count INT NULL, "
        "heat_score INT NULL, "
        "source_url VARCHAR(500) NOT NULL, "
        "source_cover_url VARCHAR(500) NULL, "
        "source_list_url VARCHAR(500) NULL, "
        "is_domestic TINYINT(1) NOT NULL DEFAULT 1, "
        "pool_status ENUM('active', 'converted', 'archived') NOT NULL DEFAULT 'active', "
        "converted_opportunity_id BIGINT UNSIGNED NULL, "
        "raw_json LONGTEXT NULL, "
        "fetched_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP, "
        "updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, "
        "UNIQUE KEY uniq_host_pool_source_external (source_site, external_id), "
        "UNIQUE KEY uniq_host_pool_source_url (source_url), "
        "INDEX idx_host_pool_status_date (pool_status, exhibition_start_date), "
        "INDEX idx_host_pool_city (city), "
        "INDEX idx_host_pool_industry (industry), "
        "CONSTRAINT fk_host_pool_converted_opp FOREIGN KEY (converted_opportunity_id) "
        "REFERENCES opportunities(id) ON DELETE SET NULL"
        ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"
      )
  except Exception:
    HOST_POOL_TABLES_READY = False
    return
  HOST_POOL_TABLES_READY = True


def _normalize_header(value):
  if value is None:
    return ""
  return re.sub(r"\s+", "", str(value))


def _resolve_import_path(filename):
  safe_name = os.path.basename(filename or "")
  if not safe_name:
    return None
  root_path = os.path.join(ROOT_DIR, safe_name)
  upload_path = os.path.join(UPLOAD_DIR, safe_name)
  if os.path.exists(upload_path):
    return upload_path
  if os.path.exists(root_path):
    return root_path
  return None


def _find_header_row(sheet):
  for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
    if not row:
      continue
    normalized = [_normalize_header(cell) for cell in row]
    if "公司名称_CN" in normalized or "CompanyName_EN" in normalized or "公司名称" in normalized:
      return row_idx, row
  return None, None


def _build_header_index_map(header_row):
  index_map = {}
  email_count = 0
  phone_count = 0
  for idx, value in enumerate(header_row):
    name = _normalize_header(value)
    if not name:
      continue
    if name in {"CompanyName_EN", "CompanyName_EN"} or name == "CompanyName_EN":
      index_map.setdefault("company_en", idx)
    elif name in {"公司名称_CN", "公司名称"}:
      index_map.setdefault("company_cn", idx)
    elif name == "联系人":
      index_map.setdefault("contact_name", idx)
    elif name == "电话号码":
      index_map.setdefault("contact_phone", idx)
    elif name == "手机号码":
      index_map.setdefault("contact_mobile", idx)
    elif name == "邮箱":
      email_count += 1
      if email_count == 1:
        index_map.setdefault("contact_email", idx)
      else:
        index_map.setdefault("contact_email_alt", idx)
    elif name in {"公司邮箱", "企业邮箱"}:
      index_map.setdefault("company_email", idx)
    elif name == "地区":
      index_map.setdefault("region", idx)
    elif name in {"官网", "网址", "网站"}:
      index_map.setdefault("website", idx)
    elif name in {"地址"}:
      index_map.setdefault("address", idx)
    elif name in {"联系地址"}:
      index_map.setdefault("contact_address", idx)
    elif name in {"联系部门"}:
      index_map.setdefault("contact_department", idx)
    elif name in {"公司电话", "联系电话"}:
      index_map.setdefault("company_phone", idx)
    elif name in {"国家", "国家/地区"}:
      index_map.setdefault("country", idx)
    elif name in {"展馆号", "展馆编号"}:
      index_map.setdefault("hall_no", idx)
    elif name in {"展位号", "展位编号"}:
      index_map.setdefault("booth_no", idx)
    elif name in {"展位类型"}:
      index_map.setdefault("booth_type", idx)
    elif name in {"展位面积"}:
      index_map.setdefault("booth_area_sqm", idx)
    elif name == "跟进人/PIC":
      index_map.setdefault("pic", idx)
    elif "电话/邮箱/陌拜" in name:
      index_map.setdefault("touch_method", idx)
    elif "是否取得联系" in name:
      index_map.setdefault("contacted", idx)
    elif name == "客户联系人":
      index_map.setdefault("contact_alt_name", idx)
    elif name == "职位":
      index_map.setdefault("contact_alt_title", idx)
    elif name == "电话":
      phone_count += 1
      if phone_count == 1:
        index_map.setdefault("contact_alt_phone", idx)
      else:
        index_map.setdefault("contact_phone_extra", idx)
    elif name == "现场反馈":
      index_map.setdefault("feedback", idx)
    elif name == "竞争对手公司":
      index_map.setdefault("competitor", idx)
    elif name == "备注":
      index_map.setdefault("remark", idx)
    elif name == "业务类型":
      index_map.setdefault("business_type", idx)
    elif name == "名称":
      index_map.setdefault("business_name", idx)
  return index_map


def _stringify_cell(value):
  if value is None:
    return None
  return str(value).strip()


def _parse_int(value):
  if value is None:
    return None
  if isinstance(value, bool):
    return None
  if isinstance(value, int):
    return value
  if isinstance(value, float):
    return int(value)
  text = str(value).strip()
  if not text:
    return None
  try:
    return int(float(text))
  except Exception:
    return None


def _build_notes(parts, max_len=255):
  notes = [item for item in parts if item]
  if not notes:
    return None
  text = "；".join(notes)
  if len(text) <= max_len:
    return text
  return text[: max_len - 1] + "…"


def _clip_text(value, max_len):
  if value is None:
    return None
  text = str(value).strip()
  if not text:
    return None
  if len(text) <= max_len:
    return text
  return text[:max_len]


def _strip_html_tags(value):
  if value is None:
    return ""
  text = re.sub(r"<[^>]+>", "", str(value), flags=re.S)
  text = html.unescape(text)
  return re.sub(r"\s+", " ", text).strip()


def _safe_parse_int(value):
  parsed = _parse_int(value)
  if parsed is None:
    return None
  return max(parsed, 0)


def _parse_qufair_date(value):
  if value is None:
    return None
  text = str(value).strip()
  if not text:
    return None
  text = text.replace(".", "-").replace("年", "-").replace("月", "-").replace("日", "")
  match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
  if not match:
    return None
  year = _safe_parse_int(match.group(1))
  month = _safe_parse_int(match.group(2))
  day = _safe_parse_int(match.group(3))
  if not year or not month or not day:
    return None
  try:
    return datetime(year, month, day).strftime("%Y-%m-%d")
  except Exception:
    return None


def _parse_qufair_date_range(value):
  if value is None:
    return (None, None)
  text = str(value).strip()
  if not text:
    return (None, None)
  parts = [item for item in re.split(r"[~/]", text) if item]
  if len(parts) >= 2:
    return (_parse_qufair_date(parts[0]), _parse_qufair_date(parts[1]))
  single = _parse_qufair_date(parts[0]) if parts else None
  return (single, single)


def _extract_qufair_external_id(url):
  parsed = urllib.parse.urlparse(str(url or ""))
  path = parsed.path or ""
  matched = re.search(r"/(\d+)(?:/|$)", path)
  if matched:
    return matched.group(1)
  matched = re.search(r"/convention/(\d+)\.shtml", path)
  if matched:
    return matched.group(1)
  return hashlib.md5(str(url or "").encode("utf-8")).hexdigest()[:24]


def _to_absolute_url(url, base_url="https://www.qufair.com/"):
  if not url:
    return None
  return urllib.parse.urljoin(base_url, str(url).strip())


def _to_bool(value, default=False):
  if value is None:
    return default
  if isinstance(value, bool):
    return value
  text = str(value).strip().lower()
  if text in {"1", "true", "yes", "y", "on"}:
    return True
  if text in {"0", "false", "no", "n", "off"}:
    return False
  return default


def _build_qufair_opener():
  cookie_jar = http.cookiejar.CookieJar()
  opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cookie_jar))
  opener.addheaders = [
    (
      "User-Agent",
      "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    ("Accept-Language", "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7")
  ]
  return opener


def _fetch_with_opener(opener, url, referer=None, timeout=20):
  headers = {}
  if referer:
    headers["Referer"] = referer
  req = urllib.request.Request(url, headers=headers)
  with opener.open(req, timeout=timeout) as response:
    raw = response.read()
    charset = None
    try:
      charset = response.headers.get_content_charset()
    except Exception:
      charset = None
  return raw.decode(charset or "utf-8", errors="ignore")


def _fetch_qufair_page(opener, url):
  html_text = _fetch_with_opener(opener, url)
  if "验证不是机器人" in html_text and "go_url" in html_text:
    parsed = urllib.parse.urlparse(url)
    robot_url = (
      "https://www.qufair.com/robot/index/?url="
      + urllib.parse.quote(parsed.path or "/", safe="")
    )
    _fetch_with_opener(opener, robot_url, referer="https://www.qufair.com/")
    html_text = _fetch_with_opener(opener, url, referer=robot_url)
  return html_text


def _parse_qufair_domestic_categories(html_text):
  categories = []
  seen = set()
  pattern = re.compile(
    r'<a class="item"[^>]*href="(?P<href>[^"]+)"[^>]*>(?P<label>[^<]+)\((?P<count>\d+)\)</a>',
    re.S
  )
  for matched in pattern.finditer(html_text or ""):
    href = _to_absolute_url(matched.group("href"))
    if not href or "/fl/" not in href:
      continue
    if href in seen:
      continue
    seen.add(href)
    label = _clip_text(_strip_html_tags(matched.group("label")), 100)
    count = _safe_parse_int(matched.group("count")) or 0
    categories.append(
      {
        "url": href,
        "industry": label,
        "estimated_count": count
      }
    )
  categories.sort(key=lambda item: item.get("estimated_count") or 0, reverse=True)
  return categories


def _parse_qufair_event_cards(html_text, source_list_url=None, fallback_industry=None):
  rows = []
  blocks = re.findall(r"<li class=\"info\">(.*?)</li>", html_text or "", flags=re.S)
  for block in blocks:
    href_match = re.search(r'href="(https?://www\.qufair\.com/[^"]+)"', block)
    if not href_match:
      href_match = re.search(r'href="(/[^"]+)"', block)
    if not href_match:
      continue
    source_url = _to_absolute_url(href_match.group(1))
    if not source_url:
      continue
    title = _clip_text(
      _strip_html_tags(
        re.search(r'class="name"[^>]*title="([^"]+)"', block).group(1)
        if re.search(r'class="name"[^>]*title="([^"]+)"', block)
        else ""
      ),
      255
    )
    if not title:
      text_match = re.search(r'class="name"[^>]*>(.*?)</a>', block, flags=re.S)
      title = _clip_text(_strip_html_tags(text_match.group(1) if text_match else ""), 255)
    datetime_match = re.search(r"<time[^>]*datetime=\"([^\"]+)\"", block)
    start_date, end_date = _parse_qufair_date_range(datetime_match.group(1) if datetime_match else None)
    heat_match = re.search(r"<em>(\d+)</em>\s*展会热度", block)
    cover_match = re.search(r"<img[^>]*src=['\"]([^'\"]+)['\"]", block)
    cover_url = _to_absolute_url(cover_match.group(1)) if cover_match else None
    if cover_url and "!" in cover_url:
      cover_url = cover_url.split("!", 1)[0]
    rows.append(
      {
        "source_site": "qufair",
        "external_id": _extract_qufair_external_id(source_url),
        "name": title,
        "exhibition_name": title,
        "industry": _clip_text(fallback_industry, 255),
        "exhibition_start_date": start_date,
        "exhibition_end_date": end_date,
        "heat_score": _safe_parse_int(heat_match.group(1) if heat_match else None),
        "source_url": source_url,
        "source_cover_url": _clip_text(cover_url, 500),
        "source_list_url": _clip_text(source_list_url, 500),
        "is_domestic": 1
      }
    )
  dedup = {}
  for row in rows:
    dedup_key = row.get("source_url")
    if not dedup_key:
      dedup_key = f"{row.get('source_site')}:{row.get('external_id')}"
    dedup[dedup_key] = row
  return list(dedup.values())


def _parse_qufair_detail_info(html_text):
  detail = {}
  cn_name_match = re.search(r'<span class="cn_name">(.*?)</span>', html_text or "", flags=re.S)
  en_name_match = re.search(r'<span class="en_name">(.*?)</span>', html_text or "", flags=re.S)
  if cn_name_match:
    detail["name"] = _clip_text(_strip_html_tags(cn_name_match.group(1)), 255)
  if en_name_match:
    detail["alias_name"] = _clip_text(_strip_html_tags(en_name_match.group(1)), 255)

  datetime_match = re.search(r"<li class=\"datebox\".*?<time[^>]*datetime=\"([^\"]+)\"", html_text or "", flags=re.S)
  start_date, end_date = _parse_qufair_date_range(datetime_match.group(1) if datetime_match else None)
  if start_date:
    detail["exhibition_start_date"] = start_date
  if end_date:
    detail["exhibition_end_date"] = end_date

  industry_block = re.search(
    r'<li class="con_hy"><span>展览行业：</span>(.*?)</li>',
    html_text or "",
    flags=re.S
  )
  if industry_block:
    industry_names = re.findall(r">([^<]+)</a>", industry_block.group(1), flags=re.S)
    normalized_industries = [_strip_html_tags(item) for item in industry_names if _strip_html_tags(item)]
    if normalized_industries:
      detail["industry"] = _clip_text(",".join(normalized_industries), 255)

  organizer_match = re.search(
    r'<li class="con_hy"><span>主办单位：</span>(.*?)</li>',
    html_text or "",
    flags=re.S
  )
  if organizer_match:
    detail["organizer_name"] = _clip_text(_strip_html_tags(organizer_match.group(1)), 255)

  location_match = re.search(
    r'<li class="site"><span class="fl">展会地点：</span><address[^>]*>(.*?)</address>',
    html_text or "",
    flags=re.S
  )
  if location_match:
    location_text = _strip_html_tags(location_match.group(1))
    location_parts = [item.strip() for item in location_text.split("-") if item and item.strip()]
    if location_parts:
      detail["country"] = _clip_text(location_parts[0], 100)
    if len(location_parts) >= 2:
      detail["city"] = _clip_text(location_parts[1], 100)
    if len(location_parts) >= 3:
      detail["venue_address"] = _clip_text(location_parts[-2], 255)
      detail["venue_name"] = _clip_text(location_parts[-1], 255)
    else:
      detail["venue_address"] = _clip_text(location_text, 255)

  data_block = re.search(r'<li class="site data_sj">(.*?)</li>', html_text or "", flags=re.S)
  if data_block:
    data_text = _strip_html_tags(data_block.group(1))
    cycle_match = re.search(r"举办周期：([^\s]+)", data_text)
    area_match = re.search(r"展览面积：(\d+)", data_text)
    exhibitors_match = re.search(r"展商数量：(\d+)", data_text)
    visitors_match = re.search(r"观众数量：(\d+)", data_text)
    if cycle_match:
      detail["cycle_text"] = _clip_text(cycle_match.group(1), 100)
    if area_match:
      detail["exhibition_area_sqm"] = _safe_parse_int(area_match.group(1))
    if exhibitors_match:
      detail["exhibitors_count"] = _safe_parse_int(exhibitors_match.group(1))
    if visitors_match:
      detail["visitors_count"] = _safe_parse_int(visitors_match.group(1))

  meta_desc_match = re.search(r'<meta name="description" content="([^"]+)"', html_text or "", flags=re.S)
  meta_desc = _strip_html_tags(meta_desc_match.group(1)) if meta_desc_match else ""
  if meta_desc:
    if not detail.get("organizer_name"):
      matched = re.search(r"主办方：([^，。]+)", meta_desc)
      if matched:
        detail["organizer_name"] = _clip_text(matched.group(1), 255)
    if not detail.get("cycle_text"):
      matched = re.search(r"举办周期：([^，。]+)", meta_desc)
      if matched:
        detail["cycle_text"] = _clip_text(matched.group(1), 100)
    if not detail.get("exhibition_area_sqm"):
      matched = re.search(r"展会面积：(\d+)", meta_desc)
      if matched:
        detail["exhibition_area_sqm"] = _safe_parse_int(matched.group(1))
    if not detail.get("visitors_count"):
      matched = re.search(r"参展观众：(\d+)", meta_desc)
      if matched:
        detail["visitors_count"] = _safe_parse_int(matched.group(1))
    if not detail.get("exhibitors_count"):
      matched = re.search(r"参展商数量及参展品牌达到(\d+)", meta_desc)
      if matched:
        detail["exhibitors_count"] = _safe_parse_int(matched.group(1))

  return {key: value for key, value in detail.items() if value not in (None, "")}


def _upsert_host_pool_event(cur, item):
  sql = (
    "INSERT INTO host_opportunity_pool_events ("
    "source_site, external_id, name, alias_name, industry, country, city, organizer_name, "
    "venue_name, venue_address, exhibition_start_date, exhibition_end_date, cycle_text, "
    "exhibition_area_sqm, exhibitors_count, visitors_count, heat_score, source_url, "
    "source_cover_url, source_list_url, is_domestic, pool_status, raw_json, fetched_at"
    ") VALUES ("
    "%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, "
    "'active', %s, CURRENT_TIMESTAMP"
    ") ON DUPLICATE KEY UPDATE "
    "name = VALUES(name), "
    "alias_name = VALUES(alias_name), "
    "industry = VALUES(industry), "
    "country = VALUES(country), "
    "city = VALUES(city), "
    "organizer_name = VALUES(organizer_name), "
    "venue_name = VALUES(venue_name), "
    "venue_address = VALUES(venue_address), "
    "exhibition_start_date = VALUES(exhibition_start_date), "
    "exhibition_end_date = VALUES(exhibition_end_date), "
    "cycle_text = VALUES(cycle_text), "
    "exhibition_area_sqm = VALUES(exhibition_area_sqm), "
    "exhibitors_count = VALUES(exhibitors_count), "
    "visitors_count = VALUES(visitors_count), "
    "heat_score = VALUES(heat_score), "
    "source_cover_url = VALUES(source_cover_url), "
    "source_list_url = VALUES(source_list_url), "
    "is_domestic = VALUES(is_domestic), "
    "raw_json = VALUES(raw_json), "
    "fetched_at = CURRENT_TIMESTAMP, "
    "updated_at = CURRENT_TIMESTAMP"
  )
  params = (
    _clip_text(item.get("source_site") or "qufair", 32),
    _clip_text(item.get("external_id"), 128),
    _clip_text(item.get("name"), 255),
    _clip_text(item.get("alias_name"), 255),
    _clip_text(item.get("industry"), 255),
    _clip_text(item.get("country"), 100),
    _clip_text(item.get("city"), 100),
    _clip_text(item.get("organizer_name"), 255),
    _clip_text(item.get("venue_name"), 255),
    _clip_text(item.get("venue_address"), 255),
    item.get("exhibition_start_date"),
    item.get("exhibition_end_date"),
    _clip_text(item.get("cycle_text"), 100),
    _safe_parse_int(item.get("exhibition_area_sqm")),
    _safe_parse_int(item.get("exhibitors_count")),
    _safe_parse_int(item.get("visitors_count")),
    _safe_parse_int(item.get("heat_score")),
    _clip_text(item.get("source_url"), 500),
    _clip_text(item.get("source_cover_url"), 500),
    _clip_text(item.get("source_list_url"), 500),
    1 if _to_bool(item.get("is_domestic"), default=True) else 0,
    json.dumps(item, ensure_ascii=False)
  )
  return cur.execute(sql, params)


def _insert_contacts(cur, opportunity_id, contacts):
  if not contacts:
    return 0
  use_role = CONTACT_ROLE_COLUMN_READY is True
  cur.execute(
    "SELECT name, phone, email FROM opportunity_contacts WHERE opportunity_id = %s",
    (opportunity_id,)
  )
  existing = {(row.get("name"), row.get("phone"), row.get("email")) for row in cur.fetchall()}
  inserted = 0
  for contact in contacts:
    name = contact.get("name")
    role = contact.get("role")
    title = contact.get("title")
    phone = contact.get("phone")
    email = contact.get("email")
    wechat = contact.get("wechat")
    key = (name, phone, email)
    if key in existing:
      continue
    if use_role:
      cur.execute(
        "INSERT INTO opportunity_contacts (opportunity_id, name, role, title, phone, email, wechat) VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (opportunity_id, name, role, title, phone, email, wechat)
      )
    else:
      cur.execute(
        "INSERT INTO opportunity_contacts (opportunity_id, name, title, phone, email, wechat) VALUES (%s, %s, %s, %s, %s, %s)",
        (opportunity_id, name, title, phone, email, wechat)
      )
    existing.add(key)
    inserted += 1
  return inserted

def _apply_analysis_defaults(analysis_data, opportunity):
  if not isinstance(analysis_data, dict):
    return analysis_data
  opportunity_type = opportunity.get("type")
  persona_defaults = HOST_PERSONA_DEFAULTS if opportunity_type == "host" else NORMAL_PERSONA_DEFAULTS
  sales_defaults = HOST_SALES_ANGLES_DEFAULTS if opportunity_type == "host" else NORMAL_SALES_ANGLES_DEFAULTS
  risk_defaults = HOST_RISK_FLAGS_DEFAULTS if opportunity_type == "host" else NORMAL_RISK_FLAGS_DEFAULTS

  persona = analysis_data.get("customer_persona")
  if not isinstance(persona, dict):
    persona = {}
  filled = False
  for key, default in persona_defaults.items():
    current = persona.get(key)
    if isinstance(default, list):
      if not isinstance(current, list) or not current:
        persona[key] = default
        filled = True
    else:
      if not current:
        persona[key] = default
        filled = True
  analysis_data["customer_persona"] = persona

  if not isinstance(analysis_data.get("sales_angles"), list) or not analysis_data.get("sales_angles"):
    analysis_data["sales_angles"] = sales_defaults
    filled = True
  if not isinstance(analysis_data.get("risk_flags"), list) or not analysis_data.get("risk_flags"):
    analysis_data["risk_flags"] = risk_defaults
    filled = True

  assumptions = analysis_data.get("assumptions")
  if not isinstance(assumptions, list):
    assumptions = []
  if filled:
    note = (
      "以下为展会型商机行业通用推测，需进一步确认"
      if opportunity_type == "host"
      else "以下为一般B2B商机行业通用推测，需进一步确认"
    )
    if note not in assumptions:
      assumptions.append(note)
    analysis_data["assumptions"] = assumptions
  return analysis_data


def _build_analysis_query(opportunity):
  parts = [
    opportunity.get("name"),
    opportunity.get("company_name"),
    opportunity.get("organizer_name"),
    opportunity.get("exhibition_name"),
    opportunity.get("city"),
    opportunity.get("industry"),
    opportunity.get("source"),
    opportunity.get("contact_name"),
    opportunity.get("contact_person"),
    opportunity.get("website"),
    "客户画像 联系方式 联系人 公司"
  ]
  return " ".join([part for part in parts if part])


def _build_analysis_messages(opportunity, sources):
  context = {
    "opportunity": {
      "name": opportunity.get("name"),
      "type": opportunity.get("type"),
      "organizer_name": opportunity.get("organizer_name"),
      "organizer_type": opportunity.get("organizer_type"),
      "exhibition_name": opportunity.get("exhibition_name"),
      "exhibition_start_date": str(opportunity.get("exhibition_start_date"))
      if opportunity.get("exhibition_start_date")
      else None,
      "exhibition_end_date": str(opportunity.get("exhibition_end_date"))
      if opportunity.get("exhibition_end_date")
      else None,
      "venue_name": opportunity.get("venue_name"),
      "venue_address": opportunity.get("venue_address"),
      "city": opportunity.get("city"),
      "industry": opportunity.get("industry"),
      "budget_range": opportunity.get("budget_range"),
      "contact_name": opportunity.get("contact_name"),
      "contact_title": opportunity.get("contact_title"),
      "contact_phone": opportunity.get("contact_phone"),
      "contact_email": opportunity.get("contact_email"),
      "company_name": opportunity.get("company_name"),
      "company_phone": opportunity.get("company_phone"),
      "company_email": opportunity.get("company_email"),
      "contact_department": opportunity.get("contact_department"),
      "contact_person": opportunity.get("contact_person"),
      "contact_address": opportunity.get("contact_address"),
      "website": opportunity.get("website"),
      "country": opportunity.get("country"),
      "hall_no": opportunity.get("hall_no"),
      "booth_no": opportunity.get("booth_no"),
      "booth_type": opportunity.get("booth_type"),
      "booth_area_sqm": opportunity.get("booth_area_sqm")
    }
  }

  blocks = []
  for idx, source in enumerate(sources, 1):
    blocks.append(
      f"[{idx}] {source.get('title') or ''}\n"
      f"URL: {source.get('url') or ''}\n"
      f"摘要: {source.get('snippet') or ''}\n"
      f"内容: {source.get('content') or ''}"
    )

  system_prompt = (
    "你是B2B展会销售的研究助手。基于提供的公开网页内容输出严格JSON，"
    "不要输出多余文字或Markdown。只能使用给定来源，无法确定就填null或空数组。"
    "允许输出公开姓名+职位+机构，并在contacts中写明source_url。"
  )
  sources_text = "\n\n".join(blocks) if blocks else "无有效来源内容"

  user_prompt = (
    f"商机信息（字段可能为空）:\n{json.dumps(context, ensure_ascii=False)}\n\n"
    f"网页来源:\n{sources_text}\n\n"
    "请返回JSON，结构如下:\n"
    f"{ANALYSIS_SCHEMA}\n"
  )
  return [
    {"role": "system", "content": system_prompt},
    {"role": "user", "content": user_prompt}
  ]


def _serialize_insight(row):
  if not row:
    return None
  analysis = _safe_json_load(row.get("analysis_json"))
  contacts = _safe_json_load(row.get("contacts_json"))
  sources = _safe_json_load(row.get("sources_json"))
  if not isinstance(contacts, list):
    contacts = []
  if not isinstance(sources, list):
    sources = []
  return {
    "analysis": analysis if isinstance(analysis, dict) else {},
    "contacts": contacts,
    "sources": sources,
    "provider": row.get("provider"),
    "model": row.get("model"),
    "updated_at": row.get("updated_at"),
    "created_at": row.get("created_at")
  }


def _workflow_json_dump(value):
  return json.dumps(value, ensure_ascii=False)


def _normalize_company_id(raw_value):
  if raw_value in (None, "", 0, "0"):
    return None
  try:
    company_id = int(raw_value)
  except (TypeError, ValueError):
    return None
  return company_id if company_id > 0 else None


def _parse_pagination_from_request(default_page_size=20, max_page_size=200):
  raw_page = request.args.get("page")
  raw_page_size = request.args.get("page_size")
  if raw_page in (None, "") and raw_page_size in (None, ""):
    return None
  try:
    page = int(raw_page or 1)
    page_size = int(raw_page_size or default_page_size)
  except (TypeError, ValueError):
    raise ValueError("invalid_pagination")
  if page <= 0 or page_size <= 0:
    raise ValueError("invalid_pagination")
  page_size = min(page_size, max_page_size)
  offset = (page - 1) * page_size
  return page, page_size, offset


def _ensure_company_exists(db, company_id):
  if not company_id:
    return True
  with db.cursor() as cur:
    cur.execute("SELECT id FROM companies WHERE id = %s", (company_id,))
    return bool(cur.fetchone())


def _normalize_unique_int_list(raw_values):
  if raw_values in (None, ""):
    return []
  if isinstance(raw_values, list):
    values = raw_values
  else:
    values = [raw_values]
  ids = []
  for raw in values:
    try:
      item_id = int(raw)
    except (TypeError, ValueError):
      continue
    if item_id > 0 and item_id not in ids:
      ids.append(item_id)
  return ids


def _normalize_org_dimension_name(raw_value):
  name = str(raw_value or "").strip()
  return name[:100]


def _normalize_org_dimension_code(raw_value):
  code = str(raw_value or "").strip()
  if not code:
    return None
  return code[:64]


def _can_view_org_dimension_scope(user, company_id):
  if is_group_admin(user):
    return True
  if is_sub_admin(user):
    own_company_id = user.get("company_id")
    return company_id in (None, own_company_id)
  return False


def _can_edit_org_dimension_scope(user, company_id):
  if is_group_admin(user):
    return True
  if is_sub_admin(user):
    own_company_id = user.get("company_id")
    return bool(own_company_id) and company_id == own_company_id
  return False


def _attach_org_dimensions_to_users(db, user_rows):
  if not isinstance(user_rows, list) or not user_rows:
    return user_rows

  user_ids = []
  for row in user_rows:
    try:
      user_id = int(row.get("id"))
    except (TypeError, ValueError, AttributeError):
      continue
    if user_id > 0 and user_id not in user_ids:
      user_ids.append(user_id)

  if not user_ids:
    for row in user_rows:
      row["org_role_ids"] = []
      row["org_role_names"] = []
      row["org_position_ids"] = []
      row["org_position_names"] = []
    return user_rows

  placeholders = ", ".join(["%s"] * len(user_ids))
  role_map = {user_id: {"ids": [], "names": []} for user_id in user_ids}
  position_map = {user_id: {"ids": [], "names": []} for user_id in user_ids}

  with db.cursor() as cur:
    cur.execute(
      f"SELECT uor.user_id, r.id AS role_id, r.name AS role_name "
      f"FROM user_org_roles uor "
      f"JOIN org_roles r ON r.id = uor.role_id "
      f"WHERE uor.user_id IN ({placeholders}) "
      f"ORDER BY uor.user_id ASC, r.name ASC, r.id ASC",
      tuple(user_ids)
    )
    for row in cur.fetchall():
      user_id = row.get("user_id")
      role_id = row.get("role_id")
      role_name = row.get("role_name")
      entry = role_map.get(user_id)
      if not entry:
        continue
      if role_id and role_id not in entry["ids"]:
        entry["ids"].append(role_id)
      if role_name and role_name not in entry["names"]:
        entry["names"].append(role_name)

    cur.execute(
      f"SELECT uop.user_id, p.id AS position_id, p.name AS position_name "
      f"FROM user_org_positions uop "
      f"JOIN org_positions p ON p.id = uop.position_id "
      f"WHERE uop.user_id IN ({placeholders}) "
      f"ORDER BY uop.user_id ASC, p.name ASC, p.id ASC",
      tuple(user_ids)
    )
    for row in cur.fetchall():
      user_id = row.get("user_id")
      position_id = row.get("position_id")
      position_name = row.get("position_name")
      entry = position_map.get(user_id)
      if not entry:
        continue
      if position_id and position_id not in entry["ids"]:
        entry["ids"].append(position_id)
      if position_name and position_name not in entry["names"]:
        entry["names"].append(position_name)

  for row in user_rows:
    user_id = row.get("id")
    role_entry = role_map.get(user_id, {"ids": [], "names": []})
    position_entry = position_map.get(user_id, {"ids": [], "names": []})
    row["org_role_ids"] = role_entry["ids"]
    row["org_role_names"] = role_entry["names"]
    row["org_position_ids"] = position_entry["ids"]
    row["org_position_names"] = position_entry["names"]

  return user_rows


def _validate_user_org_dimension_ids(db, table_name, id_column, raw_ids, user_company_id):
  ids = _normalize_unique_int_list(raw_ids)
  if not ids:
    return []

  placeholders = ", ".join(["%s"] * len(ids))
  with db.cursor() as cur:
    cur.execute(
      f"SELECT {id_column} AS item_id, company_id, status "
      f"FROM {table_name} "
      f"WHERE {id_column} IN ({placeholders})",
      tuple(ids)
    )
    rows = cur.fetchall()

  found_ids = {row.get("item_id") for row in rows}
  if any(item_id not in found_ids for item_id in ids):
    raise ValueError("invalid_org_dimension")

  allowed_company_ids = {None}
  if user_company_id:
    allowed_company_ids.add(user_company_id)

  for row in rows:
    if row.get("status") != "active":
      raise ValueError("inactive_org_dimension")
    if row.get("company_id") not in allowed_company_ids:
      raise ValueError("org_dimension_scope_mismatch")

  return ids


def _replace_user_org_dimensions(db, user_id, role_ids=None, position_ids=None):
  with db.cursor() as cur:
    if role_ids is not None:
      cur.execute("DELETE FROM user_org_roles WHERE user_id = %s", (user_id,))
      if role_ids:
        cur.executemany(
          "INSERT INTO user_org_roles (user_id, role_id) VALUES (%s, %s)",
          [(user_id, role_id) for role_id in role_ids]
        )
    if position_ids is not None:
      cur.execute("DELETE FROM user_org_positions WHERE user_id = %s", (user_id,))
      if position_ids:
        cur.executemany(
          "INSERT INTO user_org_positions (user_id, position_id) VALUES (%s, %s)",
          [(user_id, position_id) for position_id in position_ids]
        )


def _build_template_scope_sql(user, alias=""):
  prefix = f"{alias}." if alias else ""
  if is_group_admin(user):
    return "", []
  company_id = user.get("company_id")
  if company_id:
    return f"({prefix}company_id = %s OR {prefix}company_id IS NULL)", [company_id]
  return f"{prefix}company_id IS NULL", []


def _find_process_template_by_form_template(db, form_template_id, exclude_template_id=None):
  with db.cursor() as cur:
    if exclude_template_id:
      cur.execute(
        "SELECT id, name FROM approval_process_templates "
        "WHERE form_template_id = %s AND id <> %s "
        "LIMIT 1",
        (form_template_id, exclude_template_id)
      )
    else:
      cur.execute(
        "SELECT id, name FROM approval_process_templates "
        "WHERE form_template_id = %s "
        "LIMIT 1",
        (form_template_id,)
      )
    return cur.fetchone()


@contextmanager
def _db_transaction(db):
  db.begin()
  try:
    yield
  except Exception:
    db.rollback()
    raise
  else:
    db.commit()


def _build_process_owned_form_template_name(process_name):
  safe_name = str(process_name or "").strip() or "未命名流程"
  name = f"{safe_name}表单"
  return name[:255]


def _build_process_owned_form_template_description(process_name, process_description=None):
  safe_name = str(process_name or "").strip() or "未命名流程"
  safe_description = str(process_description or "").strip()
  if safe_description:
    text = f"由流程「{safe_name}」自动生成。{safe_description}"
  else:
    text = f"由流程「{safe_name}」自动生成。"
  return text[:500]


def _create_process_owned_form_template(db, process_name, process_description, company_id, schema, user_id):
  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO approval_form_templates "
      "(name, description, company_id, schema_json, status, created_by, updated_by) "
      "VALUES (%s, %s, %s, %s, 'active', %s, %s)",
      (
        _build_process_owned_form_template_name(process_name),
        _build_process_owned_form_template_description(process_name, process_description),
        company_id,
        _workflow_json_dump(schema),
        user_id,
        user_id
      )
    )
    return cur.lastrowid


def _repair_process_form_template_conflicts(db):
  with db.cursor() as cur:
    cur.execute(
      "SELECT form_template_id, COUNT(*) AS cnt "
      "FROM approval_process_templates "
      "GROUP BY form_template_id "
      "HAVING COUNT(*) > 1"
    )
    duplicate_rows = cur.fetchall()

  if not duplicate_rows:
    return 0

  repaired = 0
  with _db_transaction(db):
    for item in duplicate_rows:
      form_template_id = item.get("form_template_id")
      if not form_template_id:
        continue

      with db.cursor() as cur:
        cur.execute(
          "SELECT id, name, description, company_id, schema_json, status, created_by, updated_by "
          "FROM approval_form_templates "
          "WHERE id = %s "
          "FOR UPDATE",
          (form_template_id,)
        )
        source_form = cur.fetchone()

      if not source_form:
        continue

      with db.cursor() as cur:
        cur.execute(
          "SELECT id, name, created_by, updated_by "
          "FROM approval_process_templates "
          "WHERE form_template_id = %s "
          "ORDER BY id ASC",
          (form_template_id,)
        )
        process_rows = cur.fetchall()

      if len(process_rows) <= 1:
        continue

      for process in process_rows[1:]:
        process_id = process.get("id")
        process_name = process.get("name") or f"流程{process_id}"
        process_updated_by = process.get("updated_by") or process.get("created_by") or source_form.get("updated_by") or 1
        process_created_by = process.get("created_by") or process_updated_by
        cloned_name = _build_process_owned_form_template_name(f"{process_name}-专属")
        cloned_description = f"系统自动修复：从表单#{form_template_id}克隆，供流程#{process_id}独占使用。"

        with db.cursor() as cur:
          cur.execute(
            "INSERT INTO approval_form_templates "
            "(name, description, company_id, schema_json, status, created_by, updated_by) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (
              cloned_name,
              cloned_description[:500],
              source_form.get("company_id"),
              source_form.get("schema_json"),
              source_form.get("status") or "active",
              process_created_by,
              process_updated_by
            )
          )
          new_form_template_id = cur.lastrowid
          cur.execute(
            "UPDATE approval_process_templates "
            "SET form_template_id = %s, updated_by = %s "
            "WHERE id = %s",
            (new_form_template_id, process_updated_by, process_id)
          )
          cur.execute(
            "UPDATE approval_process_template_versions "
            "SET form_template_id = %s, updated_by = %s "
            "WHERE process_template_id = %s",
            (new_form_template_id, process_updated_by, process_id)
          )
        repaired += 1
  return repaired


def _ensure_unique_process_form_binding_index(db):
  with db.cursor() as cur:
    cur.execute("SHOW INDEX FROM approval_process_templates WHERE Key_name = 'uniq_proc_tpl_form_id'")
    if cur.fetchone():
      return False
    cur.execute(
      "ALTER TABLE approval_process_templates "
      "ADD UNIQUE KEY uniq_proc_tpl_form_id (form_template_id)"
    )
  return True


def _normalize_idempotency_key(raw_key):
  if raw_key in (None, ""):
    return ""
  key = str(raw_key).strip()
  if not key:
    return ""
  return key[:WORKFLOW_IDEMPOTENCY_KEY_MAX_LEN]


def _load_action_idempotency_response(db, idempotency_key, instance_id, actor_id, action):
  if not idempotency_key:
    return None, None
  with db.cursor() as cur:
    cur.execute(
      "SELECT response_json, status_code "
      "FROM approval_action_idempotency "
      "WHERE idem_key = %s AND instance_id = %s AND actor_id = %s AND action = %s "
      "LIMIT 1",
      (idempotency_key, instance_id, actor_id, action)
    )
    row = cur.fetchone()

  if not row:
    return None, None

  response_json = row.get("response_json")
  if response_json in (None, ""):
    return None, None

  parsed = _safe_json_load(response_json)
  if not isinstance(parsed, dict):
    return None, None

  try:
    status_code = int(row.get("status_code") or 200)
  except (TypeError, ValueError):
    status_code = 200
  if status_code < 100 or status_code > 599:
    status_code = 200
  return parsed, status_code


def _save_action_idempotency_response(db, idempotency_key, instance_id, actor_id, action, response_body, status_code=200):
  if not idempotency_key:
    return
  payload = response_body if isinstance(response_body, dict) else {"data": response_body}
  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO approval_action_idempotency "
      "(idem_key, instance_id, actor_id, action, response_json, status_code) "
      "VALUES (%s, %s, %s, %s, %s, %s) "
      "ON DUPLICATE KEY UPDATE "
      "response_json = VALUES(response_json), "
      "status_code = VALUES(status_code), "
      "updated_at = CURRENT_TIMESTAMP",
      (
        idempotency_key,
        instance_id,
        actor_id,
        action,
        _workflow_json_dump(payload),
        int(status_code or 200)
      )
    )


def _create_process_template_version(
  db,
  template_id,
  version_no,
  form_template_id,
  definition,
  status,
  user_id,
  published_at=None
):
  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO approval_process_template_versions "
      "(process_template_id, version_no, form_template_id, definition_json, status, published_at, created_by, updated_by) "
      "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
      (
        template_id,
        version_no,
        form_template_id,
        _workflow_json_dump(definition),
        status,
        published_at,
        user_id,
        user_id
      )
    )


def _publish_process_template_version(db, template_id, version_no, user_id):
  with db.cursor() as cur:
    cur.execute(
      "UPDATE approval_process_template_versions "
      "SET status = CASE WHEN version_no = %s THEN 'published' ELSE IF(status = 'published', 'archived', status) END, "
      "published_at = CASE WHEN version_no = %s THEN COALESCE(published_at, CURRENT_TIMESTAMP) ELSE published_at END, "
      "updated_by = %s "
      "WHERE process_template_id = %s",
      (version_no, version_no, user_id, template_id)
    )


def _get_process_template_version(db, template_id, version_no):
  with db.cursor() as cur:
    cur.execute(
      "SELECT * FROM approval_process_template_versions "
      "WHERE process_template_id = %s AND version_no = %s "
      "LIMIT 1",
      (template_id, version_no)
    )
    return cur.fetchone()


def _log_instance_event(db, instance_id, user_id, action, task_id=None, comment=None, detail=None):
  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO approval_instance_events "
      "(instance_id, task_id, user_id, action, detail_json, comment) "
      "VALUES (%s, %s, %s, %s, %s, %s)",
      (
        instance_id,
        task_id,
        user_id,
        action,
        _workflow_json_dump(detail) if detail not in (None, "") else None,
        comment
      )
    )


def _normalize_workflow_schema(raw_schema):
  if not isinstance(raw_schema, list) or not raw_schema:
    raise ValueError("invalid_schema")

  seen_keys = set()
  schema = []
  for idx, raw_field in enumerate(raw_schema, start=1):
    if not isinstance(raw_field, dict):
      raise ValueError("invalid_schema")

    key = str(raw_field.get("key") or "").strip()
    label = str(raw_field.get("label") or "").strip()
    field_type = str(raw_field.get("type") or "text").strip().lower()
    required = bool(raw_field.get("required"))
    placeholder = raw_field.get("placeholder")
    options = raw_field.get("options") if isinstance(raw_field.get("options"), list) else []
    columns = raw_field.get("columns") if isinstance(raw_field.get("columns"), list) else []
    max_count = raw_field.get("max_count")
    default_value = raw_field.get("default")

    if not key or not WORKFLOW_FIELD_KEY_PATTERN.match(key):
      raise ValueError("invalid_field_key")
    if key in seen_keys:
      raise ValueError("duplicated_field_key")
    if not label:
      raise ValueError("invalid_field_label")
    if field_type not in WORKFLOW_FIELD_TYPES:
      raise ValueError("invalid_field_type")

    normalized_options = []
    normalized_columns = []
    if field_type == "select":
      for option in options:
        text = str(option).strip()
        if text and text not in normalized_options:
          normalized_options.append(text)
      if not normalized_options:
        raise ValueError("invalid_field_options")
      if default_value not in (None, "") and str(default_value) not in normalized_options:
        raise ValueError("invalid_field_default")
      if default_value not in (None, ""):
        default_value = str(default_value)
    elif field_type == "table":
      seen_column_keys = set()
      for raw_column in columns:
        if not isinstance(raw_column, dict):
          raise ValueError("invalid_field_columns")
        col_key = str(raw_column.get("key") or "").strip()
        col_label = str(raw_column.get("label") or "").strip()
        col_type = str(raw_column.get("type") or "text").strip().lower()
        if not col_key or not WORKFLOW_FIELD_KEY_PATTERN.match(col_key):
          raise ValueError("invalid_field_columns")
        if col_key in seen_column_keys:
          raise ValueError("invalid_field_columns")
        if not col_label:
          raise ValueError("invalid_field_columns")
        if col_type not in {"text", "textarea", "number", "date", "select", "boolean"}:
          raise ValueError("invalid_field_columns")
        column = {
          "key": col_key,
          "label": col_label,
          "type": col_type
        }
        if col_type == "select":
          col_options = []
          for option in raw_column.get("options") or []:
            text = str(option).strip()
            if text and text not in col_options:
              col_options.append(text)
          if not col_options:
            raise ValueError("invalid_field_columns")
          column["options"] = col_options
        normalized_columns.append(column)
        seen_column_keys.add(col_key)
      if not normalized_columns:
        raise ValueError("invalid_field_columns")
      if default_value not in (None, "") and not isinstance(default_value, list):
        raise ValueError("invalid_field_default")
    else:
      normalized_options = []

    if field_type in {"text", "textarea"} and default_value not in (None, ""):
      default_value = str(default_value)
    if field_type == "number" and default_value not in (None, ""):
      if isinstance(default_value, bool):
        raise ValueError("invalid_field_default")
      try:
        default_value = float(default_value)
      except (TypeError, ValueError):
        raise ValueError("invalid_field_default")
    if field_type == "boolean" and default_value not in (None, "") and not isinstance(default_value, bool):
      raise ValueError("invalid_field_default")
    if field_type == "date" and default_value not in (None, ""):
      if not isinstance(default_value, str) or not re.match(r"^\d{4}-\d{2}-\d{2}$", default_value):
        raise ValueError("invalid_field_default")
    if field_type == "attachment":
      if default_value not in (None, "") and not isinstance(default_value, list):
        raise ValueError("invalid_field_default")
      if max_count not in (None, ""):
        try:
          max_count = int(max_count)
        except (TypeError, ValueError):
          raise ValueError("invalid_field_max_count")
        if max_count <= 0:
          raise ValueError("invalid_field_max_count")
    else:
      max_count = None

    field = {
      "key": key,
      "label": label,
      "type": field_type,
      "required": required,
      "options": normalized_options,
      "default": default_value,
      "order": idx
    }
    if normalized_columns:
      field["columns"] = normalized_columns
    if max_count:
      field["max_count"] = max_count
    if placeholder not in (None, ""):
      field["placeholder"] = str(placeholder)
    schema.append(field)
    seen_keys.add(key)

  return schema


def _normalize_workflow_steps(raw_steps):
  if not isinstance(raw_steps, list) or not raw_steps:
    raise ValueError("invalid_steps")

  steps = []
  for idx, raw_step in enumerate(raw_steps, start=1):
    if not isinstance(raw_step, dict):
      raise ValueError("invalid_steps")

    name = str(raw_step.get("name") or "").strip()
    step_type = str(raw_step.get("step_type") or "approval").strip().lower()
    raw_condition = raw_step.get("condition")

    if not name:
      raise ValueError("invalid_step_name")
    if step_type not in WORKFLOW_STEP_TYPES:
      raise ValueError("invalid_step_type")

    step = {
      "step_no": idx,
      "name": name,
      "step_type": step_type
    }

    if step_type in {"approval", "cc"}:
      approver_type = str(raw_step.get("approver_type") or "user").strip().lower()
      approval_mode = str(raw_step.get("approval_mode") or "any").strip().lower()
      approval_type = str(raw_step.get("approval_type") or "").strip().lower()
      allow_self_raw = raw_step.get("allow_self_approve")
      allow_self_approve = True if allow_self_raw is None else bool(allow_self_raw)
      allow_return_raw = raw_step.get("allow_return")
      allow_return = True if allow_return_raw is None else _normalize_bool_flag(allow_return_raw, default=True)
      timeout_hours_raw = raw_step.get("timeout_hours")
      field_permissions = _normalize_field_permissions(raw_step.get("field_permissions"))
      approver_groups = _normalize_approver_group_definitions(raw_step.get("approver_groups"))

      if approver_type not in WORKFLOW_APPROVER_TYPES:
        raise ValueError("invalid_step_approver_type")
      if approval_mode not in WORKFLOW_APPROVAL_MODES:
        raise ValueError("invalid_step_approval_mode")
      if approval_type:
        if approval_type not in WORKFLOW_APPROVAL_TYPES:
          raise ValueError("invalid_step_approval_mode")
        step["approval_type"] = approval_type
        if approval_type == "all":
          approval_mode = "all"
        elif approval_type == "any":
          approval_mode = "any"

      step["approver_type"] = approver_type
      step["approval_mode"] = approval_mode
      step["allow_self_approve"] = allow_self_approve
      step["allow_return"] = allow_return
      try:
        timeout_hours = int(timeout_hours_raw)
      except (TypeError, ValueError):
        timeout_hours = 0
      if timeout_hours > 0:
        step["timeout_hours"] = timeout_hours
      if field_permissions:
        step["field_permissions"] = field_permissions
      if approver_groups:
        step["approver_groups"] = approver_groups

      if approver_type == "user":
        user_ids = []
        for raw_user_id in raw_step.get("approver_user_ids") or []:
          try:
            user_id = int(raw_user_id)
          except (TypeError, ValueError):
            continue
          if user_id > 0 and user_id not in user_ids:
            user_ids.append(user_id)
        if not user_ids:
          raise ValueError("missing_step_user_approvers")
        step["approver_user_ids"] = user_ids

      if approver_type == "role":
        roles = []
        for role in raw_step.get("approver_roles") or []:
          role_name = str(role).strip()
          if role_name and role_name not in roles:
            roles.append(role_name)
        if not roles:
          raise ValueError("missing_step_role_approvers")
        step["approver_roles"] = roles

      if approver_type == "position":
        positions = []
        for position in raw_step.get("approver_positions") or []:
          position_name = str(position).strip()
          if position_name and position_name not in positions:
            positions.append(position_name)
        if not positions:
          raise ValueError("missing_step_position_approvers")
        step["approver_positions"] = positions

      if approver_type == "applicant_select":
        approver_field_key = str(raw_step.get("approver_field_key") or "").strip()
        if not approver_field_key or not WORKFLOW_FIELD_KEY_PATTERN.match(approver_field_key):
          raise ValueError("invalid_step_applicant_select_field")
        step["approver_field_key"] = approver_field_key

      if approver_type == "previous_handler":
        previous_step_offset_raw = raw_step.get("previous_step_offset", 1)
        try:
          previous_step_offset = int(previous_step_offset_raw)
        except (TypeError, ValueError):
          raise ValueError("invalid_step_previous_handler")
        if previous_step_offset <= 0:
          raise ValueError("invalid_step_previous_handler")
        step["previous_step_offset"] = previous_step_offset

    if step_type == "subprocess":
      try:
        subprocess_template_id = int(raw_step.get("subprocess_template_id"))
      except (TypeError, ValueError):
        raise ValueError("invalid_subprocess_template")
      if subprocess_template_id <= 0:
        raise ValueError("invalid_subprocess_template")
      step["subprocess_template_id"] = subprocess_template_id

    if raw_condition not in (None, ""):
      step["condition"] = _normalize_condition_definition(raw_condition)

    steps.append(step)

  return steps


def _normalize_condition_definition(raw_condition):
  if raw_condition in (None, ""):
    return None
  if not isinstance(raw_condition, dict):
    raise ValueError("invalid_step_condition")

  logic = str(raw_condition.get("logic") or "and").strip().lower()
  if logic not in WORKFLOW_CONDITION_LOGICS:
    raise ValueError("invalid_step_condition_logic")
  expression = raw_condition.get("expression")
  if expression in (None, ""):
    expression = None
  else:
    expression = str(expression).strip()
    if not expression:
      raise ValueError("invalid_step_condition_expression")

  normalized_rules = []
  for raw_rule in raw_condition.get("rules") or []:
    if not isinstance(raw_rule, dict):
      raise ValueError("invalid_step_condition_rule")
    field = str(raw_rule.get("field") or "").strip()
    operator = str(raw_rule.get("operator") or "eq").strip().lower()
    if not field:
      raise ValueError("invalid_step_condition_field")
    if operator not in WORKFLOW_CONDITION_OPERATORS:
      raise ValueError("invalid_step_condition_operator")
    normalized_rules.append(
      {
        "field": field,
        "operator": operator,
        "value": raw_rule.get("value")
      }
    )

  if not normalized_rules and not expression:
    raise ValueError("invalid_step_condition")

  normalized = {"logic": logic}
  if normalized_rules:
    normalized["rules"] = normalized_rules
  if expression:
    normalized["expression"] = expression
  return normalized


def _normalize_bool_flag(raw_value, default=False):
  if raw_value is None:
    return default
  if isinstance(raw_value, str):
    text = raw_value.strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
      return True
    if text in {"0", "false", "no", "n", "off"}:
      return False
  return bool(raw_value)


def _normalize_field_permissions(raw_permissions):
  if not isinstance(raw_permissions, list):
    return []
  permissions = []
  seen_keys = set()
  for raw_item in raw_permissions:
    if not isinstance(raw_item, dict):
      continue
    field_key = str(raw_item.get("field_key") or raw_item.get("key") or "").strip()
    if not field_key or not WORKFLOW_FIELD_KEY_PATTERN.match(field_key):
      continue
    if field_key in seen_keys:
      continue
    seen_keys.add(field_key)
    permissions.append(
      {
        "field_key": field_key,
        "can_view": _normalize_bool_flag(raw_item.get("can_view"), default=True),
        "can_edit": _normalize_bool_flag(raw_item.get("can_edit"), default=False),
        "required": _normalize_bool_flag(raw_item.get("required"), default=False)
      }
    )
  return permissions


def _normalize_approver_group_definitions(raw_groups):
  if not isinstance(raw_groups, list):
    return []
  groups = []
  for idx, raw_group in enumerate(raw_groups, start=1):
    if not isinstance(raw_group, dict):
      continue
    approver_type = str(raw_group.get("approver_type") or "manager").strip().lower()
    if approver_type not in WORKFLOW_APPROVER_TYPES:
      continue
    group_id = str(raw_group.get("id") or f"group_{idx}").strip()
    if not group_id:
      group_id = f"group_{idx}"
    group = {
      "id": group_id[:64],
      "name": str(raw_group.get("name") or f"审批组{idx}").strip() or f"审批组{idx}",
      "approver_type": approver_type
    }

    user_ids = []
    for raw_user_id in raw_group.get("approver_user_ids") or []:
      try:
        user_id = int(raw_user_id)
      except (TypeError, ValueError):
        continue
      if user_id > 0 and user_id not in user_ids:
        user_ids.append(user_id)
    if approver_type == "user":
      group["approver_user_ids"] = user_ids

    roles = []
    for role in raw_group.get("approver_roles") or []:
      role_name = str(role).strip()
      if role_name and role_name not in roles:
        roles.append(role_name)
    if approver_type == "role":
      group["approver_roles"] = roles

    positions = []
    for position in raw_group.get("approver_positions") or []:
      position_name = str(position).strip()
      if position_name and position_name not in positions:
        positions.append(position_name)
    if approver_type == "position":
      group["approver_positions"] = positions

    if approver_type == "applicant_select":
      approver_field_key = str(raw_group.get("approver_field_key") or "").strip()
      if approver_field_key and WORKFLOW_FIELD_KEY_PATTERN.match(approver_field_key):
        group["approver_field_key"] = approver_field_key

    if approver_type == "previous_handler":
      try:
        previous_step_offset = int(raw_group.get("previous_step_offset") or 1)
      except (TypeError, ValueError):
        previous_step_offset = 1
      if previous_step_offset > 0:
        group["previous_step_offset"] = previous_step_offset

    cc_user_ids = []
    for raw_user_id in raw_group.get("cc_user_ids") or []:
      try:
        user_id = int(raw_user_id)
      except (TypeError, ValueError):
        continue
      if user_id > 0 and user_id not in cc_user_ids:
        cc_user_ids.append(user_id)
    if cc_user_ids:
      group["cc_user_ids"] = cc_user_ids

    condition = raw_group.get("condition")
    if condition not in (None, ""):
      try:
        normalized_condition = _normalize_condition_definition(condition)
      except ValueError:
        normalized_condition = None
      if normalized_condition:
        group["condition"] = normalized_condition

    groups.append(group)
  return groups


def _normalize_workflow_graph_definition(raw_definition):
  if not isinstance(raw_definition, dict):
    raise ValueError("invalid_definition")

  raw_nodes = raw_definition.get("nodes")
  raw_edges = raw_definition.get("edges")
  if not isinstance(raw_nodes, list) or not raw_nodes:
    raise ValueError("invalid_definition_nodes")
  if not isinstance(raw_edges, list):
    raise ValueError("invalid_definition_edges")

  nodes = []
  node_ids = set()
  start_nodes = []
  end_nodes = []

  for idx, raw_node in enumerate(raw_nodes, start=1):
    if not isinstance(raw_node, dict):
      raise ValueError("invalid_definition_node")

    node_id = str(raw_node.get("id") or "").strip()
    node_type = str(raw_node.get("node_type") or "").strip().lower()
    name = str(raw_node.get("name") or "").strip()

    if not node_id or not WORKFLOW_NODE_ID_PATTERN.match(node_id):
      raise ValueError("invalid_definition_node_id")
    if node_id in node_ids:
      raise ValueError("duplicated_definition_node_id")
    if node_type not in WORKFLOW_NODE_TYPES:
      raise ValueError("invalid_definition_node_type")
    if not name:
      default_names = {
        "start": "开始",
        "end": "结束",
        "condition": "条件分支",
        "approval": "审批",
        "cc": "抄送",
        "parallel_start": "并行分支",
        "parallel_join": "并行汇聚",
        "subprocess": "子流程"
      }
      name = default_names.get(node_type, f"节点{idx}")

    node = {
      "id": node_id,
      "name": name,
      "node_type": node_type
    }

    position = raw_node.get("position")
    if isinstance(position, dict):
      try:
        node["position"] = {
          "x": float(position.get("x", 0)),
          "y": float(position.get("y", 0))
        }
      except (TypeError, ValueError):
        pass

    if node_type in {"approval", "cc"}:
      approver_groups = _normalize_approver_group_definitions(raw_node.get("approver_groups"))
      primary_group = approver_groups[0] if approver_groups else {}

      approver_type = str(
        raw_node.get("approver_type") or primary_group.get("approver_type") or "manager"
      ).strip().lower()
      raw_approval_type = str(raw_node.get("approval_type") or "").strip().lower()
      approval_mode = str(raw_node.get("approval_mode") or "any").strip().lower()
      if raw_approval_type in WORKFLOW_APPROVAL_TYPES:
        node["approval_type"] = raw_approval_type
        if raw_approval_type == "all":
          approval_mode = "all"
        elif raw_approval_type == "any":
          approval_mode = "any"
      allow_self_raw = raw_node.get("allow_self_approve")
      allow_self_approve = True if allow_self_raw is None else _normalize_bool_flag(allow_self_raw, default=True)
      allow_return_raw = raw_node.get("allow_return")
      allow_return = True if allow_return_raw is None else _normalize_bool_flag(allow_return_raw, default=True)
      timeout_hours_raw = raw_node.get("timeout_hours")
      field_permissions = _normalize_field_permissions(raw_node.get("field_permissions"))

      if approver_type not in WORKFLOW_APPROVER_TYPES:
        raise ValueError("invalid_step_approver_type")
      if approval_mode not in WORKFLOW_APPROVAL_MODES:
        raise ValueError("invalid_step_approval_mode")

      node["approver_type"] = approver_type
      node["approval_mode"] = approval_mode
      node["allow_self_approve"] = allow_self_approve
      node["allow_return"] = allow_return
      try:
        timeout_hours = int(timeout_hours_raw)
      except (TypeError, ValueError):
        timeout_hours = 0
      if timeout_hours > 0:
        node["timeout_hours"] = timeout_hours
      if field_permissions:
        node["field_permissions"] = field_permissions
      if approver_groups:
        node["approver_groups"] = approver_groups

      if approver_type == "user":
        user_ids = []
        raw_user_ids = raw_node.get("approver_user_ids")
        if not raw_user_ids and primary_group.get("approver_type") == "user":
          raw_user_ids = primary_group.get("approver_user_ids")
        for raw_user_id in raw_user_ids or []:
          try:
            user_id = int(raw_user_id)
          except (TypeError, ValueError):
            continue
          if user_id > 0 and user_id not in user_ids:
            user_ids.append(user_id)
        if not user_ids:
          raise ValueError("missing_step_user_approvers")
        node["approver_user_ids"] = user_ids

      if approver_type == "role":
        roles = []
        raw_roles = raw_node.get("approver_roles")
        if not raw_roles and primary_group.get("approver_type") == "role":
          raw_roles = primary_group.get("approver_roles")
        for role in raw_roles or []:
          role_name = str(role).strip()
          if role_name and role_name not in roles:
            roles.append(role_name)
        if not roles:
          raise ValueError("missing_step_role_approvers")
        node["approver_roles"] = roles

      if approver_type == "position":
        positions = []
        raw_positions = raw_node.get("approver_positions")
        if not raw_positions and primary_group.get("approver_type") == "position":
          raw_positions = primary_group.get("approver_positions")
        for position in raw_positions or []:
          position_name = str(position).strip()
          if position_name and position_name not in positions:
            positions.append(position_name)
        if not positions:
          raise ValueError("missing_step_position_approvers")
        node["approver_positions"] = positions

      if approver_type == "applicant_select":
        approver_field_key = str(
          raw_node.get("approver_field_key") or primary_group.get("approver_field_key") or ""
        ).strip()
        if not approver_field_key or not WORKFLOW_FIELD_KEY_PATTERN.match(approver_field_key):
          raise ValueError("invalid_step_applicant_select_field")
        node["approver_field_key"] = approver_field_key

      if approver_type == "previous_handler":
        previous_step_offset_raw = raw_node.get("previous_step_offset", primary_group.get("previous_step_offset", 1))
        try:
          previous_step_offset = int(previous_step_offset_raw)
        except (TypeError, ValueError):
          raise ValueError("invalid_step_previous_handler")
        if previous_step_offset <= 0:
          raise ValueError("invalid_step_previous_handler")
        node["previous_step_offset"] = previous_step_offset

    if node_type == "subprocess":
      try:
        subprocess_template_id = int(raw_node.get("subprocess_template_id"))
      except (TypeError, ValueError):
        raise ValueError("invalid_subprocess_template")
      if subprocess_template_id <= 0:
        raise ValueError("invalid_subprocess_template")
      node["subprocess_template_id"] = subprocess_template_id

    condition = _normalize_condition_definition(raw_node.get("condition"))
    if condition:
      node["condition"] = condition

    if node_type == "start":
      start_nodes.append(node_id)
    if node_type == "end":
      end_nodes.append(node_id)

    nodes.append(node)
    node_ids.add(node_id)

  if len(start_nodes) != 1:
    raise ValueError("invalid_start_node")
  if not end_nodes:
    raise ValueError("missing_end_node")

  start_node_id = str(raw_definition.get("start_node_id") or start_nodes[0]).strip()
  if start_node_id not in node_ids:
    raise ValueError("invalid_start_node")

  edges = []
  for idx, raw_edge in enumerate(raw_edges, start=1):
    if not isinstance(raw_edge, dict):
      raise ValueError("invalid_definition_edge")
    source = str(raw_edge.get("source") or "").strip()
    target = str(raw_edge.get("target") or "").strip()
    if source not in node_ids or target not in node_ids:
      raise ValueError("invalid_definition_edge_target")
    if source == target:
      raise ValueError("invalid_definition_edge_target")

    edge_id = str(raw_edge.get("id") or f"e_{idx}_{source}_{target}").strip()
    if not edge_id:
      edge_id = f"e_{idx}_{source}_{target}"
    try:
      priority = int(raw_edge.get("priority", idx))
    except (TypeError, ValueError):
      priority = idx
    edge = {
      "id": edge_id,
      "source": source,
      "target": target,
      "priority": priority
    }
    condition = _normalize_condition_definition(raw_edge.get("condition"))
    if condition:
      edge["condition"] = condition
    label = raw_edge.get("label")
    if label not in (None, ""):
      edge["label"] = str(label)
    is_default_raw = raw_edge.get("is_default")
    if isinstance(is_default_raw, str):
      is_default = is_default_raw.strip().lower() in {"1", "true", "yes", "y"}
    else:
      is_default = bool(is_default_raw)
    if is_default:
      edge["is_default"] = True
    edges.append(edge)

  return {
    "version": "graph_v1",
    "start_node_id": start_node_id,
    "nodes": nodes,
    "edges": edges
  }


def _steps_to_graph_definition(steps):
  nodes = [
    {"id": "start", "name": "开始", "node_type": "start"},
    {"id": "end", "name": "结束", "node_type": "end"}
  ]
  edges = []
  previous_id = "start"

  for idx, step in enumerate(steps, start=1):
    node_id = f"step_{idx}"
    node_type = str(step.get("step_type") or "approval").strip().lower()
    if node_type not in WORKFLOW_NODE_TYPES:
      node_type = "approval"
    node = {
      "id": node_id,
      "name": step.get("name") or f"节点{idx}",
      "node_type": node_type
    }
    for key in [
      "approver_type",
      "approval_mode",
      "approval_type",
      "allow_self_approve",
      "allow_return",
      "timeout_hours",
      "field_permissions",
      "approver_groups",
      "approver_user_ids",
      "approver_roles",
      "approver_positions",
      "approver_field_key",
      "previous_step_offset",
      "subprocess_template_id",
      "condition"
    ]:
      if key in step:
        node[key] = step.get(key)
    nodes.append(node)
    edges.append(
      {
        "id": f"e_{previous_id}_{node_id}",
        "source": previous_id,
        "target": node_id,
        "priority": idx
      }
    )
    previous_id = node_id

  edges.append(
    {
      "id": f"e_{previous_id}_end",
      "source": previous_id,
      "target": "end",
      "priority": len(edges) + 1
    }
  )
  return {
    "version": "graph_v1",
    "start_node_id": "start",
    "nodes": nodes,
    "edges": edges
  }


def _normalize_workflow_definition(raw_definition):
  if isinstance(raw_definition, dict) and isinstance(raw_definition.get("nodes"), list):
    return _normalize_workflow_graph_definition(raw_definition)
  steps = _normalize_workflow_steps(raw_definition)
  return _steps_to_graph_definition(steps)


def _workflow_validation_issue(code, message, nodes=None, edge_ids=None):
  issue = {"code": code, "message": message}
  if nodes:
    issue["nodes"] = sorted({str(node_id) for node_id in nodes if node_id})
  if edge_ids:
    issue["edge_ids"] = sorted({str(edge_id) for edge_id in edge_ids if edge_id})
  return issue


def _build_definition_links(definition):
  nodes_by_id, outgoing_edges = _build_definition_index(definition)
  incoming_edges = {node_id: [] for node_id in nodes_by_id.keys()}
  for source, items in outgoing_edges.items():
    if source not in incoming_edges:
      incoming_edges[source] = []
    for edge in items:
      target = edge.get("target")
      if target not in incoming_edges:
        incoming_edges[target] = []
      incoming_edges[target].append(edge)
  return nodes_by_id, outgoing_edges, incoming_edges


def _is_default_branch_edge(edge):
  if not isinstance(edge, dict):
    return False
  if bool(edge.get("is_default")):
    return True
  return not isinstance(edge.get("condition"), dict)


def _graph_has_cycle_from_start(start_node_id, outgoing_edges, allowed_nodes):
  if not start_node_id:
    return False
  visited = set()
  stack = set()

  def dfs(node_id):
    if node_id not in allowed_nodes:
      return False
    if node_id in stack:
      return True
    if node_id in visited:
      return False
    visited.add(node_id)
    stack.add(node_id)
    for edge in outgoing_edges.get(node_id, []):
      target = edge.get("target")
      if target in allowed_nodes and dfs(target):
        return True
    stack.remove(node_id)
    return False

  return dfs(start_node_id)


def _validate_workflow_definition(definition):
  errors = []
  warnings = []

  if not isinstance(definition, dict):
    errors.append(_workflow_validation_issue("invalid_definition", "流程定义结构不合法。"))
    return {"valid": False, "errors": errors, "warnings": warnings}

  nodes_by_id, outgoing_edges, incoming_edges = _build_definition_links(definition)
  if not nodes_by_id:
    errors.append(_workflow_validation_issue("invalid_definition_nodes", "流程图没有可用节点。"))
    return {"valid": False, "errors": errors, "warnings": warnings}

  start_node_id = definition.get("start_node_id")
  start_nodes = [node_id for node_id, node in nodes_by_id.items() if node.get("node_type") == "start"]
  end_nodes = [node_id for node_id, node in nodes_by_id.items() if node.get("node_type") == "end"]

  if len(start_nodes) != 1 or start_node_id not in nodes_by_id:
    errors.append(
      _workflow_validation_issue(
        "invalid_start_node",
        "流程图必须且只能有一个开始节点，并与 start_node_id 对齐。",
        nodes=start_nodes + [start_node_id]
      )
    )
  else:
    if nodes_by_id[start_node_id].get("node_type") != "start":
      errors.append(
        _workflow_validation_issue(
          "invalid_start_node",
          "start_node_id 必须指向开始节点。",
          nodes=[start_node_id]
        )
      )

  if not end_nodes:
    errors.append(_workflow_validation_issue("missing_end_node", "流程图至少需要一个结束节点。"))

  if start_node_id in nodes_by_id:
    start_incoming = incoming_edges.get(start_node_id) or []
    if start_incoming:
      errors.append(
        _workflow_validation_issue(
          "start_node_has_incoming_edge",
          "开始节点不能有入线。",
          nodes=[start_node_id],
          edge_ids=[edge.get("id") for edge in start_incoming]
        )
      )

  end_with_outgoing = [node_id for node_id in end_nodes if outgoing_edges.get(node_id)]
  if end_with_outgoing:
    edge_ids = []
    for node_id in end_with_outgoing:
      edge_ids.extend([edge.get("id") for edge in outgoing_edges.get(node_id, [])])
    errors.append(
      _workflow_validation_issue(
        "end_node_has_outgoing_edge",
        "结束节点不能配置出线。",
        nodes=end_with_outgoing,
        edge_ids=edge_ids
      )
    )

  no_outgoing_nodes = [
    node_id
    for node_id, node in nodes_by_id.items()
    if node.get("node_type") != "end" and not outgoing_edges.get(node_id)
  ]
  if no_outgoing_nodes:
    errors.append(
      _workflow_validation_issue(
        "node_missing_outgoing_edge",
        "除结束节点外，其余节点都必须至少有一条出线。",
        nodes=no_outgoing_nodes
      )
    )

  for node_id, node in nodes_by_id.items():
    node_type = node.get("node_type")
    node_outgoing = outgoing_edges.get(node_id, [])
    if node_type == "condition":
      if len(node_outgoing) < 2:
        errors.append(
          _workflow_validation_issue(
            "condition_node_requires_branches",
            "条件节点至少需要两条分支。",
            nodes=[node_id]
          )
        )
      default_edges = [edge for edge in node_outgoing if _is_default_branch_edge(edge)]
      if not default_edges:
        errors.append(
          _workflow_validation_issue(
            "condition_node_missing_default_branch",
            "条件节点至少要有一条默认分支（无条件或标记默认）。",
            nodes=[node_id]
          )
        )
      elif len(default_edges) > 1:
        errors.append(
          _workflow_validation_issue(
            "condition_node_multiple_default_branch",
            "条件节点只能有一条默认分支。",
            nodes=[node_id],
            edge_ids=[edge.get("id") for edge in default_edges]
          )
        )
    elif node_type == "parallel_start":
      if len(node_outgoing) < 2:
        errors.append(
          _workflow_validation_issue(
            "parallel_start_requires_branches",
            "并行分支节点至少需要两条分支。",
            nodes=[node_id]
          )
        )
    elif node_type == "parallel_join":
      incoming_count = len(incoming_edges.get(node_id, []))
      if incoming_count < 2:
        errors.append(
          _workflow_validation_issue(
            "parallel_join_requires_incoming",
            "并行汇聚节点至少需要两条入线。",
            nodes=[node_id]
          )
        )
    elif node_type == "subprocess":
      subprocess_template_id = node.get("subprocess_template_id")
      try:
        subprocess_template_id = int(subprocess_template_id)
      except (TypeError, ValueError):
        subprocess_template_id = 0
      if subprocess_template_id <= 0:
        errors.append(
          _workflow_validation_issue(
            "invalid_subprocess_template",
            "子流程节点缺少有效的 subprocess_template_id。",
            nodes=[node_id]
          )
        )
    elif node_type not in {"start", "end"} and len(node_outgoing) > 1:
      warnings.append(
        _workflow_validation_issue(
          "non_condition_multi_branch",
          "非条件节点存在多条分支，将按优先级和条件命中顺序路由。",
          nodes=[node_id]
        )
      )

  for source, items in outgoing_edges.items():
    for edge in items:
      if edge.get("is_default") and edge.get("condition"):
        warnings.append(
          _workflow_validation_issue(
            "default_branch_with_condition",
            "默认分支同时配置了条件，条件命中失败时仍会作为兜底分支。",
            nodes=[source, edge.get("target")],
            edge_ids=[edge.get("id")]
          )
        )

  reachable_nodes = set()
  if start_node_id in nodes_by_id:
    pending = [start_node_id]
    while pending:
      current = pending.pop()
      if current in reachable_nodes:
        continue
      reachable_nodes.add(current)
      for edge in outgoing_edges.get(current, []):
        target = edge.get("target")
        if target in nodes_by_id and target not in reachable_nodes:
          pending.append(target)

  unreachable_nodes = [node_id for node_id in nodes_by_id.keys() if node_id not in reachable_nodes]
  if unreachable_nodes:
    errors.append(
      _workflow_validation_issue(
        "unreachable_nodes",
        "存在无法从开始节点到达的节点。",
        nodes=unreachable_nodes
      )
    )

  can_reach_end = set()
  pending_end = list(end_nodes)
  while pending_end:
    current = pending_end.pop()
    if current in can_reach_end:
      continue
    can_reach_end.add(current)
    for edge in incoming_edges.get(current, []):
      source = edge.get("source")
      if source in nodes_by_id and source not in can_reach_end:
        pending_end.append(source)

  dead_end_nodes = [
    node_id
    for node_id in reachable_nodes
    if nodes_by_id.get(node_id, {}).get("node_type") != "end" and node_id not in can_reach_end
  ]
  if dead_end_nodes:
    errors.append(
      _workflow_validation_issue(
        "dead_end_nodes",
        "存在无法走到结束节点的死路节点。",
        nodes=dead_end_nodes
      )
    )

  if (
    start_node_id in nodes_by_id
    and reachable_nodes
    and _graph_has_cycle_from_start(start_node_id, outgoing_edges, reachable_nodes)
  ):
    errors.append(
      _workflow_validation_issue(
        "graph_has_cycle",
        "流程图包含环路，当前审批引擎不支持循环回路。"
      )
    )

  return {"valid": len(errors) == 0, "errors": errors, "warnings": warnings}


def _extract_steps_from_definition(definition):
  if not isinstance(definition, dict):
    return []
  nodes = definition.get("nodes")
  if not isinstance(nodes, list):
    return []
  steps = []
  for node in nodes:
    if not isinstance(node, dict):
      continue
    node_type = node.get("node_type")
    if node_type not in {"approval", "cc", "condition", "subprocess", "parallel_start", "parallel_join"}:
      continue
    step = {
      "name": node.get("name"),
      "step_type": node_type
    }
    for key in [
      "approver_type",
      "approval_mode",
      "approval_type",
      "allow_self_approve",
      "allow_return",
      "timeout_hours",
      "field_permissions",
      "approver_groups",
      "approver_user_ids",
      "approver_roles",
      "approver_positions",
      "approver_field_key",
      "previous_step_offset",
      "subprocess_template_id",
      "condition"
    ]:
      if key in node:
        step[key] = node.get(key)
    steps.append(step)
  for idx, step in enumerate(steps, start=1):
    step["step_no"] = idx
  return steps


def _validate_workflow_form_data(schema, raw_data):
  if raw_data is None:
    raw_data = {}
  if not isinstance(raw_data, dict):
    raise ValueError("invalid_form_data")

  known_keys = {field["key"] for field in schema}
  normalized = {}

  for field in schema:
    key = field["key"]
    value = raw_data.get(key)
    if value == "":
      value = None

    if value is None:
      if field.get("required"):
        raise ValueError(f"missing_required_field:{key}")
      normalized[key] = None
      continue

    field_type = field.get("type")
    if field_type in {"text", "textarea"}:
      if not isinstance(value, str):
        value = str(value)
    elif field_type == "number":
      if isinstance(value, bool):
        raise ValueError(f"invalid_field_type:{key}")
      try:
        number_value = float(value)
      except (TypeError, ValueError):
        raise ValueError(f"invalid_field_type:{key}")
      if number_value.is_integer():
        value = int(number_value)
      else:
        value = number_value
    elif field_type == "date":
      if not isinstance(value, str) or not re.match(r"^\d{4}-\d{2}-\d{2}$", value):
        raise ValueError(f"invalid_field_type:{key}")
    elif field_type == "select":
      if not isinstance(value, str):
        value = str(value)
      options = field.get("options") or []
      if options and value not in options:
        raise ValueError(f"invalid_field_option:{key}")
    elif field_type == "boolean":
      if isinstance(value, bool):
        pass
      elif isinstance(value, str) and value.lower() in {"true", "false"}:
        value = value.lower() == "true"
      elif isinstance(value, (int, float)) and value in (0, 1):
        value = bool(value)
      else:
        raise ValueError(f"invalid_field_type:{key}")
    elif field_type == "attachment":
      if not isinstance(value, list):
        raise ValueError(f"invalid_field_type:{key}")
      normalized_files = []
      for item in value:
        if isinstance(item, str):
          file_ref = item.strip()
          if file_ref:
            normalized_files.append(file_ref)
        elif isinstance(item, dict):
          file_ref = str(item.get("url") or item.get("name") or "").strip()
          if file_ref:
            normalized_files.append(file_ref)
        else:
          raise ValueError(f"invalid_field_type:{key}")
      max_count = field.get("max_count")
      if isinstance(max_count, int) and max_count > 0 and len(normalized_files) > max_count:
        raise ValueError(f"invalid_field_max_count:{key}")
      value = normalized_files
    elif field_type == "table":
      if not isinstance(value, list):
        raise ValueError(f"invalid_field_type:{key}")
      columns = field.get("columns") if isinstance(field.get("columns"), list) else []
      if not columns:
        raise ValueError(f"invalid_field_type:{key}")
      column_map = {}
      for column in columns:
        if isinstance(column, dict) and column.get("key"):
          column_map[column["key"]] = column
      normalized_rows = []
      for row in value:
        if not isinstance(row, dict):
          raise ValueError(f"invalid_field_type:{key}")
        normalized_row = {}
        for col_key, col_schema in column_map.items():
          col_value = row.get(col_key)
          if col_value in ("", None):
            normalized_row[col_key] = None
            continue
          col_type = col_schema.get("type")
          if col_type in {"text", "textarea"}:
            if not isinstance(col_value, str):
              col_value = str(col_value)
          elif col_type == "number":
            if isinstance(col_value, bool):
              raise ValueError(f"invalid_field_type:{key}")
            try:
              number_value = float(col_value)
            except (TypeError, ValueError):
              raise ValueError(f"invalid_field_type:{key}")
            col_value = int(number_value) if number_value.is_integer() else number_value
          elif col_type == "date":
            if not isinstance(col_value, str) or not re.match(r"^\d{4}-\d{2}-\d{2}$", col_value):
              raise ValueError(f"invalid_field_type:{key}")
          elif col_type == "select":
            if not isinstance(col_value, str):
              col_value = str(col_value)
            options = col_schema.get("options") if isinstance(col_schema.get("options"), list) else []
            if options and col_value not in options:
              raise ValueError(f"invalid_field_option:{key}")
          elif col_type == "boolean":
            if isinstance(col_value, bool):
              pass
            elif isinstance(col_value, str) and col_value.lower() in {"true", "false"}:
              col_value = col_value.lower() == "true"
            elif isinstance(col_value, (int, float)) and col_value in (0, 1):
              col_value = bool(col_value)
            else:
              raise ValueError(f"invalid_field_type:{key}")
          normalized_row[col_key] = col_value
        unknown_columns = [col for col in row.keys() if col not in column_map and row.get(col) not in (None, "")]
        if unknown_columns:
          raise ValueError(f"invalid_field_type:{key}")
        normalized_rows.append(normalized_row)
      value = normalized_rows

    normalized[key] = value

  unknown_keys = [key for key in raw_data.keys() if key not in known_keys and raw_data.get(key) not in (None, "")]
  if unknown_keys:
    raise ValueError("unknown_form_fields")

  return normalized


def _serialize_form_template(row, include_schema=True):
  if not row:
    return None
  schema = _safe_json_load(row.get("schema_json"))
  if not isinstance(schema, list):
    schema = []
  data = {
    "id": row.get("id"),
    "name": row.get("name"),
    "description": row.get("description"),
    "company_id": row.get("company_id"),
    "company_name": row.get("company_name"),
    "status": row.get("status"),
    "created_by": row.get("created_by"),
    "created_by_name": row.get("created_by_name"),
    "updated_by": row.get("updated_by"),
    "updated_by_name": row.get("updated_by_name"),
    "created_at": row.get("created_at"),
    "updated_at": row.get("updated_at"),
    "field_count": len(schema)
  }
  if include_schema:
    data["schema"] = schema
  return data


def _serialize_process_template(row, include_steps=True, include_form_schema=False):
  if not row:
    return None
  raw_definition = _safe_json_load(row.get("steps_json"))
  try:
    definition = _normalize_workflow_definition(raw_definition)
  except ValueError:
    definition = _steps_to_graph_definition([])
  steps = _extract_steps_from_definition(definition)
  form_schema = _safe_json_load(row.get("form_schema_json"))
  if not isinstance(form_schema, list):
    form_schema = []

  data = {
    "id": row.get("id"),
    "name": row.get("name"),
    "description": row.get("description"),
    "company_id": row.get("company_id"),
    "company_name": row.get("company_name"),
    "status": row.get("status"),
    "form_template_id": row.get("form_template_id"),
    "form_template_name": row.get("form_template_name"),
    "created_by": row.get("created_by"),
    "created_by_name": row.get("created_by_name"),
    "updated_by": row.get("updated_by"),
    "updated_by_name": row.get("updated_by_name"),
    "created_at": row.get("created_at"),
    "updated_at": row.get("updated_at"),
    "step_count": len(steps),
    "current_version": row.get("current_version"),
    "published_version": row.get("published_version")
  }
  if include_steps:
    data["steps"] = steps
  data["definition"] = definition
  if include_form_schema:
    data["form_schema"] = form_schema
  return data


def _serialize_process_template_version(row, include_definition=False, include_form_schema=False):
  if not row:
    return None
  raw_definition = _safe_json_load(row.get("definition_json"))
  try:
    definition = _normalize_workflow_definition(raw_definition)
  except ValueError:
    definition = _steps_to_graph_definition([])
  steps = _extract_steps_from_definition(definition)
  form_schema = _safe_json_load(row.get("form_schema_json"))
  if not isinstance(form_schema, list):
    form_schema = []

  data = {
    "id": row.get("id"),
    "process_template_id": row.get("process_template_id"),
    "version_no": row.get("version_no"),
    "form_template_id": row.get("form_template_id"),
    "form_template_name": row.get("form_template_name"),
    "status": row.get("status"),
    "published_at": row.get("published_at"),
    "created_by": row.get("created_by"),
    "created_by_name": row.get("created_by_name"),
    "updated_by": row.get("updated_by"),
    "updated_by_name": row.get("updated_by_name"),
    "created_at": row.get("created_at"),
    "updated_at": row.get("updated_at"),
    "step_count": len(steps),
    "steps": steps
  }
  if include_definition:
    data["definition"] = definition
  if include_form_schema:
    data["form_schema"] = form_schema
  return data


def _serialize_approval_task(row):
  return {
    "id": row.get("id"),
    "instance_id": row.get("instance_id"),
    "step_no": row.get("step_no"),
    "step_name": row.get("step_name"),
    "approval_mode": row.get("approval_mode"),
    "approver_id": row.get("approver_id"),
    "approver_name": row.get("approver_name"),
    "status": row.get("status"),
    "decision": row.get("decision"),
    "comment": row.get("comment"),
    "acted_at": row.get("acted_at"),
    "created_at": row.get("created_at"),
    "updated_at": row.get("updated_at")
  }


def _serialize_approval_instance(row, include_payload=False):
  if not row:
    return None
  data = {
    "id": row.get("id"),
    "process_template_id": row.get("process_template_id"),
    "form_template_id": row.get("form_template_id"),
    "process_name": row.get("process_name"),
    "title": row.get("title"),
    "company_id": row.get("company_id"),
    "company_name": row.get("company_name"),
    "applicant_id": row.get("applicant_id"),
    "applicant_name": row.get("applicant_name"),
    "status": row.get("status"),
    "current_step": row.get("current_step"),
    "total_steps": row.get("total_steps"),
    "current_step_name": row.get("current_step_name"),
    "current_node_id": row.get("current_node_id"),
    "pending_action": bool(row.get("pending_action")),
    "created_at": row.get("created_at"),
    "updated_at": row.get("updated_at"),
    "finished_at": row.get("finished_at")
  }
  if include_payload:
    process_snapshot = _safe_json_load(row.get("process_snapshot_json"))
    form_schema = _safe_json_load(row.get("form_schema_json"))
    form_data = _safe_json_load(row.get("form_data_json"))
    data["process_snapshot"] = process_snapshot if isinstance(process_snapshot, dict) else {}
    data["form_schema"] = form_schema if isinstance(form_schema, list) else []
    data["form_data"] = form_data if isinstance(form_data, dict) else {}
  return data


def _is_empty_value(value):
  if value is None:
    return True
  if isinstance(value, str):
    return value.strip() == ""
  if isinstance(value, (list, tuple, set, dict)):
    return len(value) == 0
  return False


def _evaluate_condition_rule(form_data, rule):
  field = rule.get("field")
  operator = rule.get("operator")
  expected = rule.get("value")
  actual = form_data.get(field)

  if operator == "is_empty":
    return _is_empty_value(actual)
  if operator == "not_empty":
    return not _is_empty_value(actual)
  if operator == "is_true":
    return actual is True
  if operator == "is_false":
    return actual is False
  if operator == "contains":
    if isinstance(actual, str):
      return str(expected or "") in actual
    if isinstance(actual, (list, tuple, set)):
      return expected in actual
    return False
  if operator == "in":
    if isinstance(expected, (list, tuple, set)):
      return actual in expected
    return actual == expected
  if operator == "not_in":
    if isinstance(expected, (list, tuple, set)):
      return actual not in expected
    return actual != expected

  if operator in {"gt", "gte", "lt", "lte"}:
    try:
      actual_num = float(actual)
      expected_num = float(expected)
    except (TypeError, ValueError):
      return False
    if operator == "gt":
      return actual_num > expected_num
    if operator == "gte":
      return actual_num >= expected_num
    if operator == "lt":
      return actual_num < expected_num
    return actual_num <= expected_num

  if operator == "neq":
    return actual != expected
  return actual == expected


def _is_safe_condition_expression(expression):
  try:
    tree = ast.parse(expression, mode="eval")
  except SyntaxError:
    return False

  allowed_node_types = (
    ast.Expression,
    ast.BoolOp,
    ast.BinOp,
    ast.UnaryOp,
    ast.Compare,
    ast.Call,
    ast.Name,
    ast.Load,
    ast.Constant,
    ast.List,
    ast.Tuple,
    ast.Dict,
    ast.Subscript,
    ast.Slice,
    ast.Index,
    ast.And,
    ast.Or,
    ast.Not,
    ast.In,
    ast.NotIn,
    ast.Eq,
    ast.NotEq,
    ast.Gt,
    ast.GtE,
    ast.Lt,
    ast.LtE,
    ast.Add,
    ast.Sub,
    ast.Mult,
    ast.Div,
    ast.FloorDiv,
    ast.Mod,
    ast.USub,
    ast.UAdd
  )
  allowed_calls = {
    "len",
    "int",
    "float",
    "str",
    "bool",
    "abs",
    "contains",
    "startswith",
    "endswith",
    "lower",
    "upper",
    "empty",
    "any",
    "all",
    "min",
    "max",
    "round",
    "field"
  }

  for node in ast.walk(tree):
    if not isinstance(node, allowed_node_types):
      return False
    if isinstance(node, ast.Name) and node.id.startswith("__"):
      return False
    if isinstance(node, ast.Call):
      if not isinstance(node.func, ast.Name):
        return False
      if node.func.id not in allowed_calls:
        return False
  return True


def _evaluate_condition_expression(form_data, expression):
  if not isinstance(expression, str):
    return False
  text = expression.strip()
  if not text:
    return False
  if not _is_safe_condition_expression(text):
    return False

  source = dict(form_data or {})
  context = dict(source)
  context.update(
    {
      "field": lambda key: source.get(str(key)),
      "contains": lambda container, item: item in container if container is not None else False,
      "startswith": lambda value, prefix: str(value).startswith(str(prefix)),
      "endswith": lambda value, suffix: str(value).endswith(str(suffix)),
      "lower": lambda value: str(value).lower(),
      "upper": lambda value: str(value).upper(),
      "empty": _is_empty_value,
      "len": len,
      "int": int,
      "float": float,
      "str": str,
      "bool": bool,
      "abs": abs,
      "any": any,
      "all": all,
      "min": min,
      "max": max,
      "round": round
    }
  )

  try:
    code = compile(ast.parse(text, mode="eval"), "<workflow_condition>", "eval")
    return bool(eval(code, {"__builtins__": {}}, context))
  except Exception:
    return False


def _evaluate_condition_definition(form_data, condition):
  if not condition:
    return True
  if not isinstance(condition, dict):
    return False
  results = []
  expression = condition.get("expression")
  if expression not in (None, ""):
    results.append(_evaluate_condition_expression(form_data, expression))
  rules = condition.get("rules")
  if isinstance(rules, list) and rules:
    results.extend([_evaluate_condition_rule(form_data, rule) for rule in rules])
  if not results:
    return False
  logic = str(condition.get("logic") or "and").lower()
  if logic == "or":
    return any(results)
  return all(results)


def _evaluate_step_condition(form_data, step):
  return _evaluate_condition_definition(form_data, step.get("condition"))


def _load_instance_form_data(instance):
  form_data = _safe_json_load(instance.get("form_data_json"))
  if isinstance(form_data, dict):
    return form_data
  return {}


def _load_instance_definition(instance):
  snapshot = _safe_json_load(instance.get("process_snapshot_json"))

  raw_definition = None
  if isinstance(snapshot, dict):
    if isinstance(snapshot.get("definition"), dict):
      raw_definition = snapshot.get("definition")
    elif isinstance(snapshot.get("nodes"), list):
      raw_definition = snapshot
    elif isinstance(snapshot.get("steps"), list):
      raw_definition = snapshot.get("steps")
  elif isinstance(snapshot, list):
    raw_definition = snapshot

  if raw_definition is None:
    raw_definition = []

  try:
    return _normalize_workflow_definition(raw_definition)
  except ValueError:
    return _steps_to_graph_definition([])


def _get_definition_node_by_id(definition, node_id):
  if not node_id or not isinstance(definition, dict):
    return None
  nodes = definition.get("nodes")
  if not isinstance(nodes, list):
    return None
  for node in nodes:
    if isinstance(node, dict) and node.get("id") == node_id:
      return node
  return None


def _get_instance_current_node(instance, definition=None):
  if definition is None:
    definition = _load_instance_definition(instance)
  current_node_id = instance.get("current_node_id")
  return _get_definition_node_by_id(definition, current_node_id)


def _build_field_permission_map(raw_permissions):
  permissions = {}
  for permission in _normalize_field_permissions(raw_permissions):
    can_view = bool(permission.get("can_view", True))
    can_edit = bool(permission.get("can_edit", False)) and can_view
    required = bool(permission.get("required", False)) and can_edit
    permissions[permission["field_key"]] = {
      "can_view": can_view,
      "can_edit": can_edit,
      "required": required
    }
  return permissions


def _build_definition_index(definition):
  nodes = definition.get("nodes") if isinstance(definition, dict) else []
  edges = definition.get("edges") if isinstance(definition, dict) else []
  if not isinstance(nodes, list):
    nodes = []
  if not isinstance(edges, list):
    edges = []

  nodes_by_id = {}
  for node in nodes:
    if isinstance(node, dict) and node.get("id"):
      nodes_by_id[node["id"]] = node

  outgoing_edges = {}
  for edge in edges:
    if not isinstance(edge, dict):
      continue
    source = edge.get("source")
    target = edge.get("target")
    if source not in nodes_by_id or target not in nodes_by_id:
      continue
    outgoing_edges.setdefault(source, []).append(edge)

  for source, items in outgoing_edges.items():
    outgoing_edges[source] = sorted(
      items,
      key=lambda edge: (int(edge.get("priority") or 9999), str(edge.get("id") or ""))
    )

  return nodes_by_id, outgoing_edges


def _resolve_step_approver_ids(
  db,
  applicant,
  template_company_id,
  step,
  form_data=None,
  instance_id=None,
  current_step=None
):
  approver_type = step.get("approver_type")
  approver_ids = []

  approver_groups = step.get("approver_groups")
  if isinstance(approver_groups, list) and approver_groups:
    matched_group = False
    merged_ids = []
    condition_form_data = form_data if isinstance(form_data, dict) else {}
    for group in approver_groups:
      if not isinstance(group, dict):
        continue
      group_condition = group.get("condition")
      if group_condition and not _evaluate_condition_definition(condition_form_data, group_condition):
        continue
      matched_group = True
      group_step = dict(step)
      group_step.pop("approver_groups", None)
      for key in [
        "approver_type",
        "approver_user_ids",
        "approver_roles",
        "approver_positions",
        "approver_field_key",
        "previous_step_offset",
        "allow_self_approve"
      ]:
        if key in group:
          group_step[key] = group.get(key)
      group_ids = _resolve_step_approver_ids(
        db,
        applicant,
        template_company_id,
        group_step,
        form_data=form_data,
        instance_id=instance_id,
        current_step=current_step
      )
      for user_id in group_ids:
        if user_id not in merged_ids:
          merged_ids.append(user_id)
    if matched_group:
      return merged_ids

  if approver_type == "user":
    candidate_ids = step.get("approver_user_ids") or []
    if not candidate_ids:
      return []
    placeholders = ", ".join(["%s"] * len(candidate_ids))
    with db.cursor() as cur:
      cur.execute(
        f"SELECT id FROM users WHERE status = 'active' AND id IN ({placeholders})",
        tuple(candidate_ids)
      )
      approver_ids = [row["id"] for row in cur.fetchall()]

  elif approver_type == "role":
    roles = step.get("approver_roles") or []
    if not roles:
      return []
    placeholders = ", ".join(["%s"] * len(roles))
    params = list(roles)
    scope_condition = " AND r.company_id IS NULL"
    user_scope_condition = ""
    if template_company_id:
      scope_condition = " AND (r.company_id = %s OR r.company_id IS NULL)"
      params.append(template_company_id)
      user_scope_condition = " AND (u.company_id = %s OR u.role = 'group_admin')"
      params.append(template_company_id)
    with db.cursor() as cur:
      cur.execute(
        f"SELECT DISTINCT u.id FROM users u "
        f"JOIN user_org_roles uor ON uor.user_id = u.id "
        f"JOIN org_roles r ON r.id = uor.role_id "
        f"WHERE u.status = 'active' AND r.status = 'active' "
        f"AND r.name IN ({placeholders})"
        f"{scope_condition}{user_scope_condition} "
        f"ORDER BY u.id ASC",
        tuple(params)
      )
      approver_ids = [row["id"] for row in cur.fetchall()]

  elif approver_type in {"manager", "department_manager"}:
    applicant_company_id = applicant.get("company_id")
    with db.cursor() as cur:
      if applicant_company_id:
        cur.execute(
          "SELECT id FROM users "
          "WHERE status = 'active' AND role = 'subsidiary_admin' AND company_id = %s "
          "ORDER BY id ASC",
          (applicant_company_id,)
        )
        approver_ids = [row["id"] for row in cur.fetchall()]
      if not approver_ids:
        cur.execute(
          "SELECT id FROM users WHERE status = 'active' AND role = 'group_admin' ORDER BY id ASC"
        )
        approver_ids = [row["id"] for row in cur.fetchall()]

  elif approver_type == "position":
    positions = step.get("approver_positions") or step.get("approver_roles") or []
    if not positions:
      return []
    placeholders = ", ".join(["%s"] * len(positions))
    params = list(positions)
    scope_condition = " AND p.company_id IS NULL"
    user_scope_condition = ""
    if template_company_id:
      scope_condition = " AND (p.company_id = %s OR p.company_id IS NULL)"
      params.append(template_company_id)
      user_scope_condition = " AND (u.company_id = %s OR u.role = 'group_admin')"
      params.append(template_company_id)
    with db.cursor() as cur:
      cur.execute(
        f"SELECT DISTINCT u.id FROM users u "
        f"JOIN user_org_positions uop ON uop.user_id = u.id "
        f"JOIN org_positions p ON p.id = uop.position_id "
        f"WHERE u.status = 'active' AND p.status = 'active' "
        f"AND p.name IN ({placeholders})"
        f"{scope_condition}{user_scope_condition} "
        f"ORDER BY u.id ASC",
        tuple(params)
      )
      approver_ids = [row["id"] for row in cur.fetchall()]

  elif approver_type == "applicant_select":
    if not isinstance(form_data, dict):
      form_data = {}
    field_key = str(step.get("approver_field_key") or "").strip()
    if not field_key:
      return []
    raw_selected = form_data.get(field_key)
    selected_ids = []
    if isinstance(raw_selected, list):
      candidates = raw_selected
    elif isinstance(raw_selected, str) and "," in raw_selected:
      candidates = [item.strip() for item in raw_selected.split(",")]
    else:
      candidates = [raw_selected]
    for item in candidates:
      try:
        user_id = int(item)
      except (TypeError, ValueError):
        continue
      if user_id > 0 and user_id not in selected_ids:
        selected_ids.append(user_id)
    if not selected_ids:
      return []
    placeholders = ", ".join(["%s"] * len(selected_ids))
    with db.cursor() as cur:
      cur.execute(
        f"SELECT id FROM users WHERE status = 'active' AND id IN ({placeholders})",
        tuple(selected_ids)
      )
      approver_ids = [row["id"] for row in cur.fetchall()]

  elif approver_type == "previous_handler":
    if not instance_id:
      return []
    if not current_step:
      return []
    try:
      previous_step_offset = int(step.get("previous_step_offset") or 1)
    except (TypeError, ValueError):
      previous_step_offset = 1
    target_step = int(current_step) - previous_step_offset
    if target_step <= 0:
      return []
    with db.cursor() as cur:
      cur.execute(
        "SELECT approver_id FROM approval_instance_tasks "
        "WHERE instance_id = %s AND step_no = %s AND decision IS NOT NULL "
        "ORDER BY id ASC",
        (instance_id, target_step)
      )
      approver_ids = [row["approver_id"] for row in cur.fetchall()]

  unique_ids = []
  for user_id in approver_ids:
    if user_id not in unique_ids:
      unique_ids.append(user_id)

  if not step.get("allow_self_approve"):
    unique_ids = [user_id for user_id in unique_ids if user_id != applicant.get("id")]
  return unique_ids


def _create_step_tasks(
  db,
  instance_id,
  step_no,
  step_name,
  approval_mode,
  approver_ids,
  task_status="pending",
  task_comment=None
):
  if not approver_ids:
    return 0
  rows = [
    (
      instance_id,
      step_no,
      step_name,
      approval_mode,
      approver_id,
      task_status,
      task_comment
    )
    for approver_id in approver_ids
  ]
  with db.cursor() as cur:
    cur.executemany(
      "INSERT INTO approval_instance_tasks "
      "(instance_id, step_no, step_name, approval_mode, approver_id, status, comment) "
      "VALUES (%s, %s, %s, %s, %s, %s, %s)",
      rows
    )
  return len(rows)


def _create_step_tasks_sequential(
  db,
  instance_id,
  step_no,
  step_name,
  approver_ids
):
  if not approver_ids:
    return 0
  rows = []
  for idx, approver_id in enumerate(approver_ids):
    rows.append(
      (
        instance_id,
        step_no,
        step_name,
        "all",
        approver_id,
        "pending" if idx == 0 else "waiting",
        None if idx == 0 else "sequential_waiting"
      )
    )
  with db.cursor() as cur:
    cur.executemany(
      "INSERT INTO approval_instance_tasks "
      "(instance_id, step_no, step_name, approval_mode, approver_id, status, comment) "
      "VALUES (%s, %s, %s, %s, %s, %s, %s)",
      rows
    )
  return len(rows)


def _mark_instance_finished(db, instance_id, status):
  with db.cursor() as cur:
    cur.execute(
      "UPDATE approval_instances "
      "SET status = %s, finished_at = CURRENT_TIMESTAMP "
      "WHERE id = %s",
      (status, instance_id)
    )


def _route_instance_forward(db, instance, start_node_id):
  instance_id = instance["id"]
  definition = _load_instance_definition(instance)
  form_data = _load_instance_form_data(instance)
  applicant = {"id": instance.get("applicant_id"), "company_id": instance.get("company_id")}
  nodes_by_id, outgoing_edges = _build_definition_index(definition)

  if not nodes_by_id:
    _mark_instance_finished(db, instance_id, "approved")
    return False

  current_node_id = start_node_id or definition.get("start_node_id")
  if current_node_id not in nodes_by_id:
    current_node_id = definition.get("start_node_id")
  if current_node_id not in nodes_by_id:
    _mark_instance_finished(db, instance_id, "rejected")
    return False

  max_hops = max(20, len(nodes_by_id) * 3)

  for _ in range(max_hops):
    candidate_edges = []
    default_edges = []
    for edge in outgoing_edges.get(current_node_id, []):
      condition = edge.get("condition")
      if condition and _evaluate_condition_definition(form_data, condition):
        candidate_edges.append(edge)
      elif _is_default_branch_edge(edge):
        default_edges.append(edge)

    next_edge = candidate_edges[0] if candidate_edges else (default_edges[0] if default_edges else None)
    if not next_edge:
      _mark_instance_finished(db, instance_id, "approved")
      return False

    next_node_id = next_edge.get("target")
    next_node = nodes_by_id.get(next_node_id)
    if not next_node:
      _mark_instance_finished(db, instance_id, "rejected")
      return False

    if not _evaluate_step_condition(form_data, next_node):
      current_node_id = next_node_id
      continue

    node_type = str(next_node.get("node_type") or "").lower()
    if node_type in {"start", "condition", "parallel_start", "parallel_join"}:
      current_node_id = next_node_id
      continue

    if node_type == "end":
      with db.cursor() as cur:
        cur.execute(
          "UPDATE approval_instances SET current_node_id = %s, current_step_name = %s WHERE id = %s",
          (next_node_id, next_node.get("name"), instance_id)
        )
      _mark_instance_finished(db, instance_id, "approved")
      return False

    step_no = int(instance.get("current_step") or 0) + 1

    if node_type == "subprocess":
      with db.cursor() as cur:
        cur.execute(
          "UPDATE approval_instances "
          "SET current_step = %s, current_step_name = %s, current_node_id = %s "
          "WHERE id = %s",
          (step_no, next_node.get("name"), next_node_id, instance_id)
        )
      _log_instance_event(
        db,
        instance_id,
        applicant.get("id"),
        "subprocess_auto",
        comment="subprocess_not_enabled_yet",
        detail={"subprocess_template_id": next_node.get("subprocess_template_id"), "node_id": next_node_id}
      )
      instance["current_step"] = step_no
      instance["current_step_name"] = next_node.get("name")
      instance["current_node_id"] = next_node_id
      current_node_id = next_node_id
      continue

    approver_ids = _resolve_step_approver_ids(
      db,
      applicant,
      instance.get("company_id"),
      next_node,
      form_data=form_data,
      instance_id=instance_id,
      current_step=step_no
    )

    if node_type == "cc":
      if approver_ids:
        _create_step_tasks(
          db,
          instance_id,
          step_no,
          next_node.get("name"),
          next_node.get("approval_mode") or "any",
          approver_ids,
          task_status="skipped",
          task_comment="cc_auto_record"
        )
      with db.cursor() as cur:
        cur.execute(
          "UPDATE approval_instances "
          "SET current_step = %s, current_step_name = %s, current_node_id = %s "
          "WHERE id = %s",
          (step_no, next_node.get("name"), next_node_id, instance_id)
        )
      instance["current_step"] = step_no
      instance["current_step_name"] = next_node.get("name")
      instance["current_node_id"] = next_node_id
      current_node_id = next_node_id
      continue

    if not approver_ids:
      _mark_instance_finished(db, instance_id, "rejected")
      return False

    approval_type = str(next_node.get("approval_type") or "").strip().lower()
    if approval_type == "sequential" and len(approver_ids) > 1:
      _create_step_tasks_sequential(
        db,
        instance_id,
        step_no,
        next_node.get("name"),
        approver_ids
      )
    else:
      _create_step_tasks(
        db,
        instance_id,
        step_no,
        next_node.get("name"),
        next_node.get("approval_mode") or "any",
        approver_ids
      )
    with db.cursor() as cur:
      cur.execute(
        "UPDATE approval_instances "
        "SET current_step = %s, current_step_name = %s, current_node_id = %s, status = 'pending', finished_at = NULL "
        "WHERE id = %s",
        (step_no, next_node.get("name"), next_node_id, instance_id)
      )
    return True

  _mark_instance_finished(db, instance_id, "rejected")
  return False


def _advance_approval_instance(db, instance):
  instance_id = instance["id"]
  current_step = instance["current_step"]
  definition = _load_instance_definition(instance)
  current_node = _get_instance_current_node(instance, definition=definition)
  approval_type = str((current_node or {}).get("approval_type") or "").strip().lower()

  with db.cursor() as cur:
    cur.execute(
      "SELECT id, status, approval_mode, step_no FROM approval_instance_tasks "
      "WHERE instance_id = %s AND step_no = %s ORDER BY id ASC",
      (instance_id, current_step)
    )
    step_tasks = cur.fetchall()

  if not step_tasks:
    return

  approval_mode = step_tasks[0].get("approval_mode") or "any"
  statuses = [task.get("status") for task in step_tasks]
  actionable_statuses = [status for status in statuses if status != "skipped"]
  approved_count = len([status for status in actionable_statuses if status == "approved"])
  pending_count = len([status for status in actionable_statuses if status == "pending"])
  waiting_tasks = [task for task in step_tasks if task.get("status") == "waiting"]

  step_done = False
  if approval_type == "sequential":
    if pending_count > 0:
      return
    if waiting_tasks:
      next_waiting_task = waiting_tasks[0]
      with db.cursor() as cur:
        cur.execute(
          "UPDATE approval_instance_tasks SET status = 'pending', comment = NULL "
          "WHERE id = %s AND status = 'waiting'",
          (next_waiting_task.get("id"),)
        )
      return
    step_done = approved_count > 0 and all(status == "approved" for status in actionable_statuses)
  elif approval_mode == "any":
    step_done = approved_count > 0
  else:
    active_statuses = [status for status in actionable_statuses if status != "waiting"]
    step_done = (
      len(active_statuses) > 0 and
      pending_count == 0 and
      not waiting_tasks and
      all(status == "approved" for status in active_statuses)
    )

  if not step_done:
    return

  with db.cursor() as cur:
    cur.execute(
      "UPDATE approval_instance_tasks SET status = 'skipped' "
      "WHERE instance_id = %s AND step_no = %s AND status IN ('pending', 'waiting')",
      (instance_id, current_step)
    )
  start_node_id = instance.get("current_node_id")
  if not start_node_id and current_step:
    start_node_id = f"step_{current_step}"
  _route_instance_forward(db, instance, start_node_id)


def _can_access_instance(user, instance_row):
  if is_group_admin(user):
    return True
  if instance_row.get("applicant_id") == user.get("id"):
    return True
  return bool(instance_row.get("has_task_access"))


def _get_instance_detail(db, instance_id, user):
  with db.cursor() as cur:
    cur.execute(
      "SELECT ai.*, c.name AS company_name, au.name AS applicant_name, "
      "EXISTS(SELECT 1 FROM approval_instance_tasks ait WHERE ait.instance_id = ai.id AND ait.approver_id = %s) AS has_task_access, "
      "EXISTS(SELECT 1 FROM approval_instance_tasks ait WHERE ait.instance_id = ai.id AND ait.approver_id = %s AND ait.status = 'pending') AS pending_action "
      "FROM approval_instances ai "
      "LEFT JOIN companies c ON c.id = ai.company_id "
      "LEFT JOIN users au ON au.id = ai.applicant_id "
      "WHERE ai.id = %s",
      (user.get("id"), user.get("id"), instance_id)
    )
    instance_row = cur.fetchone()
    if not instance_row:
      return None

    if not _can_access_instance(user, instance_row):
      return None

    cur.execute(
      "SELECT ait.*, u.name AS approver_name "
      "FROM approval_instance_tasks ait "
      "LEFT JOIN users u ON u.id = ait.approver_id "
      "WHERE ait.instance_id = %s "
      "ORDER BY ait.step_no ASC, ait.id ASC",
      (instance_id,)
    )
    task_rows = cur.fetchall()
    cur.execute(
      "SELECT aie.*, u.name AS user_name "
      "FROM approval_instance_events aie "
      "LEFT JOIN users u ON u.id = aie.user_id "
      "WHERE aie.instance_id = %s "
      "ORDER BY aie.id DESC",
      (instance_id,)
    )
    event_rows = cur.fetchall()

  data = _serialize_approval_instance(instance_row, include_payload=True)
  data["tasks"] = [_serialize_approval_task(row) for row in task_rows]
  data["events"] = [
    {
      "id": row.get("id"),
      "task_id": row.get("task_id"),
      "user_id": row.get("user_id"),
      "user_name": row.get("user_name"),
      "action": row.get("action"),
      "comment": row.get("comment"),
      "detail": _safe_json_load(row.get("detail_json")),
      "created_at": row.get("created_at")
    }
    for row in event_rows
  ]
  definition = _load_instance_definition(instance_row)
  current_node = _get_instance_current_node(instance_row, definition=definition)
  permission_map = _build_field_permission_map((current_node or {}).get("field_permissions"))
  if permission_map:
    current_step = instance_row.get("current_step")
    current_user_id = user.get("id")
    pending_current_task = any(
      row.get("step_no") == current_step and row.get("approver_id") == current_user_id and row.get("status") == "pending"
      for row in task_rows
    )
    form_schema = data.get("form_schema") if isinstance(data.get("form_schema"), list) else []
    form_data = data.get("form_data") if isinstance(data.get("form_data"), dict) else {}
    effective_map = {}
    filtered_schema = []
    filtered_data = {}
    for field in form_schema:
      if not isinstance(field, dict):
        continue
      field_key = str(field.get("key") or "").strip()
      if not field_key:
        continue
      permission = permission_map.get(
        field_key,
        {
          "can_view": True,
          "can_edit": False,
          "required": False
        }
      )
      effective_map[field_key] = permission
      if pending_current_task and not permission.get("can_view"):
        continue
      field_copy = dict(field)
      if pending_current_task:
        field_copy["can_edit"] = bool(permission.get("can_edit"))
        field_copy["required"] = bool(permission.get("required")) or bool(field_copy.get("required"))
      filtered_schema.append(field_copy)
      if field_key in form_data and (not pending_current_task or permission.get("can_view")):
        filtered_data[field_key] = form_data.get(field_key)

    if pending_current_task:
      data["form_schema"] = filtered_schema
      data["form_data"] = filtered_data
    data["field_permissions"] = effective_map
  return data


def _open_db_connection():
  last_error = None
  for attempt in range(DB_CONNECT_RETRIES):
    try:
      return pymysql.connect(
        host=os.getenv("DB_HOST"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME"),
        port=int(os.getenv("DB_PORT", "3306")),
        autocommit=True,
        connect_timeout=6,
        read_timeout=10,
        write_timeout=10,
        cursorclass=DictCursor
      )
    except OperationalError as err:
      last_error = err
      if attempt < DB_CONNECT_RETRIES - 1:
        time.sleep(DB_RETRY_DELAY_SECONDS * (attempt + 1))
        continue
      raise
  if last_error:
    raise last_error
  raise OperationalError("db_connect_failed")


def get_db():
  if "db" not in g:
    g.db = _open_db_connection()
    _ensure_contact_role_column(g.db)
    _ensure_workflow_tables(g.db)
    _ensure_org_dimension_tables(g.db)
    _ensure_host_pool_tables(g.db)
  else:
    try:
      g.db.ping(reconnect=True)
    except OperationalError:
      g.db = _open_db_connection()
      _ensure_contact_role_column(g.db)
      _ensure_workflow_tables(g.db)
      _ensure_org_dimension_tables(g.db)
      _ensure_host_pool_tables(g.db)
  return g.db


def close_db(_error=None):
  db = g.pop("db", None)
  if db:
    db.close()


def is_group_admin(user):
  return user and user.get("role") == "group_admin"


def is_sub_admin(user):
  return user and user.get("role") == "subsidiary_admin"


def ensure_group_admin():
  if not is_group_admin(g.user):
    return jsonify({"error": "forbidden"}), 403
  return None


def ensure_org_access():
  if not (is_group_admin(g.user) or is_sub_admin(g.user)):
    return jsonify({"error": "forbidden"}), 403
  return None


def require_user(fn):
  @wraps(fn)
  def wrapper(*args, **kwargs):
    if request.method == "OPTIONS":
      return "", 204

    raw_user_id = request.headers.get("x-user-id")
    try:
      user_id = int(raw_user_id)
    except (TypeError, ValueError):
      return jsonify({"error": "missing_x_user_id"}), 401

    db = get_db()
    with db.cursor() as cur:
      cur.execute(
        "SELECT id, name, role, company_id FROM users WHERE id = %s AND status = 'active'",
        (user_id,)
      )
      user = cur.fetchone()

    if not user:
      return jsonify({"error": "invalid_user"}), 401

    g.user = user
    return fn(*args, **kwargs)

  return wrapper


app = Flask(__name__)
app.teardown_appcontext(close_db)


@app.errorhandler(OperationalError)
def handle_db_operational_error(_err):
  return jsonify({"error": "db_unavailable"}), 503


@app.errorhandler(Exception)
def handle_unexpected_exception(err):
  if isinstance(err, HTTPException):
    return err
  return jsonify({"error": "internal_server_error"}), 500


@app.after_request
def add_cors_headers(response):
  response.headers["Access-Control-Allow-Origin"] = os.getenv("CORS_ORIGIN", "*")
  response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-User-Id"
  response.headers["Access-Control-Allow-Methods"] = "GET,POST,PATCH,PUT,DELETE,OPTIONS"
  return response


@app.route("/health")
def health():
  return jsonify({"status": "ok"})


@app.route("/me", methods=["GET"])
@require_user
def get_me():
  return jsonify({"data": g.user})


@app.route("/login", methods=["POST"])
def login():
  body = request.get_json(silent=True) or {}
  username = body.get("username")
  password = body.get("password")

  if not username or not password:
    return jsonify({"error": "missing_credentials"}), 400

  db = get_db()
  admin_username = os.getenv("ADMIN_USERNAME", "admin-pico")
  admin_password = os.getenv("ADMIN_PASSWORD", "pico@2026")

  if username == admin_username and password == admin_password:
    with db.cursor() as cur:
      cur.execute(
        "SELECT id, name, role, company_id, status FROM users WHERE name = %s AND role = 'group_admin' LIMIT 1",
        (username,)
      )
      user = cur.fetchone()

      if not user:
        cur.execute(
          "INSERT INTO users (name, role, status) VALUES (%s, 'group_admin', 'active')",
          (username,)
        )
        user_id = cur.lastrowid
        cur.execute(
          "SELECT id, name, role, company_id, status FROM users WHERE id = %s",
          (user_id,)
        )
        user = cur.fetchone()

    if not user or user.get("status") != "active":
      return jsonify({"error": "user_inactive"}), 403

    return jsonify({"data": user})

  with db.cursor() as cur:
    cur.execute(
      "SELECT id, name, role, company_id, status, password_hash "
      "FROM users WHERE name = %s OR email = %s LIMIT 1",
      (username, username)
    )
    user = cur.fetchone()

  if not user:
    return jsonify({"error": "invalid_credentials"}), 401
  if user.get("status") != "active":
    return jsonify({"error": "user_inactive"}), 403

  password_hash = user.get("password_hash")
  if password_hash:
    if not check_password_hash(password_hash, password):
      return jsonify({"error": "invalid_credentials"}), 401
  else:
    if password != DEFAULT_USER_PASSWORD:
      return jsonify({"error": "invalid_credentials"}), 401
    hashed = hash_password(DEFAULT_USER_PASSWORD)
    with db.cursor() as cur:
      cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (hashed, user["id"]))

  user.pop("password_hash", None)
  return jsonify({"data": user})


@app.route("/companies", methods=["GET"])
@require_user
def list_companies():
  guard = ensure_org_access()
  if guard:
    return guard

  filters = []
  params = []

  if is_sub_admin(g.user):
    if not g.user.get("company_id"):
      return jsonify({"error": "user_missing_company"}), 400
    filters.append("id = %s")
    params.append(g.user["company_id"])

  parent_id = request.args.get("parent_id", type=int)
  if parent_id is not None:
    if parent_id == 0:
      filters.append("parent_id IS NULL")
    else:
      filters.append("parent_id = %s")
      params.append(parent_id)

  status = request.args.get("status")
  if status:
    if status not in COMPANY_STATUSES:
      return jsonify({"error": "invalid_status"}), 400
    filters.append("status = %s")
    params.append(status)

  name = request.args.get("name")
  if name:
    filters.append("name LIKE %s")
    params.append(f"%{name}%")

  code = request.args.get("code")
  if code:
    filters.append("code = %s")
    params.append(code)

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

  db = get_db()
  with db.cursor() as cur:
    cur.execute(f"SELECT * FROM companies {where_clause} ORDER BY id ASC", params)
    rows = cur.fetchall()

  return jsonify({"data": rows})


@app.route("/companies", methods=["POST"])
@require_user
def create_company():
  guard = ensure_group_admin()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  name = body.get("name")
  code = body.get("code") or None
  parent_id = body.get("parent_id")
  if parent_id == 0:
    parent_id = None
  status = body.get("status", "active")

  if not name:
    return jsonify({"error": "missing_name"}), 400
  if status not in COMPANY_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  if parent_id:
    db = get_db()
    with db.cursor() as cur:
      cur.execute("SELECT id FROM companies WHERE id = %s", (parent_id,))
      if not cur.fetchone():
        return jsonify({"error": "invalid_parent"}), 400

  db = get_db()
  try:
    with db.cursor() as cur:
      cur.execute(
        "INSERT INTO companies (name, code, parent_id, status) VALUES (%s, %s, %s, %s)",
        (name, code, parent_id, status)
      )
      company_id = cur.lastrowid
      cur.execute("SELECT * FROM companies WHERE id = %s", (company_id,))
      created = cur.fetchone()
  except IntegrityError:
    return jsonify({"error": "company_code_exists"}), 409

  return jsonify({"data": created}), 201


@app.route("/companies/<int:company_id>", methods=["PATCH"])
@require_user
def update_company(company_id):
  guard = ensure_group_admin()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  allowed = {"name", "code", "parent_id", "status"}

  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM companies WHERE id = %s", (company_id,))
    existing = cur.fetchone()

  if not existing:
    return jsonify({"error": "not_found"}), 404

  updates = []
  params = []

  for key in allowed:
    if key not in body:
      continue
    value = body.get(key)
    if key == "status":
      if value not in COMPANY_STATUSES:
        return jsonify({"error": "invalid_status"}), 400
    if key == "parent_id":
      if value == 0:
        value = None
      if value == company_id:
        return jsonify({"error": "invalid_parent"}), 400
      if value:
        with db.cursor() as cur:
          cur.execute("SELECT id FROM companies WHERE id = %s", (value,))
          if not cur.fetchone():
            return jsonify({"error": "invalid_parent"}), 400
    updates.append(f"{key} = %s")
    params.append(value)

  if not updates:
    return jsonify({"error": "no_updates"}), 400

  params.append(company_id)
  try:
    with db.cursor() as cur:
      cur.execute(f"UPDATE companies SET {', '.join(updates)} WHERE id = %s", params)
      cur.execute("SELECT * FROM companies WHERE id = %s", (company_id,))
      updated = cur.fetchone()
  except IntegrityError:
    return jsonify({"error": "company_code_exists"}), 409

  return jsonify({"data": updated})


@app.route("/companies/<int:company_id>", methods=["DELETE"])
@require_user
def delete_company(company_id):
  guard = ensure_group_admin()
  if guard:
    return guard

  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT id FROM companies WHERE id = %s", (company_id,))
    existing = cur.fetchone()

  if not existing:
    return jsonify({"error": "not_found"}), 404

  with db.cursor() as cur:
    cur.execute("SELECT id FROM companies WHERE parent_id = %s LIMIT 1", (company_id,))
    if cur.fetchone():
      return jsonify({"error": "company_has_children"}), 400
    cur.execute("SELECT id FROM users WHERE company_id = %s LIMIT 1", (company_id,))
    if cur.fetchone():
      return jsonify({"error": "company_has_users"}), 400
    cur.execute("SELECT id FROM opportunities WHERE company_id = %s LIMIT 1", (company_id,))
    if cur.fetchone():
      return jsonify({"error": "company_has_opportunities"}), 400
    cur.execute("DELETE FROM companies WHERE id = %s", (company_id,))

  return jsonify({"data": {"id": company_id}})


@app.route("/org/roles", methods=["GET"])
@require_user
def list_org_roles():
  guard = ensure_org_access()
  if guard:
    return guard

  filters = []
  params = []

  company_id = request.args.get("company_id", type=int)
  status = request.args.get("status")
  name = request.args.get("name")

  if status:
    if status not in ORG_DIMENSION_STATUSES:
      return jsonify({"error": "invalid_status"}), 400
    filters.append("r.status = %s")
    params.append(status)

  if name:
    filters.append("r.name LIKE %s")
    params.append(f"%{name.strip()}%")

  if is_sub_admin(g.user):
    own_company_id = g.user.get("company_id")
    if not own_company_id:
      return jsonify({"error": "user_missing_company"}), 400
    if company_id is None:
      filters.append("(r.company_id = %s OR r.company_id IS NULL)")
      params.append(own_company_id)
    elif company_id == 0:
      filters.append("r.company_id IS NULL")
    elif company_id == own_company_id:
      filters.append("r.company_id = %s")
      params.append(own_company_id)
    else:
      return jsonify({"error": "invalid_company"}), 400
  else:
    if company_id is not None:
      if company_id == 0:
        filters.append("r.company_id IS NULL")
      elif company_id > 0:
        filters.append("r.company_id = %s")
        params.append(company_id)
      else:
        return jsonify({"error": "invalid_company"}), 400

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

  db = get_db()
  with db.cursor() as cur:
    cur.execute(
      "SELECT r.id, r.name, r.code, r.company_id, c.name AS company_name, "
      "r.status, r.created_at, r.updated_at "
      "FROM org_roles r "
      "LEFT JOIN companies c ON c.id = r.company_id "
      f"{where_clause} "
      "ORDER BY (r.company_id IS NULL) DESC, r.company_id ASC, r.name ASC, r.id ASC",
      params
    )
    rows = cur.fetchall()

  return jsonify({"data": rows})


@app.route("/org/roles", methods=["POST"])
@require_user
def create_org_role():
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  name = _normalize_org_dimension_name(body.get("name"))
  code = _normalize_org_dimension_code(body.get("code"))
  status = str(body.get("status") or "active").strip().lower()
  company_id = _normalize_company_id(body.get("company_id"))
  if body.get("company_id") in (0, "0"):
    company_id = None

  if not name:
    return jsonify({"error": "missing_name"}), 400
  if status not in ORG_DIMENSION_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  db = get_db()
  if is_sub_admin(g.user):
    own_company_id = g.user.get("company_id")
    if not own_company_id:
      return jsonify({"error": "user_missing_company"}), 400
    company_id = own_company_id
  elif company_id and not _ensure_company_exists(db, company_id):
    return jsonify({"error": "invalid_company"}), 400

  with db.cursor() as cur:
    if company_id:
      cur.execute(
        "SELECT id FROM org_roles WHERE name = %s AND company_id = %s LIMIT 1",
        (name, company_id)
      )
    else:
      cur.execute(
        "SELECT id FROM org_roles WHERE name = %s AND company_id IS NULL LIMIT 1",
        (name,)
      )
    if cur.fetchone():
      return jsonify({"error": "role_name_exists"}), 409

  try:
    with db.cursor() as cur:
      cur.execute(
        "INSERT INTO org_roles (name, code, company_id, status, created_by, updated_by) "
        "VALUES (%s, %s, %s, %s, %s, %s)",
        (name, code, company_id, status, g.user["id"], g.user["id"])
      )
      role_id = cur.lastrowid
      cur.execute(
        "SELECT r.id, r.name, r.code, r.company_id, c.name AS company_name, "
        "r.status, r.created_at, r.updated_at "
        "FROM org_roles r LEFT JOIN companies c ON c.id = r.company_id "
        "WHERE r.id = %s",
        (role_id,)
      )
      created = cur.fetchone()
  except IntegrityError:
    return jsonify({"error": "role_code_exists"}), 409

  return jsonify({"data": created}), 201


@app.route("/org/roles/<int:role_id>", methods=["PATCH"])
@require_user
def update_org_role(role_id):
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM org_roles WHERE id = %s", (role_id,))
    existing = cur.fetchone()
  if not existing:
    return jsonify({"error": "not_found"}), 404
  if not _can_edit_org_dimension_scope(g.user, existing.get("company_id")):
    return jsonify({"error": "forbidden"}), 403

  next_name = _normalize_org_dimension_name(body.get("name", existing.get("name")))
  next_code = (
    _normalize_org_dimension_code(body.get("code"))
    if "code" in body
    else _normalize_org_dimension_code(existing.get("code"))
  )
  next_status = str(body.get("status", existing.get("status")) or "").strip().lower()
  next_company_id = existing.get("company_id")

  if "company_id" in body:
    requested_company_id = body.get("company_id")
    if requested_company_id in (0, "0", None, ""):
      next_company_id = None
    else:
      next_company_id = _normalize_company_id(requested_company_id)
      if not next_company_id:
        return jsonify({"error": "invalid_company"}), 400

  if is_sub_admin(g.user):
    own_company_id = g.user.get("company_id")
    if not own_company_id:
      return jsonify({"error": "user_missing_company"}), 400
    if next_company_id != own_company_id:
      return jsonify({"error": "invalid_company"}), 400
  elif next_company_id and not _ensure_company_exists(db, next_company_id):
    return jsonify({"error": "invalid_company"}), 400

  if not next_name:
    return jsonify({"error": "missing_name"}), 400
  if next_status not in ORG_DIMENSION_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  with db.cursor() as cur:
    if next_company_id:
      cur.execute(
        "SELECT id FROM org_roles WHERE name = %s AND company_id = %s AND id <> %s LIMIT 1",
        (next_name, next_company_id, role_id)
      )
    else:
      cur.execute(
        "SELECT id FROM org_roles WHERE name = %s AND company_id IS NULL AND id <> %s LIMIT 1",
        (next_name, role_id)
      )
    if cur.fetchone():
      return jsonify({"error": "role_name_exists"}), 409

  try:
    with db.cursor() as cur:
      cur.execute(
        "UPDATE org_roles "
        "SET name = %s, code = %s, company_id = %s, status = %s, updated_by = %s "
        "WHERE id = %s",
        (next_name, next_code, next_company_id, next_status, g.user["id"], role_id)
      )
      cur.execute(
        "SELECT r.id, r.name, r.code, r.company_id, c.name AS company_name, "
        "r.status, r.created_at, r.updated_at "
        "FROM org_roles r LEFT JOIN companies c ON c.id = r.company_id "
        "WHERE r.id = %s",
        (role_id,)
      )
      updated = cur.fetchone()
  except IntegrityError:
    return jsonify({"error": "role_code_exists"}), 409

  return jsonify({"data": updated})


@app.route("/org/roles/<int:role_id>", methods=["DELETE"])
@require_user
def delete_org_role(role_id):
  guard = ensure_org_access()
  if guard:
    return guard

  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT id, company_id FROM org_roles WHERE id = %s", (role_id,))
    existing = cur.fetchone()
  if not existing:
    return jsonify({"error": "not_found"}), 404
  if not _can_edit_org_dimension_scope(g.user, existing.get("company_id")):
    return jsonify({"error": "forbidden"}), 403

  with db.cursor() as cur:
    cur.execute("DELETE FROM org_roles WHERE id = %s", (role_id,))
  return jsonify({"data": {"id": role_id}})


@app.route("/org/positions", methods=["GET"])
@require_user
def list_org_positions():
  guard = ensure_org_access()
  if guard:
    return guard

  filters = []
  params = []

  company_id = request.args.get("company_id", type=int)
  status = request.args.get("status")
  name = request.args.get("name")

  if status:
    if status not in ORG_DIMENSION_STATUSES:
      return jsonify({"error": "invalid_status"}), 400
    filters.append("p.status = %s")
    params.append(status)

  if name:
    filters.append("p.name LIKE %s")
    params.append(f"%{name.strip()}%")

  if is_sub_admin(g.user):
    own_company_id = g.user.get("company_id")
    if not own_company_id:
      return jsonify({"error": "user_missing_company"}), 400
    if company_id is None:
      filters.append("(p.company_id = %s OR p.company_id IS NULL)")
      params.append(own_company_id)
    elif company_id == 0:
      filters.append("p.company_id IS NULL")
    elif company_id == own_company_id:
      filters.append("p.company_id = %s")
      params.append(own_company_id)
    else:
      return jsonify({"error": "invalid_company"}), 400
  else:
    if company_id is not None:
      if company_id == 0:
        filters.append("p.company_id IS NULL")
      elif company_id > 0:
        filters.append("p.company_id = %s")
        params.append(company_id)
      else:
        return jsonify({"error": "invalid_company"}), 400

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

  db = get_db()
  with db.cursor() as cur:
    cur.execute(
      "SELECT p.id, p.name, p.code, p.company_id, c.name AS company_name, "
      "p.status, p.created_at, p.updated_at "
      "FROM org_positions p "
      "LEFT JOIN companies c ON c.id = p.company_id "
      f"{where_clause} "
      "ORDER BY (p.company_id IS NULL) DESC, p.company_id ASC, p.name ASC, p.id ASC",
      params
    )
    rows = cur.fetchall()

  return jsonify({"data": rows})


@app.route("/org/positions", methods=["POST"])
@require_user
def create_org_position():
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  name = _normalize_org_dimension_name(body.get("name"))
  code = _normalize_org_dimension_code(body.get("code"))
  status = str(body.get("status") or "active").strip().lower()
  company_id = _normalize_company_id(body.get("company_id"))
  if body.get("company_id") in (0, "0"):
    company_id = None

  if not name:
    return jsonify({"error": "missing_name"}), 400
  if status not in ORG_DIMENSION_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  db = get_db()
  if is_sub_admin(g.user):
    own_company_id = g.user.get("company_id")
    if not own_company_id:
      return jsonify({"error": "user_missing_company"}), 400
    company_id = own_company_id
  elif company_id and not _ensure_company_exists(db, company_id):
    return jsonify({"error": "invalid_company"}), 400

  with db.cursor() as cur:
    if company_id:
      cur.execute(
        "SELECT id FROM org_positions WHERE name = %s AND company_id = %s LIMIT 1",
        (name, company_id)
      )
    else:
      cur.execute(
        "SELECT id FROM org_positions WHERE name = %s AND company_id IS NULL LIMIT 1",
        (name,)
      )
    if cur.fetchone():
      return jsonify({"error": "position_name_exists"}), 409

  try:
    with db.cursor() as cur:
      cur.execute(
        "INSERT INTO org_positions (name, code, company_id, status, created_by, updated_by) "
        "VALUES (%s, %s, %s, %s, %s, %s)",
        (name, code, company_id, status, g.user["id"], g.user["id"])
      )
      position_id = cur.lastrowid
      cur.execute(
        "SELECT p.id, p.name, p.code, p.company_id, c.name AS company_name, "
        "p.status, p.created_at, p.updated_at "
        "FROM org_positions p LEFT JOIN companies c ON c.id = p.company_id "
        "WHERE p.id = %s",
        (position_id,)
      )
      created = cur.fetchone()
  except IntegrityError:
    return jsonify({"error": "position_code_exists"}), 409

  return jsonify({"data": created}), 201


@app.route("/org/positions/<int:position_id>", methods=["PATCH"])
@require_user
def update_org_position(position_id):
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM org_positions WHERE id = %s", (position_id,))
    existing = cur.fetchone()
  if not existing:
    return jsonify({"error": "not_found"}), 404
  if not _can_edit_org_dimension_scope(g.user, existing.get("company_id")):
    return jsonify({"error": "forbidden"}), 403

  next_name = _normalize_org_dimension_name(body.get("name", existing.get("name")))
  next_code = (
    _normalize_org_dimension_code(body.get("code"))
    if "code" in body
    else _normalize_org_dimension_code(existing.get("code"))
  )
  next_status = str(body.get("status", existing.get("status")) or "").strip().lower()
  next_company_id = existing.get("company_id")

  if "company_id" in body:
    requested_company_id = body.get("company_id")
    if requested_company_id in (0, "0", None, ""):
      next_company_id = None
    else:
      next_company_id = _normalize_company_id(requested_company_id)
      if not next_company_id:
        return jsonify({"error": "invalid_company"}), 400

  if is_sub_admin(g.user):
    own_company_id = g.user.get("company_id")
    if not own_company_id:
      return jsonify({"error": "user_missing_company"}), 400
    if next_company_id != own_company_id:
      return jsonify({"error": "invalid_company"}), 400
  elif next_company_id and not _ensure_company_exists(db, next_company_id):
    return jsonify({"error": "invalid_company"}), 400

  if not next_name:
    return jsonify({"error": "missing_name"}), 400
  if next_status not in ORG_DIMENSION_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  with db.cursor() as cur:
    if next_company_id:
      cur.execute(
        "SELECT id FROM org_positions WHERE name = %s AND company_id = %s AND id <> %s LIMIT 1",
        (next_name, next_company_id, position_id)
      )
    else:
      cur.execute(
        "SELECT id FROM org_positions WHERE name = %s AND company_id IS NULL AND id <> %s LIMIT 1",
        (next_name, position_id)
      )
    if cur.fetchone():
      return jsonify({"error": "position_name_exists"}), 409

  try:
    with db.cursor() as cur:
      cur.execute(
        "UPDATE org_positions "
        "SET name = %s, code = %s, company_id = %s, status = %s, updated_by = %s "
        "WHERE id = %s",
        (next_name, next_code, next_company_id, next_status, g.user["id"], position_id)
      )
      cur.execute(
        "SELECT p.id, p.name, p.code, p.company_id, c.name AS company_name, "
        "p.status, p.created_at, p.updated_at "
        "FROM org_positions p LEFT JOIN companies c ON c.id = p.company_id "
        "WHERE p.id = %s",
        (position_id,)
      )
      updated = cur.fetchone()
  except IntegrityError:
    return jsonify({"error": "position_code_exists"}), 409

  return jsonify({"data": updated})


@app.route("/org/positions/<int:position_id>", methods=["DELETE"])
@require_user
def delete_org_position(position_id):
  guard = ensure_org_access()
  if guard:
    return guard

  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT id, company_id FROM org_positions WHERE id = %s", (position_id,))
    existing = cur.fetchone()
  if not existing:
    return jsonify({"error": "not_found"}), 404
  if not _can_edit_org_dimension_scope(g.user, existing.get("company_id")):
    return jsonify({"error": "forbidden"}), 403

  with db.cursor() as cur:
    cur.execute("DELETE FROM org_positions WHERE id = %s", (position_id,))
  return jsonify({"data": {"id": position_id}})


@app.route("/users", methods=["GET"])
@require_user
def list_users():
  guard = ensure_org_access()
  if guard:
    return guard

  filters = []
  params = []

  if is_sub_admin(g.user):
    if not g.user.get("company_id"):
      return jsonify({"error": "user_missing_company"}), 400
    filters.append("company_id = %s")
    params.append(g.user["company_id"])
  else:
    company_id = request.args.get("company_id", type=int)
    if company_id is not None:
      if company_id == 0:
        filters.append("company_id IS NULL")
      elif company_id:
        filters.append("company_id = %s")
        params.append(company_id)

  role = request.args.get("role")
  if role:
    if role not in USER_ROLES:
      return jsonify({"error": "invalid_role"}), 400
    filters.append("role = %s")
    params.append(role)

  status = request.args.get("status")
  if status:
    if status not in USER_STATUSES:
      return jsonify({"error": "invalid_status"}), 400
    filters.append("status = %s")
    params.append(status)

  name = request.args.get("name")
  if name:
    filters.append("name LIKE %s")
    params.append(f"%{name}%")

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

  db = get_db()
  with db.cursor() as cur:
    cur.execute(
      f"SELECT id, name, email, role, company_id, status, created_at FROM users {where_clause} ORDER BY id ASC",
      params
    )
    rows = cur.fetchall()
  _attach_org_dimensions_to_users(db, rows)

  return jsonify({"data": rows})


@app.route("/users", methods=["POST"])
@require_user
def create_user():
  body = request.get_json(silent=True) or {}
  name = body.get("name")
  role = body.get("role")
  email = body.get("email")
  status = body.get("status", "active")
  company_id = body.get("company_id")
  password = body.get("password")
  raw_org_role_ids = body.get("org_role_ids")
  raw_org_position_ids = body.get("org_position_ids")
  if company_id == 0:
    company_id = None

  if not name:
    return jsonify({"error": "missing_name"}), 400
  if not role or role not in USER_ROLES:
    return jsonify({"error": "invalid_role"}), 400
  if status not in USER_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  db = get_db()
  if is_group_admin(g.user):
    if role != "group_admin":
      if not company_id:
        return jsonify({"error": "company_id_required"}), 400
    if company_id:
      with db.cursor() as cur:
        cur.execute("SELECT id FROM companies WHERE id = %s", (company_id,))
        if not cur.fetchone():
          return jsonify({"error": "invalid_company"}), 400
  elif is_sub_admin(g.user):
    if role == "group_admin":
      return jsonify({"error": "invalid_role"}), 400
    if not g.user.get("company_id"):
      return jsonify({"error": "user_missing_company"}), 400
    company_id = g.user.get("company_id")
  else:
    return jsonify({"error": "forbidden"}), 403

  try:
    org_role_ids = _validate_user_org_dimension_ids(db, "org_roles", "id", raw_org_role_ids, company_id)
  except ValueError as err:
    code = str(err)
    if code == "inactive_org_dimension":
      return jsonify({"error": "inactive_org_role"}), 400
    if code == "org_dimension_scope_mismatch":
      return jsonify({"error": "invalid_org_role_scope"}), 400
    return jsonify({"error": "invalid_org_role_ids"}), 400

  try:
    org_position_ids = _validate_user_org_dimension_ids(
      db,
      "org_positions",
      "id",
      raw_org_position_ids,
      company_id
    )
  except ValueError as err:
    code = str(err)
    if code == "inactive_org_dimension":
      return jsonify({"error": "inactive_org_position"}), 400
    if code == "org_dimension_scope_mismatch":
      return jsonify({"error": "invalid_org_position_scope"}), 400
    return jsonify({"error": "invalid_org_position_ids"}), 400

  password_hash = hash_password(password or DEFAULT_USER_PASSWORD)

  with _db_transaction(db):
    with db.cursor() as cur:
      cur.execute(
        "INSERT INTO users (name, email, role, company_id, status, password_hash) VALUES (%s, %s, %s, %s, %s, %s)",
        (name, email, role, company_id, status, password_hash)
      )
      user_id = cur.lastrowid
    _replace_user_org_dimensions(db, user_id, role_ids=org_role_ids, position_ids=org_position_ids)
    with db.cursor() as cur:
      cur.execute(
        "SELECT id, name, email, role, company_id, status, created_at FROM users WHERE id = %s",
        (user_id,)
      )
      created = cur.fetchone()
  _attach_org_dimensions_to_users(db, [created])

  return jsonify({"data": created}), 201


@app.route("/users/<int:user_id>", methods=["PATCH"])
@require_user
def update_user(user_id):
  body = request.get_json(silent=True) or {}
  allowed = {"name", "email", "role", "company_id", "status", "password"}
  patch_role_ids = "org_role_ids" in body
  patch_position_ids = "org_position_ids" in body

  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM users WHERE id = %s", (user_id,))
    existing = cur.fetchone()

  if not existing:
    return jsonify({"error": "not_found"}), 404

  if not (is_group_admin(g.user) or is_sub_admin(g.user)):
    return jsonify({"error": "forbidden"}), 403

  if is_sub_admin(g.user):
    if existing.get("role") == "group_admin":
      return jsonify({"error": "forbidden"}), 403
    if not g.user.get("company_id") or existing.get("company_id") != g.user.get("company_id"):
      return jsonify({"error": "forbidden"}), 403

  updates = []
  params = []

  next_role = body.get("role", existing.get("role"))
  next_company_id = body.get("company_id", existing.get("company_id"))

  if is_sub_admin(g.user):
    if next_role == "group_admin":
      return jsonify({"error": "invalid_role"}), 400
    if next_company_id not in (None, 0, g.user.get("company_id")):
      return jsonify({"error": "invalid_company"}), 400
    next_company_id = g.user.get("company_id")

  if next_role not in USER_ROLES:
    return jsonify({"error": "invalid_role"}), 400
  if next_role != "group_admin" and not next_company_id:
    return jsonify({"error": "company_id_required"}), 400

  if "status" in body:
    status_value = body.get("status")
    if status_value not in USER_STATUSES:
      return jsonify({"error": "invalid_status"}), 400

  if "company_id" in body:
    if next_company_id == 0:
      next_company_id = None
    if not is_sub_admin(g.user):
      with db.cursor() as cur:
        if next_company_id:
          cur.execute("SELECT id FROM companies WHERE id = %s", (next_company_id,))
          if not cur.fetchone():
            return jsonify({"error": "invalid_company"}), 400

  role_ids_to_apply = None
  position_ids_to_apply = None
  if patch_role_ids:
    try:
      role_ids_to_apply = _validate_user_org_dimension_ids(
        db,
        "org_roles",
        "id",
        body.get("org_role_ids"),
        next_company_id
      )
    except ValueError as err:
      code = str(err)
      if code == "inactive_org_dimension":
        return jsonify({"error": "inactive_org_role"}), 400
      if code == "org_dimension_scope_mismatch":
        return jsonify({"error": "invalid_org_role_scope"}), 400
      return jsonify({"error": "invalid_org_role_ids"}), 400

  if patch_position_ids:
    try:
      position_ids_to_apply = _validate_user_org_dimension_ids(
        db,
        "org_positions",
        "id",
        body.get("org_position_ids"),
        next_company_id
      )
    except ValueError as err:
      code = str(err)
      if code == "inactive_org_dimension":
        return jsonify({"error": "inactive_org_position"}), 400
      if code == "org_dimension_scope_mismatch":
        return jsonify({"error": "invalid_org_position_scope"}), 400
      return jsonify({"error": "invalid_org_position_ids"}), 400

  for key in allowed:
    if key not in body:
      continue
    if key == "password":
      password_value = body.get("password")
      if not password_value:
        continue
      updates.append("password_hash = %s")
      params.append(hash_password(password_value))
      continue
    value = body.get(key)
    if key == "company_id":
      if is_sub_admin(g.user):
        value = g.user.get("company_id")
      elif value == 0:
        value = None
    updates.append(f"{key} = %s")
    params.append(value)

  if not updates and not patch_role_ids and not patch_position_ids:
    return jsonify({"error": "no_updates"}), 400

  with _db_transaction(db):
    if updates:
      params.append(user_id)
      with db.cursor() as cur:
        cur.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = %s", params)
    _replace_user_org_dimensions(
      db,
      user_id,
      role_ids=role_ids_to_apply if patch_role_ids else None,
      position_ids=position_ids_to_apply if patch_position_ids else None
    )
    with db.cursor() as cur:
      cur.execute(
        "SELECT id, name, email, role, company_id, status, created_at FROM users WHERE id = %s",
        (user_id,)
      )
      updated = cur.fetchone()
  _attach_org_dimensions_to_users(db, [updated])

  return jsonify({"data": updated})


@app.route("/me/password", methods=["PATCH"])
@require_user
def update_my_password():
  body = request.get_json(silent=True) or {}
  current_password = body.get("current_password")
  new_password = body.get("new_password")

  if not current_password or not new_password:
    return jsonify({"error": "missing_password"}), 400

  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT id, password_hash FROM users WHERE id = %s", (g.user["id"],))
    user = cur.fetchone()

  if not user:
    return jsonify({"error": "not_found"}), 404

  password_hash = user.get("password_hash")
  if password_hash:
    if not check_password_hash(password_hash, current_password):
      return jsonify({"error": "invalid_password"}), 400
  else:
    if current_password != DEFAULT_USER_PASSWORD:
      return jsonify({"error": "invalid_password"}), 400

  new_hash = hash_password(new_password)
  with db.cursor() as cur:
    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (new_hash, g.user["id"]))

  return jsonify({"data": {"id": g.user["id"]}})


@app.route("/opportunities", methods=["GET"])
@require_user
def list_opportunities():
  user = g.user
  filters = []
  params = []

  company_id = request.args.get("company_id", type=int)
  if is_group_admin(user):
    if company_id:
      filters.append("company_id = %s")
      params.append(company_id)
  else:
    if not user.get("company_id"):
      return jsonify({"error": "user_missing_company"}), 400
    filters.append("company_id = %s")
    params.append(user["company_id"])

  for field in ["type", "status", "stage", "source", "city", "industry"]:
    value = request.args.get(field)
    if value:
      filters.append(f"{field} = %s")
      params.append(value)

  owner_id = request.args.get("owner_id", type=int)
  if owner_id:
    filters.append("owner_id = %s")
    params.append(owner_id)

  name = request.args.get("name")
  if name:
    filters.append("name LIKE %s")
    params.append(f"%{name}%")

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""
  page = max(request.args.get("page", default=1, type=int), 1)
  page_size = request.args.get("page_size", type=int)
  if page_size:
    limit = min(max(page_size, 1), 500)
  else:
    limit = min(request.args.get("limit", default=200, type=int), 500)
  offset = (page - 1) * limit

  db = get_db()
  with db.cursor() as cur:
    cur.execute(
      f"SELECT COUNT(*) AS total, "
      "SUM(CASE WHEN status = 'valid' THEN 1 ELSE 0 END) AS valid_count, "
      "SUM(CASE WHEN status = 'in_progress' THEN 1 ELSE 0 END) AS in_progress_count, "
      "SUM(CASE WHEN stage = 'ready_for_handoff' THEN 1 ELSE 0 END) AS ready_for_handoff_count, "
      "SUM(CASE WHEN type = 'host' THEN 1 ELSE 0 END) AS host_count "
      f"FROM opportunities {where_clause}",
      params
    )
    summary_row = cur.fetchone() or {}
    cur.execute(
      f"SELECT COUNT(*) AS follow_up_count FROM activities "
      f"WHERE opportunity_id IN (SELECT id FROM opportunities {where_clause})",
      params
    )
    follow_up_row = cur.fetchone() or {}
    cur.execute(
      f"SELECT COUNT(*) AS contact_count FROM opportunity_contacts "
      f"WHERE opportunity_id IN (SELECT id FROM opportunities {where_clause})",
      params
    )
    contact_row = cur.fetchone() or {}
    cur.execute(
      f"SELECT COUNT(*) AS persona_count FROM opportunity_insights "
      f"WHERE opportunity_id IN (SELECT id FROM opportunities {where_clause})",
      params
    )
    persona_row = cur.fetchone() or {}
    total = summary_row.get("total", 0)
    cur.execute(
      f"SELECT * FROM opportunities {where_clause} ORDER BY updated_at DESC LIMIT %s OFFSET %s",
      params + [limit, offset]
    )
    rows = cur.fetchall()

  summary = {
    "total": _to_int(total),
    "valid": _to_int(summary_row.get("valid_count", 0)),
    "in_progress": _to_int(summary_row.get("in_progress_count", 0)),
    "ready_for_handoff": _to_int(summary_row.get("ready_for_handoff_count", 0)),
    "host": _to_int(summary_row.get("host_count", 0)),
    "follow_ups": _to_int(follow_up_row.get("follow_up_count", 0)),
    "contacts": _to_int(contact_row.get("contact_count", 0)),
    "personas": _to_int(persona_row.get("persona_count", 0))
  }

  return jsonify(
    {
      "data": rows,
      "total": _to_int(total),
      "page": page,
      "page_size": limit,
      "summary": summary
    }
  )


@app.route("/opportunities", methods=["POST"])
@require_user
def create_opportunity():
  user = g.user
  body = request.get_json(silent=True) or {}

  if not body.get("name") or not body.get("type") or not body.get("source"):
    return jsonify({"error": "missing_required_fields"}), 400
  if body["type"] not in OPPORTUNITY_TYPES:
    return jsonify({"error": "invalid_type"}), 400
  if body.get("status") and body["status"] not in OPPORTUNITY_STATUSES:
    return jsonify({"error": "invalid_status"}), 400
  if body.get("stage") and body["stage"] not in OPPORTUNITY_STAGES:
    return jsonify({"error": "invalid_stage"}), 400
  if body.get("organizer_type") and body["organizer_type"] not in ORGANIZER_TYPES:
    return jsonify({"error": "invalid_organizer_type"}), 400

  if is_group_admin(user):
    company_id = int(body.get("company_id") or 0)
    if not company_id:
      return jsonify({"error": "company_id_required"}), 400
    owner_id = int(body.get("owner_id") or user["id"])
  else:
    company_id = user.get("company_id")
    if not company_id:
      return jsonify({"error": "user_missing_company"}), 400
    owner_id = user["id"]

  contacts = _normalize_contacts(body.get("contacts"))
  if contacts:
    primary = contacts[0]
    body.setdefault("contact_name", primary.get("name"))
    body.setdefault("contact_title", primary.get("title"))
    body.setdefault("contact_phone", primary.get("phone"))
    body.setdefault("contact_email", primary.get("email"))
    body.setdefault("contact_wechat", primary.get("wechat"))
  if body.get("contact_person") and not body.get("contact_name"):
    body["contact_name"] = body.get("contact_person")

  sql = (
    "INSERT INTO opportunities ("
    "name, type, source, industry, city, status, stage, "
    "owner_id, company_id, "
    "organizer_name, organizer_type, exhibition_name, "
    "exhibition_start_date, exhibition_end_date, venue_name, venue_address, "
    "booth_count, exhibition_area_sqm, expected_visitors, exhibition_theme, budget_range, "
    "risk_notes, "
    "contact_name, contact_title, contact_phone, contact_email, contact_wechat, "
    "company_name, company_phone, company_email, contact_department, contact_person, "
    "contact_address, website, country, hall_no, booth_no, booth_type, booth_area_sqm, "
    "invalid_reason"
    ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
  )
  params = (
    body["name"],
    body["type"],
    body["source"],
    body.get("industry"),
    body.get("city"),
    body.get("status", "new"),
    body.get("stage", "cold"),
    owner_id,
    company_id,
    body.get("organizer_name"),
    body.get("organizer_type"),
    body.get("exhibition_name"),
    body.get("exhibition_start_date"),
    body.get("exhibition_end_date"),
    body.get("venue_name"),
    body.get("venue_address"),
    body.get("booth_count"),
    body.get("exhibition_area_sqm"),
    body.get("expected_visitors"),
    body.get("exhibition_theme"),
    body.get("budget_range"),
    body.get("risk_notes"),
    body.get("contact_name"),
    body.get("contact_title"),
    body.get("contact_phone"),
    body.get("contact_email"),
    body.get("contact_wechat"),
    body.get("company_name"),
    body.get("company_phone"),
    body.get("company_email"),
    body.get("contact_department"),
    body.get("contact_person"),
    body.get("contact_address"),
    body.get("website"),
    body.get("country"),
    body.get("hall_no"),
    body.get("booth_no"),
    body.get("booth_type"),
    body.get("booth_area_sqm"),
    body.get("invalid_reason")
  )

  db = get_db()
  with db.cursor() as cur:
    cur.execute(sql, params)
    new_id = cur.lastrowid
    if contacts:
      _insert_contacts(cur, new_id, contacts)
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (new_id,))
    created = cur.fetchone()

  return jsonify({"data": created}), 201


@app.route("/opportunities/<int:opportunity_id>", methods=["GET"])
@require_user
def get_opportunity(opportunity_id):
  user = g.user
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    opportunity = cur.fetchone()

  if not opportunity:
    return jsonify({"error": "not_found"}), 404

  if not is_group_admin(user) and opportunity.get("company_id") != user.get("company_id"):
    return jsonify({"error": "not_found"}), 404

  return jsonify({"data": opportunity})


@app.route("/opportunities/<int:opportunity_id>", methods=["PATCH"])
@require_user
def update_opportunity(opportunity_id):
  user = g.user
  db = get_db()

  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    opportunity = cur.fetchone()

  if not opportunity:
    return jsonify({"error": "not_found"}), 404

  if not is_group_admin(user) and opportunity.get("company_id") != user.get("company_id"):
    return jsonify({"error": "not_found"}), 404

  body = request.get_json(silent=True) or {}
  contacts_in_body = "contacts" in body
  contacts = _normalize_contacts(body.get("contacts"))
  allowed = {
    "name",
    "type",
    "source",
    "industry",
    "city",
    "status",
    "stage",
    "organizer_name",
    "organizer_type",
    "exhibition_name",
    "exhibition_start_date",
    "exhibition_end_date",
    "venue_name",
    "venue_address",
    "booth_count",
    "exhibition_area_sqm",
    "expected_visitors",
    "exhibition_theme",
    "budget_range",
    "risk_notes",
    "contact_name",
    "contact_title",
    "contact_phone",
    "contact_email",
    "contact_wechat",
    "company_name",
    "company_phone",
    "company_email",
    "contact_department",
    "contact_person",
    "contact_address",
    "website",
    "country",
    "hall_no",
    "booth_no",
    "booth_type",
    "booth_area_sqm",
    "invalid_reason",
    "owner_id"
  }

  updates = []
  params = []

  for key in allowed:
    if key not in body:
      continue
    if key == "type" and body[key] not in OPPORTUNITY_TYPES:
      return jsonify({"error": "invalid_type"}), 400
    if key == "status" and body[key] not in OPPORTUNITY_STATUSES:
      return jsonify({"error": "invalid_status"}), 400
    if key == "stage" and body[key] not in OPPORTUNITY_STAGES:
      return jsonify({"error": "invalid_stage"}), 400
    if key == "organizer_type" and body[key] and body[key] not in ORGANIZER_TYPES:
      return jsonify({"error": "invalid_organizer_type"}), 400
    if key == "owner_id" and not is_group_admin(user):
      continue

    updates.append(f"{key} = %s")
    params.append(body[key])

  if not updates:
    if not contacts_in_body:
      return jsonify({"error": "no_updates"}), 400

  params.append(opportunity_id)

  with db.cursor() as cur:
    if updates:
      cur.execute(f"UPDATE opportunities SET {', '.join(updates)} WHERE id = %s", params)
    if contacts_in_body:
      cur.execute("DELETE FROM opportunity_contacts WHERE opportunity_id = %s", (opportunity_id,))
      _insert_contacts(cur, opportunity_id, contacts)
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    updated = cur.fetchone()

  return jsonify({"data": updated})


@app.route("/opportunities/<int:opportunity_id>/analysis", methods=["GET"])
@require_user
def get_opportunity_analysis(opportunity_id):
  user = g.user
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    opportunity = cur.fetchone()

  if not opportunity:
    return jsonify({"error": "not_found"}), 404

  if not is_group_admin(user) and opportunity.get("company_id") != user.get("company_id"):
    return jsonify({"error": "not_found"}), 404

  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunity_insights WHERE opportunity_id = %s", (opportunity_id,))
    insight = cur.fetchone()

  return jsonify({"data": _serialize_insight(insight)})


@app.route("/opportunities/<int:opportunity_id>/analysis", methods=["POST"])
@require_user
def analyze_opportunity(opportunity_id):
  user = g.user
  db = get_db()

  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    opportunity = cur.fetchone()

  if not opportunity:
    return jsonify({"error": "not_found"}), 404

  if not is_group_admin(user) and opportunity.get("company_id") != user.get("company_id"):
    return jsonify({"error": "not_found"}), 404

  if not opportunity.get("name"):
    return jsonify({"error": "missing_opportunity_name"}), 400

  query = _build_analysis_query(opportunity)
  try:
    search_results = _search_web(query, num=6)
  except RuntimeError as err:
    return jsonify({"error": str(err)}), 400
  except Exception:
    return jsonify({"error": "search_failed"}), 500

  if not search_results:
    fallback_terms = [
      opportunity.get("exhibition_name"),
      opportunity.get("organizer_name"),
      opportunity.get("name")
    ]
    for term in fallback_terms:
      if not term:
        continue
      try:
        search_results = _search_web(str(term), num=6)
      except Exception:
        continue
      if search_results:
        break

  if not search_results:
    return jsonify({"error": "no_search_results"}), 400

  sources_payload = []
  sources_context = []
  seen_urls = set()

  for item in search_results:
    url = item.get("url")
    if not url or url in seen_urls:
      continue
    seen_urls.add(url)
    title = item.get("title") or ""
    snippet = item.get("snippet") or ""
    sources_payload.append({"title": title, "url": url, "snippet": snippet})

    if len(sources_context) < 4:
      content = _fetch_url_text(url)
      sources_context.append(
        {
          "title": title,
          "url": url,
          "snippet": snippet,
          "content": content or ""
        }
      )

  messages = _build_analysis_messages(opportunity, sources_context or sources_payload)
  try:
    raw = _call_chat_completion(messages, temperature=0.1)
  except Exception:
    return jsonify({"error": "analysis_failed"}), 500

  analysis_data = _safe_json_load(raw)
  if not isinstance(analysis_data, dict) or not analysis_data:
    try:
      fix_messages = [
        {
          "role": "system",
          "content": "你是JSON修复助手。将输入转换为严格JSON，只输出JSON，不要解释。"
        },
        {
          "role": "user",
          "content": f"请转换为以下结构的JSON：\n{ANALYSIS_SCHEMA}\n\n原始内容:\n{raw}"
        }
      ]
      fixed = _call_chat_completion(fix_messages, temperature=0.0)
      analysis_data = _safe_json_load(fixed)
    except Exception:
      analysis_data = {}

  if not isinstance(analysis_data, dict) or not analysis_data:
    analysis_data = {"_parse_warning": "model_output_invalid_json"}

  contacts = analysis_data.get("contacts")
  if not isinstance(contacts, list):
    contacts = []
  analysis_data["contacts"] = contacts
  analysis_data = _apply_analysis_defaults(analysis_data, opportunity)

  use_azure = os.getenv("AZURE_OPENAI_DEFAULT_FLAG", "").upper() == "Y"
  provider = "azure_openai" if use_azure else "openai"
  model = (
    os.getenv("AZURE_OPENAI_MODEL_ID", DEFAULT_ANALYSIS_MODEL)
    if use_azure
    else os.getenv("OPENAI_MODEL_ID", DEFAULT_ANALYSIS_MODEL)
  )

  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO opportunity_insights (opportunity_id, analysis_json, contacts_json, sources_json, provider, model) "
      "VALUES (%s, %s, %s, %s, %s, %s) "
      "ON DUPLICATE KEY UPDATE analysis_json = VALUES(analysis_json), contacts_json = VALUES(contacts_json), "
      "sources_json = VALUES(sources_json), provider = VALUES(provider), model = VALUES(model), "
      "updated_at = CURRENT_TIMESTAMP",
      (
        opportunity_id,
        json.dumps(analysis_data, ensure_ascii=False),
        json.dumps(contacts, ensure_ascii=False),
        json.dumps(sources_payload, ensure_ascii=False),
        provider,
        model
      )
    )
    cur.execute("SELECT * FROM opportunity_insights WHERE opportunity_id = %s", (opportunity_id,))
    insight = cur.fetchone()

  return jsonify({"data": _serialize_insight(insight)})


@app.route("/imports/sheets", methods=["GET"])
@require_user
def list_import_sheets():
  filename = request.args.get("filename") or DEFAULT_IMPORT_FILE
  path = _resolve_import_path(filename)
  if not path:
    return jsonify({"error": "file_not_found"}), 404
  try:
    workbook = load_workbook(path, read_only=True, data_only=True)
    return jsonify({"data": {"sheets": workbook.sheetnames}})
  except Exception:
    return jsonify({"error": "invalid_excel"}), 400


@app.route("/imports/upload", methods=["POST"])
@require_user
def upload_import_file():
  if "file" not in request.files:
    return jsonify({"error": "file_required"}), 400
  file = request.files["file"]
  if not file or not file.filename:
    return jsonify({"error": "file_required"}), 400
  ext = os.path.splitext(file.filename)[1].lower()
  if ext not in ALLOWED_IMPORT_EXTENSIONS:
    return jsonify({"error": "invalid_file_type"}), 400

  os.makedirs(UPLOAD_DIR, exist_ok=True)
  safe_name = secure_filename(file.filename)
  timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
  stored_name = f"{timestamp}_{safe_name}" if safe_name else f"{timestamp}.xlsx"
  stored_path = os.path.join(UPLOAD_DIR, stored_name)
  file.save(stored_path)

  return jsonify({"data": {"filename": stored_name}})


@app.route("/imports/opportunities", methods=["POST"])
@require_user
def import_opportunities():
  user = g.user
  body = request.get_json(silent=True) or {}
  filename = body.get("filename") or DEFAULT_IMPORT_FILE
  sheet_name = body.get("sheet") or "总表"
  company_id = body.get("company_id") if "company_id" in body else user.get("company_id")
  if company_id == "":
    company_id = None
  if company_id is None and is_group_admin(user):
    company_id = 0

  if company_id is None:
    return jsonify({"error": "company_required"}), 400

  if str(company_id) == "0":
    with get_db().cursor() as cur:
      cur.execute("SELECT id FROM companies WHERE parent_id IS NULL ORDER BY id ASC LIMIT 1")
      root = cur.fetchone()
    if not root:
      return jsonify({"error": "group_company_not_found"}), 400
    company_id = root["id"]

  default_company_id = company_id

  path = _resolve_import_path(filename)
  if not path:
    return jsonify({"error": "file_not_found"}), 404

  try:
    workbook = load_workbook(path, read_only=True, data_only=True)
  except Exception:
    return jsonify({"error": "invalid_excel"}), 400

  if sheet_name not in workbook.sheetnames:
    return jsonify({"error": "sheet_not_found"}), 404

  sheet = workbook[sheet_name]
  header_row_index, header_row = _find_header_row(sheet)
  if not header_row_index:
    return jsonify({"error": "header_not_found"}), 400

  header_map = _build_header_index_map(header_row)
  inserted = 0
  updated = 0
  skipped = 0
  contacts_added = 0

  def cell_value(row, key):
    idx = header_map.get(key)
    if idx is None or idx >= len(row):
      return None
    return _stringify_cell(row[idx])

  with get_db().cursor() as cur:
    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
      if not row or not any(row):
        continue
      company_cn = cell_value(row, "company_cn")
      company_en = cell_value(row, "company_en")
      name = company_cn or company_en
      if not name:
        skipped += 1
        continue

      city = cell_value(row, "region")
      industry = cell_value(row, "business_type")
      source = f"Excel导入:{sheet_name}"
      pic = cell_value(row, "pic")

      target_company_id = default_company_id
      target_owner_id = user.get("id")
      if is_group_admin(user) and pic:
        cur.execute(
          "SELECT id, company_id FROM users WHERE name = %s AND status = 'active' LIMIT 1",
          (pic,)
        )
        pic_user = cur.fetchone()
        if pic_user and pic_user.get("company_id"):
          target_company_id = pic_user["company_id"]
          target_owner_id = pic_user["id"]

      primary_contact_name = cell_value(row, "contact_name")
      primary_phone = cell_value(row, "contact_mobile") or cell_value(row, "contact_phone")
      primary_email = cell_value(row, "contact_email")

      secondary_contact_name = cell_value(row, "contact_alt_name")
      secondary_contact_title = cell_value(row, "contact_alt_title")
      secondary_contact_email = cell_value(row, "contact_email_alt")
      secondary_contact_phone = cell_value(row, "contact_alt_phone")

      company_name = name
      company_phone = cell_value(row, "company_phone") or primary_phone
      company_email = cell_value(row, "company_email") or primary_email
      contact_department = cell_value(row, "contact_department")
      contact_person = cell_value(row, "contact_person") or primary_contact_name
      contact_address = cell_value(row, "contact_address") or cell_value(row, "address")
      website = cell_value(row, "website")
      country = cell_value(row, "country")
      hall_no = cell_value(row, "hall_no")
      booth_no = cell_value(row, "booth_no")
      booth_type = cell_value(row, "booth_type")
      booth_area_sqm = _parse_int(cell_value(row, "booth_area_sqm"))

      notes = _build_notes(
        [
          f"英文名: {company_en}" if company_en and company_en != company_cn else None,
          f"官网: {cell_value(row, 'website')}" if cell_value(row, "website") else None,
          f"地址: {cell_value(row, 'address')}" if cell_value(row, "address") else None,
          f"跟进人/PIC: {cell_value(row, 'pic')}" if cell_value(row, "pic") else None,
          f"触达方式: {cell_value(row, 'touch_method')}" if cell_value(row, "touch_method") else None,
          f"是否取得联系: {cell_value(row, 'contacted')}" if cell_value(row, "contacted") else None,
          f"现场反馈: {cell_value(row, 'feedback')}" if cell_value(row, "feedback") else None,
          f"竞争对手: {cell_value(row, 'competitor')}" if cell_value(row, "competitor") else None,
          f"备注: {cell_value(row, 'remark')}" if cell_value(row, "remark") else None,
          f"业务名称: {cell_value(row, 'business_name')}" if cell_value(row, "business_name") else None
        ]
      )

      if is_group_admin(user):
        cur.execute(
          "SELECT * FROM opportunities WHERE name = %s ORDER BY id DESC LIMIT 1",
          (name,)
        )
      else:
        cur.execute(
          "SELECT * FROM opportunities WHERE name = %s AND company_id = %s ORDER BY id DESC LIMIT 1",
          (name, target_company_id)
        )
      existing = cur.fetchone()

      if existing:
        updates = {}
        if is_group_admin(user) and target_company_id and existing.get("company_id") != target_company_id:
          updates["company_id"] = target_company_id
        if is_group_admin(user) and target_owner_id and existing.get("owner_id") != target_owner_id:
          updates["owner_id"] = target_owner_id
        if city and not existing.get("city"):
          updates["city"] = city
        if industry and not existing.get("industry"):
          updates["industry"] = industry
        if primary_contact_name and not existing.get("contact_name"):
          updates["contact_name"] = primary_contact_name
        if primary_phone and not existing.get("contact_phone"):
          updates["contact_phone"] = primary_phone
        if primary_email and not existing.get("contact_email"):
          updates["contact_email"] = primary_email
        if company_name and not existing.get("company_name"):
          updates["company_name"] = company_name
        if company_phone and not existing.get("company_phone"):
          updates["company_phone"] = company_phone
        if company_email and not existing.get("company_email"):
          updates["company_email"] = company_email
        if contact_department and not existing.get("contact_department"):
          updates["contact_department"] = contact_department
        if contact_person and not existing.get("contact_person"):
          updates["contact_person"] = contact_person
        if contact_address and not existing.get("contact_address"):
          updates["contact_address"] = contact_address
        if website and not existing.get("website"):
          updates["website"] = website
        if country and not existing.get("country"):
          updates["country"] = country
        if hall_no and not existing.get("hall_no"):
          updates["hall_no"] = hall_no
        if booth_no and not existing.get("booth_no"):
          updates["booth_no"] = booth_no
        if booth_type and not existing.get("booth_type"):
          updates["booth_type"] = booth_type
        if booth_area_sqm and not existing.get("booth_area_sqm"):
          updates["booth_area_sqm"] = booth_area_sqm
        if notes and not existing.get("risk_notes"):
          updates["risk_notes"] = notes
        if updates:
          set_clause = ", ".join([f"{key} = %s" for key in updates.keys()])
          cur.execute(
            f"UPDATE opportunities SET {set_clause} WHERE id = %s",
            (*updates.values(), existing["id"])
          )
          updated += 1
        opportunity_id = existing["id"]
      else:
        cur.execute(
          "INSERT INTO opportunities (name, type, source, industry, city, owner_id, company_id, contact_name, contact_phone, contact_email, "
          "company_name, company_phone, company_email, contact_department, contact_person, contact_address, website, country, hall_no, booth_no, booth_type, booth_area_sqm, "
          "risk_notes) "
          "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
          (
            name,
            "normal",
            source,
            industry,
            city,
            target_owner_id,
            target_company_id,
            primary_contact_name,
            primary_phone,
            primary_email,
            company_name,
            company_phone,
            company_email,
            contact_department,
            contact_person,
            contact_address,
            website,
            country,
            hall_no,
            booth_no,
            booth_type,
            booth_area_sqm,
            notes
          )
        )
        opportunity_id = cur.lastrowid
        inserted += 1

      contacts = []
      if primary_contact_name or primary_phone or primary_email:
        contacts.append(
          {
            "name": primary_contact_name,
            "title": None,
            "phone": primary_phone,
            "email": primary_email,
            "wechat": None
          }
        )
      if secondary_contact_name or secondary_contact_title or secondary_contact_phone or secondary_contact_email:
        contacts.append(
          {
            "name": secondary_contact_name,
            "title": secondary_contact_title,
            "phone": secondary_contact_phone,
            "email": secondary_contact_email,
            "wechat": None
          }
        )
      contacts_added += _insert_contacts(cur, opportunity_id, contacts)

  return jsonify(
    {
      "data": {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "contacts_added": contacts_added
      }
    }
  )


@app.route("/host-pool/events", methods=["GET"])
@require_user
def list_host_pool_events():
  filters = []
  params = []

  keyword = (request.args.get("keyword") or "").strip()
  if keyword:
    filters.append("(name LIKE %s OR alias_name LIKE %s OR industry LIKE %s OR city LIKE %s OR organizer_name LIKE %s)")
    like_value = f"%{keyword}%"
    params.extend([like_value, like_value, like_value, like_value, like_value])

  city = (request.args.get("city") or "").strip()
  if city:
    filters.append("city = %s")
    params.append(city)

  industry = (request.args.get("industry") or "").strip()
  if industry:
    filters.append("industry LIKE %s")
    params.append(f"%{industry}%")

  pool_status = (request.args.get("pool_status") or "").strip()
  if pool_status:
    if pool_status not in HOST_POOL_STATUSES:
      return jsonify({"error": "invalid_pool_status"}), 400
    filters.append("pool_status = %s")
    params.append(pool_status)

  is_domestic_raw = request.args.get("is_domestic")
  if is_domestic_raw not in (None, ""):
    is_domestic = 1 if _to_bool(is_domestic_raw, default=True) else 0
    filters.append("is_domestic = %s")
    params.append(is_domestic)

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""
  page = max(request.args.get("page", default=1, type=int), 1)
  page_size = request.args.get("page_size", type=int)
  if page_size:
    limit = min(max(page_size, 1), 500)
  else:
    limit = min(request.args.get("limit", default=20, type=int), 500)
  offset = (page - 1) * limit

  db = get_db()
  with db.cursor() as cur:
    cur.execute(f"SELECT COUNT(*) AS total FROM host_opportunity_pool_events {where_clause}", params)
    total_row = cur.fetchone() or {}
    cur.execute(
      f"SELECT "
      "SUM(CASE WHEN pool_status = 'active' THEN 1 ELSE 0 END) AS active_count, "
      "SUM(CASE WHEN pool_status = 'converted' THEN 1 ELSE 0 END) AS converted_count, "
      "SUM(CASE WHEN pool_status = 'archived' THEN 1 ELSE 0 END) AS archived_count "
      f"FROM host_opportunity_pool_events {where_clause}",
      params
    )
    summary_row = cur.fetchone() or {}
    cur.execute(
      f"SELECT * FROM host_opportunity_pool_events {where_clause} "
      "ORDER BY "
      "CASE WHEN exhibition_start_date IS NULL THEN 1 ELSE 0 END, "
      "exhibition_start_date ASC, id DESC "
      "LIMIT %s OFFSET %s",
      params + [limit, offset]
    )
    rows = cur.fetchall()

  summary = {
    "total": _to_int(total_row.get("total")),
    "active": _to_int(summary_row.get("active_count")),
    "converted": _to_int(summary_row.get("converted_count")),
    "archived": _to_int(summary_row.get("archived_count"))
  }
  return jsonify(
    {
      "data": rows,
      "total": _to_int(total_row.get("total")),
      "page": page,
      "page_size": limit,
      "summary": summary
    }
  )


@app.route("/host-pool/events/sync", methods=["POST"])
@require_user
def sync_host_pool_events():
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  source = (body.get("source") or "qufair").strip().lower()
  if source != "qufair":
    return jsonify({"error": "unsupported_source"}), 400

  domestic_url = str(body.get("domestic_url") or QUFAIR_DEFAULT_DOMESTIC_URL).strip()
  max_categories = min(max(_safe_parse_int(body.get("max_categories")) or 12, 1), 80)
  max_events_per_category = min(max(_safe_parse_int(body.get("max_events_per_category")) or 80, 1), 500)
  fetch_detail = _to_bool(body.get("fetch_detail"), default=True)

  category_urls = body.get("category_urls")
  selected_categories = []
  if isinstance(category_urls, list) and category_urls:
    for raw_url in category_urls:
      if not raw_url:
        continue
      url = _to_absolute_url(raw_url)
      if not url or "/fl/" not in url:
        continue
      selected_categories.append({"url": url, "industry": None, "estimated_count": 0})

  try:
    opener = _build_qufair_opener()
    domestic_html = _fetch_qufair_page(opener, domestic_url)
  except Exception:
    return jsonify({"error": "qufair_unreachable"}), 502

  discovered_categories = _parse_qufair_domestic_categories(domestic_html)
  if not selected_categories:
    selected_categories = discovered_categories[:max_categories]
  else:
    selected_categories = selected_categories[:max_categories]

  if not selected_categories:
    return jsonify({"error": "no_categories_found"}), 400

  inserted = 0
  updated = 0
  skipped = 0
  detail_failed = 0
  categories_processed = 0
  events_discovered = 0
  errors = []

  db = get_db()
  with db.cursor() as cur:
    for category in selected_categories:
      category_url = category.get("url")
      if not category_url:
        continue
      try:
        category_html = _fetch_qufair_page(opener, category_url)
      except Exception:
        errors.append(f"分类抓取失败: {category_url}")
        continue
      categories_processed += 1
      events = _parse_qufair_event_cards(
        category_html,
        source_list_url=category_url,
        fallback_industry=category.get("industry")
      )[:max_events_per_category]
      events_discovered += len(events)

      for event in events:
        if fetch_detail and event.get("source_url"):
          try:
            detail_html = _fetch_qufair_page(opener, event.get("source_url"))
            detail = _parse_qufair_detail_info(detail_html)
            event.update(detail)
          except Exception:
            detail_failed += 1
        if not event.get("external_id"):
          event["external_id"] = _extract_qufair_external_id(event.get("source_url"))
        if not event.get("name"):
          event["name"] = event.get("exhibition_name")
        if not event.get("name"):
          skipped += 1
          continue
        try:
          affected = _upsert_host_pool_event(cur, event)
          if affected == 1:
            inserted += 1
          elif affected >= 2:
            updated += 1
          else:
            skipped += 1
        except IntegrityError:
          skipped += 1
        except Exception:
          skipped += 1

  return jsonify(
    {
      "data": {
        "source": source,
        "domestic_url": domestic_url,
        "categories_found": len(discovered_categories),
        "categories_processed": categories_processed,
        "events_discovered": events_discovered,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "detail_failed": detail_failed,
        "errors": errors[:20]
      }
    }
  )


@app.route("/host-pool/events/<int:event_id>/convert", methods=["POST"])
@require_user
def convert_host_pool_event(event_id):
  user = g.user
  body = request.get_json(silent=True) or {}
  db = get_db()

  with db.cursor() as cur:
    cur.execute("SELECT * FROM host_opportunity_pool_events WHERE id = %s", (event_id,))
    event = cur.fetchone()

  if not event:
    return jsonify({"error": "not_found"}), 404

  if event.get("pool_status") == "archived":
    return jsonify({"error": "event_archived"}), 400

  if is_group_admin(user):
    company_id = _normalize_company_id(body.get("company_id", user.get("company_id")))
    if not company_id:
      with db.cursor() as cur:
        cur.execute("SELECT id FROM companies WHERE parent_id IS NULL ORDER BY id ASC LIMIT 1")
        root_company = cur.fetchone()
      company_id = (root_company or {}).get("id")
    if not company_id:
      return jsonify({"error": "company_id_required"}), 400
    if not _ensure_company_exists(db, company_id):
      return jsonify({"error": "invalid_company_id"}), 400
  else:
    company_id = user.get("company_id")
    if not company_id:
      return jsonify({"error": "user_missing_company"}), 400

  owner_id = user.get("id")
  if is_group_admin(user) and body.get("owner_id"):
    try:
      requested_owner = int(body.get("owner_id"))
    except (TypeError, ValueError):
      requested_owner = None
    if requested_owner:
      with db.cursor() as cur:
        cur.execute("SELECT id FROM users WHERE id = %s AND status = 'active' LIMIT 1", (requested_owner,))
        owner = cur.fetchone()
      if not owner:
        return jsonify({"error": "invalid_owner_id"}), 400
      owner_id = requested_owner

  source_url = event.get("source_url")
  with db.cursor() as cur:
    if source_url:
      cur.execute(
        "SELECT * FROM opportunities WHERE type = 'host' AND company_id = %s AND website = %s ORDER BY id DESC LIMIT 1",
        (company_id, source_url)
      )
      existing = cur.fetchone()
    else:
      existing = None
    if not existing:
      cur.execute(
        "SELECT * FROM opportunities WHERE type = 'host' AND company_id = %s AND name = %s ORDER BY id DESC LIMIT 1",
        (company_id, event.get("name"))
      )
      existing = cur.fetchone()

  if existing:
    with db.cursor() as cur:
      cur.execute(
        "UPDATE host_opportunity_pool_events "
        "SET pool_status = 'converted', converted_opportunity_id = %s, updated_at = CURRENT_TIMESTAMP "
        "WHERE id = %s",
        (existing["id"], event_id)
      )
    return jsonify({"data": {"already_exists": True, "opportunity": existing}})

  risk_notes = _build_notes(
    [
      f"公海池来源: {event.get('source_site')}" if event.get("source_site") else "公海池来源: qufair",
      f"来源链接: {source_url}" if source_url else None
    ],
    max_len=255
  )

  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO opportunities ("
      "name, type, source, industry, city, status, stage, owner_id, company_id, "
      "organizer_name, exhibition_name, exhibition_start_date, exhibition_end_date, "
      "venue_name, venue_address, exhibition_area_sqm, expected_visitors, risk_notes, website, country"
      ") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
      (
        event.get("name"),
        "host",
        "去展网公海池",
        event.get("industry"),
        event.get("city"),
        "new",
        "cold",
        owner_id,
        company_id,
        event.get("organizer_name"),
        event.get("name"),
        event.get("exhibition_start_date"),
        event.get("exhibition_end_date"),
        event.get("venue_name"),
        event.get("venue_address"),
        event.get("exhibition_area_sqm"),
        event.get("visitors_count"),
        risk_notes,
        source_url,
        event.get("country")
      )
    )
    new_opportunity_id = cur.lastrowid
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (new_opportunity_id,))
    created = cur.fetchone()
    cur.execute(
      "UPDATE host_opportunity_pool_events "
      "SET pool_status = 'converted', converted_opportunity_id = %s, updated_at = CURRENT_TIMESTAMP "
      "WHERE id = %s",
      (new_opportunity_id, event_id)
    )

  return jsonify({"data": {"already_exists": False, "opportunity": created}})


@app.route("/opportunities/<int:opportunity_id>/activities", methods=["POST"])
@require_user
def add_activity(opportunity_id):
  user = g.user
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    opportunity = cur.fetchone()

  if not opportunity:
    return jsonify({"error": "not_found"}), 404

  if not is_group_admin(user) and opportunity.get("company_id") != user.get("company_id"):
    return jsonify({"error": "not_found"}), 404

  body = request.get_json(silent=True) or {}
  comment = body.get("comment")
  channel = body.get("channel")
  if comment and not channel:
    channel = "other"
  if channel not in ACTIVITY_CHANNELS:
    return jsonify({"error": "invalid_channel"}), 400

  result = body.get("result")
  next_step = body.get("next_step")
  if comment:
    result = comment
    next_step = None

  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO activities (opportunity_id, user_id, channel, result, next_step, follow_up_at) VALUES (%s, %s, %s, %s, %s, %s)",
      (
        opportunity_id,
        user["id"],
        channel,
        result,
        next_step,
        body.get("follow_up_at")
      )
    )
    activity_id = cur.lastrowid
    cur.execute(
      "UPDATE opportunities SET last_follow_up_at = CURRENT_TIMESTAMP WHERE id = %s",
      (opportunity_id,)
    )
    cur.execute("SELECT * FROM activities WHERE id = %s", (activity_id,))
    created = cur.fetchone()

  return jsonify({"data": created}), 201


@app.route("/opportunities/<int:opportunity_id>/activities", methods=["GET"])
@require_user
def list_activities(opportunity_id):
  user = g.user
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    opportunity = cur.fetchone()

  if not opportunity:
    return jsonify({"error": "not_found"}), 404

  if not is_group_admin(user) and opportunity.get("company_id") != user.get("company_id"):
    return jsonify({"error": "not_found"}), 404

  with db.cursor() as cur:
    cur.execute(
      "SELECT * FROM activities WHERE opportunity_id = %s ORDER BY created_at DESC",
      (opportunity_id,)
    )
    rows = cur.fetchall()

  return jsonify({"data": rows})


@app.route("/opportunities/<int:opportunity_id>/contacts", methods=["GET"])
@require_user
def list_contacts(opportunity_id):
  user = g.user
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    opportunity = cur.fetchone()

  if not opportunity:
    return jsonify({"error": "not_found"}), 404

  if not is_group_admin(user) and opportunity.get("company_id") != user.get("company_id"):
    return jsonify({"error": "not_found"}), 404

  with db.cursor() as cur:
    if CONTACT_ROLE_COLUMN_READY is True:
      cur.execute(
        "SELECT id, name, role, title, phone, email, wechat, created_at FROM opportunity_contacts "
        "WHERE opportunity_id = %s ORDER BY id ASC",
        (opportunity_id,)
      )
    else:
      cur.execute(
        "SELECT id, name, title, phone, email, wechat, created_at FROM opportunity_contacts "
        "WHERE opportunity_id = %s ORDER BY id ASC",
        (opportunity_id,)
      )
    rows = cur.fetchall()

  return jsonify({"data": rows})


@app.route("/opportunities/<int:opportunity_id>/tags", methods=["PUT"])
@require_user
def replace_tags(opportunity_id):
  user = g.user
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM opportunities WHERE id = %s", (opportunity_id,))
    opportunity = cur.fetchone()

  if not opportunity:
    return jsonify({"error": "not_found"}), 404

  if not is_group_admin(user) and opportunity.get("company_id") != user.get("company_id"):
    return jsonify({"error": "not_found"}), 404

  body = request.get_json(silent=True) or {}
  tag_ids = body.get("tag_ids")
  if not isinstance(tag_ids, list):
    return jsonify({"error": "tag_ids_required"}), 400

  unique_ids = sorted({int(tag_id) for tag_id in tag_ids if tag_id})

  with db.cursor() as cur:
    cur.execute("DELETE FROM opportunity_tags WHERE opportunity_id = %s", (opportunity_id,))
    if unique_ids:
      cur.executemany(
        "INSERT INTO opportunity_tags (opportunity_id, tag_id) VALUES (%s, %s)",
        [(opportunity_id, tag_id) for tag_id in unique_ids]
      )

  return jsonify({"ok": True})


@app.route("/tags", methods=["GET"])
@require_user
def list_tags():
  filters = []
  params = []

  tag_type = request.args.get("type")
  if tag_type:
    filters.append("type = %s")
    params.append(tag_type)

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

  db = get_db()
  with db.cursor() as cur:
    cur.execute(f"SELECT * FROM tags {where_clause} ORDER BY name ASC", params)
    rows = cur.fetchall()

  return jsonify({"data": rows})


@app.route("/tags", methods=["POST"])
@require_user
def create_tag():
  body = request.get_json(silent=True) or {}
  name = body.get("name")
  tag_type = body.get("type", "custom")

  if not name:
    return jsonify({"error": "missing_name"}), 400

  db = get_db()
  try:
    with db.cursor() as cur:
      cur.execute("INSERT INTO tags (name, type) VALUES (%s, %s)", (name, tag_type))
      tag_id = cur.lastrowid
      cur.execute("SELECT * FROM tags WHERE id = %s", (tag_id,))
      created = cur.fetchone()
    return jsonify({"data": created}), 201
  except IntegrityError:
    with db.cursor() as cur:
      cur.execute("SELECT * FROM tags WHERE name = %s AND type = %s", (name, tag_type))
      existing = cur.fetchone()
    return jsonify({"data": existing}), 200


@app.route("/approval/form-templates", methods=["GET"])
@require_user
def list_approval_form_templates():
  include_schema = request.args.get("include_schema", "1") != "0"
  status = request.args.get("status")
  company_id = request.args.get("company_id", type=int)
  try:
    pagination = _parse_pagination_from_request(default_page_size=20, max_page_size=200)
  except ValueError:
    return jsonify({"error": "invalid_pagination"}), 400

  if status and status not in WORKFLOW_TEMPLATE_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  filters = []
  params = []
  scope_sql, scope_params = _build_template_scope_sql(g.user, "aft")
  if scope_sql:
    filters.append(scope_sql)
    params.extend(scope_params)

  if status:
    filters.append("aft.status = %s")
    params.append(status)

  if company_id is not None:
    if company_id == 0:
      filters.append("aft.company_id IS NULL")
    else:
      filters.append("aft.company_id = %s")
      params.append(company_id)

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

  db = get_db()
  total = None
  if pagination:
    page, page_size, offset = pagination
    with db.cursor() as cur:
      cur.execute(
        "SELECT COUNT(1) AS total "
        "FROM approval_form_templates aft "
        f"{where_clause}",
        tuple(params)
      )
      total = int((cur.fetchone() or {}).get("total") or 0)
      cur.execute(
        "SELECT aft.*, c.name AS company_name, cu.name AS created_by_name, uu.name AS updated_by_name "
        "FROM approval_form_templates aft "
        "LEFT JOIN companies c ON c.id = aft.company_id "
        "LEFT JOIN users cu ON cu.id = aft.created_by "
        "LEFT JOIN users uu ON uu.id = aft.updated_by "
        f"{where_clause} "
        "ORDER BY aft.updated_at DESC, aft.id DESC "
        "LIMIT %s OFFSET %s",
        tuple([*params, page_size, offset])
      )
      rows = cur.fetchall()
  else:
    with db.cursor() as cur:
      cur.execute(
        "SELECT aft.*, c.name AS company_name, cu.name AS created_by_name, uu.name AS updated_by_name "
        "FROM approval_form_templates aft "
        "LEFT JOIN companies c ON c.id = aft.company_id "
        "LEFT JOIN users cu ON cu.id = aft.created_by "
        "LEFT JOIN users uu ON uu.id = aft.updated_by "
        f"{where_clause} "
        "ORDER BY aft.updated_at DESC, aft.id DESC",
        tuple(params)
      )
      rows = cur.fetchall()

  response = {"data": [_serialize_form_template(row, include_schema=include_schema) for row in rows]}
  if pagination:
    response.update({"page": page, "page_size": page_size, "total": total})
  return jsonify(response)


@app.route("/approval/form-templates/<int:template_id>", methods=["GET"])
@require_user
def get_approval_form_template(template_id):
  db = get_db()
  with db.cursor() as cur:
    cur.execute(
      "SELECT aft.*, c.name AS company_name, cu.name AS created_by_name, uu.name AS updated_by_name "
      "FROM approval_form_templates aft "
      "LEFT JOIN companies c ON c.id = aft.company_id "
      "LEFT JOIN users cu ON cu.id = aft.created_by "
      "LEFT JOIN users uu ON uu.id = aft.updated_by "
      "WHERE aft.id = %s",
      (template_id,)
    )
    row = cur.fetchone()

  if not row:
    return jsonify({"error": "not_found"}), 404

  scope_sql, scope_params = _build_template_scope_sql(g.user)
  if scope_sql:
    company_id = row.get("company_id")
    allowed_company_id = scope_params[0] if scope_params else None
    if not (company_id is None or company_id == allowed_company_id):
      return jsonify({"error": "not_found"}), 404

  return jsonify({"data": _serialize_form_template(row, include_schema=True)})


@app.route("/approval/form-templates", methods=["POST"])
@require_user
def create_approval_form_template():
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  name = str(body.get("name") or "").strip()
  description = body.get("description")
  status = body.get("status", "active")
  schema_raw = body.get("schema")
  company_id = _normalize_company_id(body.get("company_id"))

  if not name:
    return jsonify({"error": "missing_name"}), 400
  if status not in WORKFLOW_TEMPLATE_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  db = get_db()
  if is_sub_admin(g.user):
    user_company_id = g.user.get("company_id")
    if not user_company_id:
      return jsonify({"error": "user_missing_company"}), 400
    if company_id and company_id != user_company_id:
      return jsonify({"error": "invalid_company"}), 400
    company_id = user_company_id
  elif company_id and not _ensure_company_exists(db, company_id):
    return jsonify({"error": "invalid_company"}), 400

  try:
    schema = _normalize_workflow_schema(schema_raw)
  except ValueError as err:
    return jsonify({"error": str(err)}), 400

  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO approval_form_templates "
      "(name, description, company_id, schema_json, status, created_by, updated_by) "
      "VALUES (%s, %s, %s, %s, %s, %s, %s)",
      (
        name,
        description,
        company_id,
        _workflow_json_dump(schema),
        status,
        g.user["id"],
        g.user["id"]
      )
    )
    template_id = cur.lastrowid

  with db.cursor() as cur:
    cur.execute(
      "SELECT aft.*, c.name AS company_name, cu.name AS created_by_name, uu.name AS updated_by_name "
      "FROM approval_form_templates aft "
      "LEFT JOIN companies c ON c.id = aft.company_id "
      "LEFT JOIN users cu ON cu.id = aft.created_by "
      "LEFT JOIN users uu ON uu.id = aft.updated_by "
      "WHERE aft.id = %s",
      (template_id,)
    )
    row = cur.fetchone()

  return jsonify({"data": _serialize_form_template(row, include_schema=True)}), 201


@app.route("/approval/form-templates/<int:template_id>", methods=["PATCH"])
@require_user
def update_approval_form_template(template_id):
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  db = get_db()
  with db.cursor() as cur:
    cur.execute("SELECT * FROM approval_form_templates WHERE id = %s", (template_id,))
    existing = cur.fetchone()

  if not existing:
    return jsonify({"error": "not_found"}), 404

  if is_sub_admin(g.user):
    user_company_id = g.user.get("company_id")
    if not user_company_id or existing.get("company_id") != user_company_id:
      return jsonify({"error": "forbidden"}), 403

  updates = []
  params = []

  if "name" in body:
    name = str(body.get("name") or "").strip()
    if not name:
      return jsonify({"error": "missing_name"}), 400
    updates.append("name = %s")
    params.append(name)

  if "description" in body:
    updates.append("description = %s")
    params.append(body.get("description"))

  if "status" in body:
    status = body.get("status")
    if status not in WORKFLOW_TEMPLATE_STATUSES:
      return jsonify({"error": "invalid_status"}), 400
    updates.append("status = %s")
    params.append(status)

  if "company_id" in body:
    if is_sub_admin(g.user):
      company_id = g.user.get("company_id")
    else:
      company_id = _normalize_company_id(body.get("company_id"))
      if company_id and not _ensure_company_exists(db, company_id):
        return jsonify({"error": "invalid_company"}), 400
    updates.append("company_id = %s")
    params.append(company_id)

  if "schema" in body:
    try:
      schema = _normalize_workflow_schema(body.get("schema"))
    except ValueError as err:
      return jsonify({"error": str(err)}), 400
    updates.append("schema_json = %s")
    params.append(_workflow_json_dump(schema))

  if not updates:
    return jsonify({"error": "no_updates"}), 400

  updates.append("updated_by = %s")
  params.append(g.user["id"])
  params.append(template_id)

  with db.cursor() as cur:
    cur.execute(
      f"UPDATE approval_form_templates SET {', '.join(updates)} WHERE id = %s",
      params
    )
    cur.execute(
      "SELECT aft.*, c.name AS company_name, cu.name AS created_by_name, uu.name AS updated_by_name "
      "FROM approval_form_templates aft "
      "LEFT JOIN companies c ON c.id = aft.company_id "
      "LEFT JOIN users cu ON cu.id = aft.created_by "
      "LEFT JOIN users uu ON uu.id = aft.updated_by "
      "WHERE aft.id = %s",
      (template_id,)
    )
    row = cur.fetchone()

  return jsonify({"data": _serialize_form_template(row, include_schema=True)})


@app.route("/approval/process-templates", methods=["GET"])
@require_user
def list_approval_process_templates():
  include_steps = request.args.get("include_steps", "1") != "0"
  include_form_schema = request.args.get("include_form_schema", "0") == "1"
  status = request.args.get("status")
  company_id = request.args.get("company_id", type=int)
  try:
    pagination = _parse_pagination_from_request(default_page_size=20, max_page_size=200)
  except ValueError:
    return jsonify({"error": "invalid_pagination"}), 400

  if status and status not in WORKFLOW_TEMPLATE_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  filters = []
  params = []
  scope_sql, scope_params = _build_template_scope_sql(g.user, "apt")
  if scope_sql:
    filters.append(scope_sql)
    params.extend(scope_params)

  if status:
    filters.append("apt.status = %s")
    params.append(status)
  elif not (is_group_admin(g.user) or is_sub_admin(g.user)):
    filters.append("apt.status = 'active'")

  if company_id is not None:
    if company_id == 0:
      filters.append("apt.company_id IS NULL")
    else:
      filters.append("apt.company_id = %s")
      params.append(company_id)

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

  db = get_db()
  total = None
  if pagination:
    page, page_size, offset = pagination
    with db.cursor() as cur:
      cur.execute(
        "SELECT COUNT(1) AS total "
        "FROM approval_process_templates apt "
        f"{where_clause}",
        tuple(params)
      )
      total = int((cur.fetchone() or {}).get("total") or 0)
      cur.execute(
        "SELECT apt.*, c.name AS company_name, ft.name AS form_template_name, ft.schema_json AS form_schema_json, "
        "cu.name AS created_by_name, uu.name AS updated_by_name "
        "FROM approval_process_templates apt "
        "JOIN approval_form_templates ft ON ft.id = apt.form_template_id "
        "LEFT JOIN companies c ON c.id = apt.company_id "
        "LEFT JOIN users cu ON cu.id = apt.created_by "
        "LEFT JOIN users uu ON uu.id = apt.updated_by "
        f"{where_clause} "
        "ORDER BY apt.updated_at DESC, apt.id DESC "
        "LIMIT %s OFFSET %s",
        tuple([*params, page_size, offset])
      )
      rows = cur.fetchall()
  else:
    with db.cursor() as cur:
      cur.execute(
        "SELECT apt.*, c.name AS company_name, ft.name AS form_template_name, ft.schema_json AS form_schema_json, "
        "cu.name AS created_by_name, uu.name AS updated_by_name "
        "FROM approval_process_templates apt "
        "JOIN approval_form_templates ft ON ft.id = apt.form_template_id "
        "LEFT JOIN companies c ON c.id = apt.company_id "
        "LEFT JOIN users cu ON cu.id = apt.created_by "
        "LEFT JOIN users uu ON uu.id = apt.updated_by "
        f"{where_clause} "
        "ORDER BY apt.updated_at DESC, apt.id DESC",
        tuple(params)
      )
      rows = cur.fetchall()

  response = {
    "data": [
      _serialize_process_template(
        row,
        include_steps=include_steps,
        include_form_schema=include_form_schema
      )
      for row in rows
    ]
  }
  if pagination:
    response.update({"page": page, "page_size": page_size, "total": total})
  return jsonify(response)


@app.route("/approval/process-templates/<int:template_id>", methods=["GET"])
@require_user
def get_approval_process_template(template_id):
  include_form_schema = request.args.get("include_form_schema", "1") != "0"

  db = get_db()
  with db.cursor() as cur:
    cur.execute(
      "SELECT apt.*, c.name AS company_name, ft.name AS form_template_name, ft.schema_json AS form_schema_json, "
      "cu.name AS created_by_name, uu.name AS updated_by_name "
      "FROM approval_process_templates apt "
      "JOIN approval_form_templates ft ON ft.id = apt.form_template_id "
      "LEFT JOIN companies c ON c.id = apt.company_id "
      "LEFT JOIN users cu ON cu.id = apt.created_by "
      "LEFT JOIN users uu ON uu.id = apt.updated_by "
      "WHERE apt.id = %s",
      (template_id,)
    )
    row = cur.fetchone()

  if not row:
    return jsonify({"error": "not_found"}), 404

  scope_sql, scope_params = _build_template_scope_sql(g.user)
  if scope_sql:
    company_scope = scope_params[0] if scope_params else None
    if row.get("company_id") not in (None, company_scope):
      return jsonify({"error": "not_found"}), 404

  if not (is_group_admin(g.user) or is_sub_admin(g.user)) and row.get("status") != "active":
    return jsonify({"error": "not_found"}), 404

  return jsonify(
    {
      "data": _serialize_process_template(
        row,
        include_steps=True,
        include_form_schema=include_form_schema
      )
    }
  )


@app.route("/approval/process-templates/<int:template_id>/versions", methods=["GET"])
@require_user
def list_approval_process_template_versions(template_id):
  include_definition = request.args.get("include_definition", "0") == "1"
  include_form_schema = request.args.get("include_form_schema", "0") == "1"

  db = get_db()
  with db.cursor() as cur:
    cur.execute(
      "SELECT apt.* FROM approval_process_templates apt WHERE apt.id = %s",
      (template_id,)
    )
    template = cur.fetchone()

  if not template:
    return jsonify({"error": "not_found"}), 404

  scope_sql, scope_params = _build_template_scope_sql(g.user)
  if scope_sql:
    company_scope = scope_params[0] if scope_params else None
    if template.get("company_id") not in (None, company_scope):
      return jsonify({"error": "not_found"}), 404

  if not (is_group_admin(g.user) or is_sub_admin(g.user)) and template.get("status") != "active":
    return jsonify({"error": "not_found"}), 404

  with db.cursor() as cur:
    cur.execute(
      "SELECT apv.*, aft.name AS form_template_name, aft.schema_json AS form_schema_json, "
      "cu.name AS created_by_name, uu.name AS updated_by_name "
      "FROM approval_process_template_versions apv "
      "LEFT JOIN approval_form_templates aft ON aft.id = apv.form_template_id "
      "LEFT JOIN users cu ON cu.id = apv.created_by "
      "LEFT JOIN users uu ON uu.id = apv.updated_by "
      "WHERE apv.process_template_id = %s "
      "ORDER BY apv.version_no DESC",
      (template_id,)
    )
    rows = cur.fetchall()

  return jsonify(
    {
      "data": [
        _serialize_process_template_version(
          row,
          include_definition=include_definition,
          include_form_schema=include_form_schema
        )
        for row in rows
      ]
    }
  )


@app.route("/approval/process-templates/validate", methods=["POST"])
@require_user
def validate_approval_process_template():
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  raw_definition = body.get("definition") if "definition" in body else body.get("steps")
  try:
    definition = _normalize_workflow_definition(raw_definition)
  except ValueError as err:
    return jsonify(
      {
        "data": {
          "valid": False,
          "errors": [_workflow_validation_issue(str(err), "流程定义格式不合法。")],
          "warnings": []
        }
      }
    )

  validation = _validate_workflow_definition(definition)
  return jsonify({"data": validation})


@app.route("/approval/conditions/validate-expression", methods=["POST"])
@require_user
def validate_workflow_condition_expression():
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  expression = str(body.get("expression") or "").strip()
  form_data = body.get("form_data") if isinstance(body.get("form_data"), dict) else None

  if not expression:
    return jsonify({"data": {"valid": True, "result": None}})

  if not _is_safe_condition_expression(expression):
    return jsonify(
      {
        "data": {
          "valid": False,
          "result": None,
          "message": "表达式语法不合法，或包含不允许的函数/语句。"
        }
      }
    )

  eval_result = None
  if form_data is not None:
    eval_result = _evaluate_condition_expression(form_data, expression)
  return jsonify({"data": {"valid": True, "result": eval_result}})


@app.route("/approval/process-templates", methods=["POST"])
@require_user
def create_approval_process_template():
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  name = str(body.get("name") or "").strip()
  description = body.get("description")
  status = body.get("status", WORKFLOW_PROCESS_DEFAULT_STATUS)
  raw_form_template_id = body.get("form_template_id")
  form_schema_provided = "form_schema" in body
  company_id = _normalize_company_id(body.get("company_id"))

  if not name:
    return jsonify({"error": "missing_name"}), 400
  if status not in WORKFLOW_TEMPLATE_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  if raw_form_template_id in (None, "", 0, "0"):
    form_template_id = None
  else:
    try:
      form_template_id = int(raw_form_template_id)
    except (TypeError, ValueError):
      return jsonify({"error": "invalid_form_template_id"}), 400
    if form_template_id <= 0:
      return jsonify({"error": "invalid_form_template_id"}), 400

  if form_template_id is None and not form_schema_provided:
    return jsonify({"error": "invalid_form_template_id"}), 400

  form_schema = None
  if form_schema_provided:
    try:
      form_schema = _normalize_workflow_schema(body.get("form_schema"))
    except ValueError as err:
      return jsonify({"error": str(err)}), 400

  db = get_db()
  if is_sub_admin(g.user):
    user_company_id = g.user.get("company_id")
    if not user_company_id:
      return jsonify({"error": "user_missing_company"}), 400
    if company_id and company_id != user_company_id:
      return jsonify({"error": "invalid_company"}), 400
    company_id = user_company_id
  elif company_id and not _ensure_company_exists(db, company_id):
    return jsonify({"error": "invalid_company"}), 400

  raw_definition = body.get("definition") if "definition" in body else body.get("steps")
  try:
    definition = _normalize_workflow_definition(raw_definition)
  except ValueError as err:
    return jsonify({"error": str(err)}), 400

  if status == "active":
    validation = _validate_workflow_definition(definition)
    if not validation["valid"]:
      first_issue = validation["errors"][0]
      return jsonify({"error": first_issue["code"], "details": validation["errors"]}), 400

  effective_form_template_id = form_template_id
  try:
    with _db_transaction(db):
      if effective_form_template_id is None:
        effective_form_template_id = _create_process_owned_form_template(
          db,
          name,
          description,
          company_id,
          form_schema or [],
          g.user["id"]
        )

      with db.cursor() as cur:
        cur.execute(
          "SELECT id, company_id, status FROM approval_form_templates WHERE id = %s FOR UPDATE",
          (effective_form_template_id,)
        )
        form_template = cur.fetchone()

      if not form_template:
        return jsonify({"error": "invalid_form_template"}), 400

      form_company_id = form_template.get("company_id")
      if company_id is None and form_company_id is not None:
        return jsonify({"error": "invalid_form_template_scope"}), 400
      if company_id is not None and form_company_id not in (None, company_id):
        return jsonify({"error": "invalid_form_template_scope"}), 400

      bound_template = _find_process_template_by_form_template(db, effective_form_template_id)
      if bound_template:
        if form_schema_provided:
          # Backward-compatibility: when client submits an already-bound form_template_id together with
          # form schema, allocate a new process-owned form template instead of failing hard.
          effective_form_template_id = _create_process_owned_form_template(
            db,
            name,
            description,
            company_id,
            form_schema or [],
            g.user["id"]
          )
        else:
          return jsonify(
            {
              "error": "form_template_already_bound",
              "details": {
                "process_template_id": bound_template.get("id"),
                "process_template_name": bound_template.get("name")
              }
            }
          ), 400

      if form_schema_provided and form_template_id is not None:
        with db.cursor() as cur:
          cur.execute(
            "UPDATE approval_form_templates "
            "SET schema_json = %s, updated_by = %s "
            "WHERE id = %s",
            (_workflow_json_dump(form_schema), g.user["id"], effective_form_template_id)
          )

      initial_version_no = 1
      published_version_no = initial_version_no if status == "active" else None
      with db.cursor() as cur:
        cur.execute(
          "INSERT INTO approval_process_templates "
          "(name, description, company_id, form_template_id, steps_json, current_version, published_version, status, created_by, updated_by) "
          "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
          (
            name,
            description,
            company_id,
            effective_form_template_id,
            _workflow_json_dump(definition),
            initial_version_no,
            published_version_no,
            status,
            g.user["id"],
            g.user["id"]
          )
        )
        template_id = cur.lastrowid

      _create_process_template_version(
        db,
        template_id,
        initial_version_no,
        effective_form_template_id,
        definition,
        "published" if status == "active" else "draft",
        g.user["id"],
        published_at=datetime.now() if status == "active" else None
      )
      if status == "active":
        _publish_process_template_version(db, template_id, initial_version_no, g.user["id"])

      with db.cursor() as cur:
        cur.execute(
          "SELECT apt.*, c.name AS company_name, ft.name AS form_template_name, ft.schema_json AS form_schema_json, "
          "cu.name AS created_by_name, uu.name AS updated_by_name "
          "FROM approval_process_templates apt "
          "JOIN approval_form_templates ft ON ft.id = apt.form_template_id "
          "LEFT JOIN companies c ON c.id = apt.company_id "
          "LEFT JOIN users cu ON cu.id = apt.created_by "
          "LEFT JOIN users uu ON uu.id = apt.updated_by "
          "WHERE apt.id = %s",
          (template_id,)
        )
        row = cur.fetchone()

    return jsonify({"data": _serialize_process_template(row, include_steps=True, include_form_schema=True)}), 201
  except IntegrityError as err:
    if "uniq_proc_tpl_form_id" in str(err).lower():
      lookup_form_template_id = effective_form_template_id or form_template_id
      bound_template = None
      if lookup_form_template_id:
        bound_template = _find_process_template_by_form_template(db, lookup_form_template_id)
      return jsonify(
        {
          "error": "form_template_already_bound",
          "details": {
            "process_template_id": (bound_template or {}).get("id"),
            "process_template_name": (bound_template or {}).get("name")
          }
        }
      ), 400
    raise


@app.route("/approval/process-templates/<int:template_id>", methods=["PATCH"])
@require_user
def update_approval_process_template(template_id):
  guard = ensure_org_access()
  if guard:
    return guard

  body = request.get_json(silent=True) or {}
  form_schema_provided = "form_schema" in body
  form_template_id_provided = "form_template_id" in body
  form_schema = None
  if form_schema_provided:
    try:
      form_schema = _normalize_workflow_schema(body.get("form_schema"))
    except ValueError as err:
      return jsonify({"error": str(err)}), 400

  db = get_db()
  try:
    with _db_transaction(db):
      with db.cursor() as cur:
        cur.execute("SELECT * FROM approval_process_templates WHERE id = %s FOR UPDATE", (template_id,))
        existing = cur.fetchone()

      if not existing:
        return jsonify({"error": "not_found"}), 404

      if is_sub_admin(g.user):
        user_company_id = g.user.get("company_id")
        if not user_company_id or existing.get("company_id") != user_company_id:
          return jsonify({"error": "forbidden"}), 403

      updates = []
      params = []
      next_status = existing.get("status")
      next_company_id = existing.get("company_id")
      next_form_template_id = existing.get("form_template_id")
      existing_form_template_id = existing.get("form_template_id")
      existing_current_version = int(existing.get("current_version") or 1)
      existing_published_version = existing.get("published_version")
      try:
        next_definition = _normalize_workflow_definition(_safe_json_load(existing.get("steps_json")))
      except ValueError:
        next_definition = _steps_to_graph_definition([])
      existing_definition_dump = _workflow_json_dump(next_definition)
      form_template_changed = False
      definition_changed = False

      if "name" in body:
        name = str(body.get("name") or "").strip()
        if not name:
          return jsonify({"error": "missing_name"}), 400
        updates.append("name = %s")
        params.append(name)

      if "description" in body:
        updates.append("description = %s")
        params.append(body.get("description"))

      if "status" in body:
        status = body.get("status")
        if status not in WORKFLOW_TEMPLATE_STATUSES:
          return jsonify({"error": "invalid_status"}), 400
        next_status = status
        updates.append("status = %s")
        params.append(status)

      if "company_id" in body:
        if is_sub_admin(g.user):
          next_company_id = g.user.get("company_id")
        else:
          next_company_id = _normalize_company_id(body.get("company_id"))
          if next_company_id and not _ensure_company_exists(db, next_company_id):
            return jsonify({"error": "invalid_company"}), 400
        updates.append("company_id = %s")
        params.append(next_company_id)

      if form_template_id_provided:
        try:
          next_form_template_id = int(body.get("form_template_id"))
        except (TypeError, ValueError):
          return jsonify({"error": "invalid_form_template_id"}), 400
        if next_form_template_id <= 0:
          return jsonify({"error": "invalid_form_template_id"}), 400
        form_template_changed = next_form_template_id != existing_form_template_id

      if "steps" in body or "definition" in body:
        raw_definition = body.get("definition") if "definition" in body else body.get("steps")
        try:
          definition = _normalize_workflow_definition(raw_definition)
        except ValueError as err:
          return jsonify({"error": str(err)}), 400
        next_definition = definition
        definition_changed = _workflow_json_dump(next_definition) != existing_definition_dump

      with db.cursor() as cur:
        cur.execute(
          "SELECT id, company_id, schema_json FROM approval_form_templates WHERE id = %s FOR UPDATE",
          (next_form_template_id,)
        )
        form_template = cur.fetchone()

      if not form_template:
        return jsonify({"error": "invalid_form_template"}), 400

      form_company_id = form_template.get("company_id")
      if next_company_id is None and form_company_id is not None:
        return jsonify({"error": "invalid_form_template_scope"}), 400
      if next_company_id is not None and form_company_id not in (None, next_company_id):
        return jsonify({"error": "invalid_form_template_scope"}), 400

      bound_template = _find_process_template_by_form_template(
        db,
        next_form_template_id,
        exclude_template_id=template_id
      )
      if bound_template:
        stale_form_binding = (
          form_template_id_provided and
          existing_form_template_id and
          next_form_template_id != existing_form_template_id
        )
        if stale_form_binding:
          # Backward-compatibility: older clients may submit stale form_template_id from cached drafts.
          # Keep current process->form binding and continue applying other updates.
          next_form_template_id = existing_form_template_id
          form_template_changed = False
          with db.cursor() as cur:
            cur.execute(
              "SELECT id, company_id, schema_json FROM approval_form_templates WHERE id = %s FOR UPDATE",
              (next_form_template_id,)
            )
            form_template = cur.fetchone()
          if not form_template:
            return jsonify({"error": "invalid_form_template"}), 400

          form_company_id = form_template.get("company_id")
          if next_company_id is None and form_company_id is not None:
            return jsonify({"error": "invalid_form_template_scope"}), 400
          if next_company_id is not None and form_company_id not in (None, next_company_id):
            return jsonify({"error": "invalid_form_template_scope"}), 400
        else:
          return jsonify(
            {
              "error": "form_template_already_bound",
              "details": {
                "process_template_id": bound_template.get("id"),
                "process_template_name": bound_template.get("name")
              }
            }
          ), 400

      form_schema_changed = False
      if form_schema_provided:
        existing_schema = _safe_json_load(form_template.get("schema_json"))
        if not isinstance(existing_schema, list):
          existing_schema = []
        form_schema_changed = _workflow_json_dump(existing_schema) != _workflow_json_dump(form_schema)

      version_changed = form_template_changed or definition_changed or form_schema_changed
      publish_requested = "status" in body and next_status == "active"
      if not updates and not version_changed and not publish_requested:
        return jsonify({"error": "no_updates"}), 400

      if publish_requested:
        validation = _validate_workflow_definition(next_definition)
        if not validation["valid"]:
          first_issue = validation["errors"][0]
          return jsonify({"error": first_issue["code"], "details": validation["errors"]}), 400

      if form_schema_changed:
        with db.cursor() as cur:
          cur.execute(
            "UPDATE approval_form_templates "
            "SET schema_json = %s, updated_by = %s "
            "WHERE id = %s",
            (_workflow_json_dump(form_schema), g.user["id"], next_form_template_id)
          )

      next_current_version = existing_current_version
      next_published_version = existing_published_version
      if version_changed:
        next_current_version = existing_current_version + 1
        _create_process_template_version(
          db,
          template_id,
          next_current_version,
          next_form_template_id,
          next_definition,
          "published" if publish_requested else "draft",
          g.user["id"],
          published_at=datetime.now() if publish_requested else None
        )
        updates.append("form_template_id = %s")
        params.append(next_form_template_id)
        updates.append("steps_json = %s")
        params.append(_workflow_json_dump(next_definition))
        updates.append("current_version = %s")
        params.append(next_current_version)
        if publish_requested:
          next_published_version = next_current_version
      elif publish_requested:
        next_published_version = existing_current_version

      if publish_requested:
        updates.append("published_version = %s")
        params.append(next_published_version)

      updates.append("updated_by = %s")
      params.append(g.user["id"])
      params.append(template_id)

      with db.cursor() as cur:
        cur.execute(
          f"UPDATE approval_process_templates SET {', '.join(updates)} WHERE id = %s",
          params
        )

      if publish_requested:
        _publish_process_template_version(db, template_id, next_published_version, g.user["id"])

      with db.cursor() as cur:
        cur.execute(
          "SELECT apt.*, c.name AS company_name, ft.name AS form_template_name, ft.schema_json AS form_schema_json, "
          "cu.name AS created_by_name, uu.name AS updated_by_name "
          "FROM approval_process_templates apt "
          "JOIN approval_form_templates ft ON ft.id = apt.form_template_id "
          "LEFT JOIN companies c ON c.id = apt.company_id "
          "LEFT JOIN users cu ON cu.id = apt.created_by "
          "LEFT JOIN users uu ON uu.id = apt.updated_by "
          "WHERE apt.id = %s",
          (template_id,)
        )
        row = cur.fetchone()

    return jsonify({"data": _serialize_process_template(row, include_steps=True, include_form_schema=True)})
  except IntegrityError as err:
    if "uniq_proc_tpl_form_id" in str(err).lower():
      lookup_form_template_id = None
      try:
        lookup_form_template_id = int(body.get("form_template_id") or 0)
      except (TypeError, ValueError):
        lookup_form_template_id = None
      bound_template = (
        _find_process_template_by_form_template(
          db,
          lookup_form_template_id,
          exclude_template_id=template_id
        )
        if lookup_form_template_id and lookup_form_template_id > 0
        else None
      )
      return jsonify(
        {
          "error": "form_template_already_bound",
          "details": {
            "process_template_id": (bound_template or {}).get("id"),
            "process_template_name": (bound_template or {}).get("name")
          }
        }
      ), 400
    raise


@app.route("/approval/instances", methods=["GET"])
@require_user
def list_approval_instances():
  scope = request.args.get("scope", "all")
  status = request.args.get("status")
  try:
    pagination = _parse_pagination_from_request(default_page_size=20, max_page_size=200)
  except ValueError:
    return jsonify({"error": "invalid_pagination"}), 400

  if scope not in {"all", "mine", "pending"}:
    return jsonify({"error": "invalid_scope"}), 400
  if status and status not in WORKFLOW_INSTANCE_STATUSES:
    return jsonify({"error": "invalid_status"}), 400

  filters = []
  filter_params = []

  if is_group_admin(g.user):
    if scope == "mine":
      filters.append("ai.applicant_id = %s")
      filter_params.append(g.user["id"])
    elif scope == "pending":
      filters.append(
        "EXISTS(SELECT 1 FROM approval_instance_tasks apt2 "
        "WHERE apt2.instance_id = ai.id AND apt2.approver_id = %s AND apt2.status = 'pending')"
      )
      filter_params.append(g.user["id"])
  else:
    if scope == "mine":
      filters.append("ai.applicant_id = %s")
      filter_params.append(g.user["id"])
    elif scope == "pending":
      filters.append(
        "EXISTS(SELECT 1 FROM approval_instance_tasks apt2 "
        "WHERE apt2.instance_id = ai.id AND apt2.approver_id = %s AND apt2.status = 'pending')"
      )
      filter_params.append(g.user["id"])
    else:
      filters.append(
        "(ai.applicant_id = %s OR EXISTS("
        "SELECT 1 FROM approval_instance_tasks apt2 "
        "WHERE apt2.instance_id = ai.id AND apt2.approver_id = %s"
        "))"
      )
      filter_params.extend([g.user["id"], g.user["id"]])

  if status:
    filters.append("ai.status = %s")
    filter_params.append(status)

  where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

  db = get_db()
  total = None
  if pagination:
    page, page_size, offset = pagination
    with db.cursor() as cur:
      cur.execute(
        "SELECT COUNT(1) AS total "
        "FROM approval_instances ai "
        f"{where_clause}",
        tuple(filter_params)
      )
      total = int((cur.fetchone() or {}).get("total") or 0)
      cur.execute(
        "SELECT ai.*, c.name AS company_name, au.name AS applicant_name, "
        "EXISTS(SELECT 1 FROM approval_instance_tasks ait "
        "WHERE ait.instance_id = ai.id AND ait.approver_id = %s AND ait.status = 'pending') AS pending_action "
        "FROM approval_instances ai "
        "LEFT JOIN companies c ON c.id = ai.company_id "
        "LEFT JOIN users au ON au.id = ai.applicant_id "
        f"{where_clause} "
        "ORDER BY ai.created_at DESC, ai.id DESC "
        "LIMIT %s OFFSET %s",
        tuple([g.user["id"], *filter_params, page_size, offset])
      )
      rows = cur.fetchall()
  else:
    with db.cursor() as cur:
      cur.execute(
        "SELECT ai.*, c.name AS company_name, au.name AS applicant_name, "
        "EXISTS(SELECT 1 FROM approval_instance_tasks ait "
        "WHERE ait.instance_id = ai.id AND ait.approver_id = %s AND ait.status = 'pending') AS pending_action "
        "FROM approval_instances ai "
        "LEFT JOIN companies c ON c.id = ai.company_id "
        "LEFT JOIN users au ON au.id = ai.applicant_id "
        f"{where_clause} "
        "ORDER BY ai.created_at DESC, ai.id DESC "
        "LIMIT 500",
        tuple([g.user["id"], *filter_params])
      )
      rows = cur.fetchall()

  response = {"data": [_serialize_approval_instance(row, include_payload=False) for row in rows]}
  if pagination:
    response.update({"page": page, "page_size": page_size, "total": total})
  return jsonify(response)


@app.route("/approval/instances", methods=["POST"])
@require_user
def create_approval_instance():
  body = request.get_json(silent=True) or {}
  process_template_id = body.get("process_template_id")
  title = str(body.get("title") or "").strip()
  form_data_raw = body.get("form_data")

  try:
    process_template_id = int(process_template_id)
  except (TypeError, ValueError):
    return jsonify({"error": "invalid_process_template_id"}), 400

  db = get_db()
  with db.cursor() as cur:
    cur.execute(
      "SELECT apt.* "
      "FROM approval_process_templates apt "
      "WHERE apt.id = %s",
      (process_template_id,)
    )
    process_template = cur.fetchone()

  if not process_template:
    return jsonify({"error": "invalid_process_template"}), 400
  if process_template.get("status") != "active":
    return jsonify({"error": "process_template_inactive"}), 400
  published_version = process_template.get("published_version")
  try:
    published_version = int(published_version)
  except (TypeError, ValueError):
    published_version = 0
  if published_version <= 0:
    return jsonify({"error": "process_template_not_published_version"}), 400

  template_company_id = process_template.get("company_id")
  if not is_group_admin(g.user):
    user_company_id = g.user.get("company_id")
    if template_company_id not in (None, user_company_id):
      return jsonify({"error": "forbidden"}), 403

  version_row = _get_process_template_version(db, process_template_id, published_version)
  if not version_row:
    return jsonify({"error": "process_template_not_published_version"}), 400

  with db.cursor() as cur:
    cur.execute(
      "SELECT id, company_id, schema_json FROM approval_form_templates WHERE id = %s",
      (version_row.get("form_template_id"),)
    )
    form_template = cur.fetchone()

  if not form_template:
    return jsonify({"error": "invalid_form_template"}), 400

  schema = _safe_json_load(form_template.get("schema_json"))
  if not isinstance(schema, list):
    return jsonify({"error": "invalid_form_schema"}), 400

  try:
    normalized_form_data = _validate_workflow_form_data(schema, form_data_raw)
  except ValueError as err:
    return jsonify({"error": str(err)}), 400

  raw_definition = _safe_json_load(version_row.get("definition_json"))
  try:
    definition = _normalize_workflow_definition(raw_definition)
  except ValueError:
    return jsonify({"error": "invalid_process_steps"}), 400

  if not definition.get("nodes"):
    return jsonify({"error": "invalid_process_steps"}), 400
  validation = _validate_workflow_definition(definition)
  if not validation["valid"]:
    first_issue = validation["errors"][0]
    return jsonify({"error": first_issue["code"], "details": validation["errors"]}), 400

  instance_company_id = template_company_id or g.user.get("company_id")

  if not title:
    title = f"{process_template.get('name')} - {datetime.now().strftime('%Y-%m-%d %H:%M')}"

  process_snapshot = {
    "id": process_template.get("id"),
    "name": process_template.get("name"),
    "version": published_version,
    "definition": definition
  }

  with db.cursor() as cur:
    cur.execute(
      "INSERT INTO approval_instances "
      "(process_template_id, form_template_id, process_name, title, company_id, applicant_id, "
      "process_snapshot_json, form_schema_json, form_data_json, status, current_step, total_steps, current_step_name, current_node_id) "
      "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending', 0, %s, NULL, %s)",
      (
        process_template.get("id"),
        version_row.get("form_template_id"),
        process_template.get("name"),
        title,
        instance_company_id,
        g.user.get("id"),
        _workflow_json_dump(process_snapshot),
        _workflow_json_dump(schema),
        _workflow_json_dump(normalized_form_data),
        len(_extract_steps_from_definition(definition)),
        definition.get("start_node_id")
      )
    )
    instance_id = cur.lastrowid

  with db.cursor() as cur:
    cur.execute("SELECT * FROM approval_instances WHERE id = %s", (instance_id,))
    instance_row = cur.fetchone()

  _route_instance_forward(db, instance_row, definition.get("start_node_id"))

  data = _get_instance_detail(db, instance_id, g.user)
  return jsonify({"data": data}), 201


@app.route("/approval/instances/<int:instance_id>", methods=["GET"])
@require_user
def get_approval_instance(instance_id):
  db = get_db()
  data = _get_instance_detail(db, instance_id, g.user)
  if not data:
    return jsonify({"error": "not_found"}), 404
  return jsonify({"data": data})


@app.route("/approval/instances/<int:instance_id>/actions", methods=["POST"])
@require_user
def handle_approval_instance_action(instance_id):
  body = request.get_json(silent=True) or {}
  action = str(body.get("action") or "").strip().lower()
  comment = body.get("comment")
  target_user_id_raw = body.get("target_user_id")
  target_user_ids_raw = body.get("target_user_ids")
  form_data_patch = body.get("form_data")
  idempotency_key = _normalize_idempotency_key(request.headers.get("Idempotency-Key"))

  if action not in WORKFLOW_INSTANCE_ACTIONS:
    return jsonify({"error": "invalid_action"}), 400

  db = get_db()
  def respond_success(payload, status_code=200):
    response_body = {"data": payload}
    if idempotency_key:
      _save_action_idempotency_response(
        db,
        idempotency_key,
        instance_id,
        g.user.get("id"),
        action,
        response_body,
        status_code
      )
    return jsonify(response_body), status_code

  with _db_transaction(db):
    with db.cursor() as cur:
      cur.execute(
        "SELECT ai.*, "
        "EXISTS(SELECT 1 FROM approval_instance_tasks ait WHERE ait.instance_id = ai.id AND ait.approver_id = %s) AS has_task_access "
        "FROM approval_instances ai "
        "WHERE ai.id = %s "
        "FOR UPDATE",
        (g.user["id"], instance_id)
      )
      instance = cur.fetchone()

    if not instance:
      return jsonify({"error": "not_found"}), 404
    if not _can_access_instance(g.user, instance):
      return jsonify({"error": "not_found"}), 404

    if idempotency_key:
      cached_body, cached_status = _load_action_idempotency_response(
        db,
        idempotency_key,
        instance_id,
        g.user.get("id"),
        action
      )
      if cached_body is not None:
        return jsonify(cached_body), cached_status

    if instance.get("status") != "pending":
      return jsonify({"error": "invalid_instance_status"}), 400

    current_definition = _load_instance_definition(instance)
    current_node = _get_instance_current_node(instance, definition=current_definition)
    current_approval_type = str((current_node or {}).get("approval_type") or "").strip().lower()
    current_field_permission_map = _build_field_permission_map((current_node or {}).get("field_permissions"))

    if action == "withdraw":
      if instance.get("applicant_id") != g.user.get("id"):
        return jsonify({"error": "forbidden"}), 403
      with db.cursor() as cur:
        cur.execute(
          "UPDATE approval_instances SET status = 'withdrawn', finished_at = CURRENT_TIMESTAMP WHERE id = %s",
          (instance_id,)
        )
        cur.execute(
          "UPDATE approval_instance_tasks SET status = 'skipped' "
          "WHERE instance_id = %s AND status IN ('pending', 'waiting')",
          (instance_id,)
        )
      _log_instance_event(
        db,
        instance_id,
        g.user.get("id"),
        "withdraw",
        comment=comment
      )
      data = _get_instance_detail(db, instance_id, g.user)
      return respond_success(data)

    if action == "remind":
      with db.cursor() as cur:
        cur.execute(
          "SELECT approver_id FROM approval_instance_tasks "
          "WHERE instance_id = %s AND step_no = %s AND status = 'pending' "
          "ORDER BY id ASC",
          (instance_id, instance.get("current_step"))
        )
        pending_rows = cur.fetchall()
      if not pending_rows:
        return jsonify({"error": "no_pending_task"}), 400
      reminded_user_ids = sorted({row.get("approver_id") for row in pending_rows if row.get("approver_id")})
      _log_instance_event(
        db,
        instance_id,
        g.user.get("id"),
        "remind",
        comment=comment,
        detail={"reminded_user_ids": reminded_user_ids}
      )
      data = _get_instance_detail(db, instance_id, g.user)
      return respond_success(data)

    with db.cursor() as cur:
      cur.execute(
        "SELECT * FROM approval_instance_tasks "
        "WHERE instance_id = %s AND step_no = %s AND approver_id = %s AND status = 'pending' "
        "ORDER BY id ASC LIMIT 1 "
        "FOR UPDATE",
        (instance_id, instance.get("current_step"), g.user.get("id"))
      )
      task = cur.fetchone()

    if not task:
      return jsonify({"error": "no_pending_task"}), 400

    updated_form_fields = []
    if action == "approve" and form_data_patch not in (None, ""):
      if not isinstance(form_data_patch, dict):
        return jsonify({"error": "invalid_form_data"}), 400
      editable_keys = {
        field_key
        for field_key, permission in current_field_permission_map.items()
        if permission.get("can_edit")
      }
      if not editable_keys:
        return jsonify({"error": "field_update_not_allowed"}), 400
      invalid_keys = [key for key in form_data_patch.keys() if key not in editable_keys]
      if invalid_keys:
        return jsonify({"error": "field_update_forbidden", "details": {"fields": invalid_keys}}), 400

      schema = _safe_json_load(instance.get("form_schema_json"))
      if not isinstance(schema, list):
        return jsonify({"error": "invalid_form_schema"}), 400
      current_form_data = _load_instance_form_data(instance)
      merged_form_data = dict(current_form_data)
      for key, value in form_data_patch.items():
        merged_form_data[key] = value
        updated_form_fields.append(key)
      try:
        normalized_form_data = _validate_workflow_form_data(schema, merged_form_data)
      except ValueError as err:
        return jsonify({"error": str(err)}), 400
      for field_key, permission in current_field_permission_map.items():
        if permission.get("required") and permission.get("can_edit"):
          if _is_empty_value(normalized_form_data.get(field_key)):
            return jsonify({"error": f"missing_required_field:{field_key}"}), 400
      with db.cursor() as cur:
        cur.execute(
          "UPDATE approval_instances SET form_data_json = %s WHERE id = %s",
          (_workflow_json_dump(normalized_form_data), instance_id)
        )
      instance["form_data_json"] = _workflow_json_dump(normalized_form_data)

    if action == "transfer":
      try:
        target_user_id = int(target_user_id_raw)
      except (TypeError, ValueError):
        return jsonify({"error": "invalid_target_user"}), 400
      if target_user_id <= 0 or target_user_id == g.user.get("id"):
        return jsonify({"error": "invalid_target_user"}), 400
      with db.cursor() as cur:
        cur.execute("SELECT id FROM users WHERE id = %s AND status = 'active'", (target_user_id,))
        target_user = cur.fetchone()
      if not target_user:
        return jsonify({"error": "invalid_target_user"}), 400
      with db.cursor() as cur:
        cur.execute(
          "SELECT id FROM approval_instance_tasks "
          "WHERE instance_id = %s AND step_no = %s AND approver_id = %s "
          "LIMIT 1",
          (instance_id, instance.get("current_step"), target_user_id)
        )
        target_exists = cur.fetchone()
        if target_exists:
          return jsonify({"error": "target_user_task_exists"}), 400
        cur.execute(
          "UPDATE approval_instance_tasks "
          "SET status = 'skipped', comment = %s, acted_at = CURRENT_TIMESTAMP "
          "WHERE id = %s",
          (f"transfer_to:{target_user_id}" + (f"; {comment}" if comment else ""), task.get("id"))
        )
        cur.execute(
          "INSERT INTO approval_instance_tasks "
          "(instance_id, step_no, step_name, approval_mode, approver_id, status, comment) "
          "VALUES (%s, %s, %s, %s, %s, 'pending', %s)",
          (
            instance_id,
            instance.get("current_step"),
            task.get("step_name"),
            task.get("approval_mode") or "any",
            target_user_id,
            "transferred_from_current_approver"
          )
        )
        new_task_id = cur.lastrowid
      _log_instance_event(
        db,
        instance_id,
        g.user.get("id"),
        "transfer",
        task_id=new_task_id,
        comment=comment,
        detail={"from_user_id": g.user.get("id"), "to_user_id": target_user_id}
      )
      data = _get_instance_detail(db, instance_id, g.user)
      return respond_success(data)

    if action == "add_sign":
      raw_targets = []
      if isinstance(target_user_ids_raw, list):
        raw_targets = target_user_ids_raw
      elif target_user_id_raw not in (None, ""):
        raw_targets = [target_user_id_raw]
      target_user_ids = []
      for raw_user_id in raw_targets:
        try:
          user_id = int(raw_user_id)
        except (TypeError, ValueError):
          continue
        if user_id > 0 and user_id not in target_user_ids:
          target_user_ids.append(user_id)
      if not target_user_ids:
        return jsonify({"error": "invalid_target_user"}), 400

      placeholders = ", ".join(["%s"] * len(target_user_ids))
      with db.cursor() as cur:
        cur.execute(
          f"SELECT id FROM users WHERE status = 'active' AND id IN ({placeholders})",
          tuple(target_user_ids)
        )
        valid_target_ids = [row["id"] for row in cur.fetchall()]
      if not valid_target_ids:
        return jsonify({"error": "invalid_target_user"}), 400

      with db.cursor() as cur:
        cur.execute(
          "SELECT approver_id FROM approval_instance_tasks "
          "WHERE instance_id = %s AND step_no = %s",
          (instance_id, instance.get("current_step"))
        )
        existing_approver_ids = {row.get("approver_id") for row in cur.fetchall() if row.get("approver_id")}
        insert_rows = [
          (
            instance_id,
            instance.get("current_step"),
            task.get("step_name"),
            task.get("approval_mode") or "any",
            approver_id,
            "waiting" if current_approval_type == "sequential" else "pending",
            "add_sign_added"
          )
          for approver_id in valid_target_ids
          if approver_id not in existing_approver_ids
        ]
        if not insert_rows:
          return jsonify({"error": "target_user_task_exists"}), 400
        cur.executemany(
          "INSERT INTO approval_instance_tasks "
          "(instance_id, step_no, step_name, approval_mode, approver_id, status, comment) "
          "VALUES (%s, %s, %s, %s, %s, %s, %s)",
          insert_rows
        )
      _log_instance_event(
        db,
        instance_id,
        g.user.get("id"),
        "add_sign",
        task_id=task.get("id"),
        comment=comment,
        detail={"added_user_ids": [row[4] for row in insert_rows]}
      )
      data = _get_instance_detail(db, instance_id, g.user)
      return respond_success(data)

    if action in {"reject", "return"}:
      action_comment = comment
      if action == "return" and not action_comment:
        action_comment = "returned_by_approver"
      with db.cursor() as cur:
        cur.execute(
          "UPDATE approval_instance_tasks "
          "SET status = 'rejected', decision = 'reject', comment = %s, acted_at = CURRENT_TIMESTAMP "
          "WHERE id = %s",
          (action_comment, task.get("id"))
        )
        cur.execute(
          "UPDATE approval_instance_tasks SET status = 'skipped' "
          "WHERE instance_id = %s AND step_no = %s AND status IN ('pending', 'waiting')",
          (instance_id, instance.get("current_step"))
        )
      _mark_instance_finished(db, instance_id, "rejected")
      _log_instance_event(
        db,
        instance_id,
        g.user.get("id"),
        action,
        task_id=task.get("id"),
        comment=action_comment
      )
      data = _get_instance_detail(db, instance_id, g.user)
      return respond_success(data)

    with db.cursor() as cur:
      cur.execute(
        "UPDATE approval_instance_tasks "
        "SET status = 'approved', decision = 'approve', comment = %s, acted_at = CURRENT_TIMESTAMP "
        "WHERE id = %s",
        (comment, task.get("id"))
      )
    _log_instance_event(
      db,
      instance_id,
      g.user.get("id"),
      "approve",
      task_id=task.get("id"),
      comment=comment,
      detail={"updated_fields": sorted(set(updated_form_fields))} if updated_form_fields else None
    )

    with db.cursor() as cur:
      cur.execute("SELECT * FROM approval_instances WHERE id = %s", (instance_id,))
      latest_instance = cur.fetchone()
    _advance_approval_instance(db, latest_instance)

    data = _get_instance_detail(db, instance_id, g.user)
    return respond_success(data)


if __name__ == "__main__":
  port = int(os.getenv("PORT", "3000"))
  app.run(host="0.0.0.0", port=port)
